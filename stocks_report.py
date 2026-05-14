#!/usr/bin/env python3
"""
Daily multi-index stock report — S&P 500, Nikkei 225, FTSE 100, DAX 40, CAC 40, BEL 20.

Outputs a single .xlsx with one row per stock containing:
  - Current price + prices 1d / 1w / 1mo / 6mo / 1y / 5y ago, all in EUR at TODAY's FX rate
  - Trailing P/E and Forward P/E
  - Native currency
  - Comma-separated list of Yahoo watchlists the stock appears in, from:
      Largest 52-Week Gains, Crowded Hedge Fund Positions, Berkshire Hathaway Portfolio,
      Smart Money Stocks, Most Sold by Activist Hedge Funds,
      Most Bought by Activist Hedge Funds, Most Bought by Hedge Funds,
      Activist Hedge Fund Positions

Sources: yfinance (Yahoo Finance), Wikipedia for index constituents, Yahoo watchlist pages.

USAGE
-----
First-time setup:
    pip install -r requirements.txt

Smoke test (BEL 20 only, ~20 tickers, finishes in ~1 min):
    python stocks_report.py --indexes BEL20

Full run (all indexes, ~925 tickers, 10–20 min):
    python stocks_report.py

Custom output path:
    python stocks_report.py --output ~/reports/stocks.xlsx

CAVEATS
-------
- yfinance is unofficial; Yahoo can break it. If a previously-working run fails,
  try: pip install --upgrade yfinance
- Watchlist scraping depends on Yahoo's HTML structure and may break silently.
  The report still completes; failed lists log a warning.
- Tickers younger than 5y will have a blank "5Y ago" cell — expected.
- FX conversion uses TODAY's rate for all historical prices, as requested.
  This mixes price movement with FX movement for non-EUR stocks.
"""

from __future__ import annotations

import argparse
import datetime as dt
import io
import logging
import re
import sqlite3
import sys
import threading
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Callable, Optional

import pandas as pd
import requests
import yfinance as yf
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("stocks_report")

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
HTTP_HEADERS = {"User-Agent": USER_AGENT, "Accept-Language": "en-US,en;q=0.9"}

# Wikipedia URLs per index. Constituent table is auto-detected by column names.
INDEX_WIKI = {
    "SP500":     "https://en.wikipedia.org/wiki/List_of_S%26P_500_companies",
    "NIKKEI225": "https://en.wikipedia.org/wiki/Nikkei_225",
    "FTSE100":   "https://en.wikipedia.org/wiki/FTSE_100_Index",
    "DAX":       "https://en.wikipedia.org/wiki/DAX",
    "CAC40":     "https://en.wikipedia.org/wiki/CAC_40",
    "BEL20":     "https://en.wikipedia.org/wiki/BEL_20",
    "ESTOXX50":  "https://en.wikipedia.org/wiki/EURO_STOXX_50",
}

# Native currency fallback if yfinance .info doesn't give us one.
INDEX_DEFAULT_CCY = {
    "SP500": "USD", "NIKKEI225": "JPY", "FTSE100": "GBP",
    "DAX": "EUR", "CAC40": "EUR", "BEL20": "EUR", "ESTOXX50": "EUR",
}

# ETFs to surface in Market alongside the index constituents. Curated;
# Yahoo tickers are stable so this list rarely needs touching.
# Schema: (Yahoo ticker, display name, label-for-the-Indexes-column).
ETF_LIST: list[tuple[str, str, str]] = [
    # One per tracked index — gives a single-line proxy for each universe.
    ("SPY",   "SPDR S&P 500 ETF Trust",                          "ETF — SP500"),
    ("EWJ",   "iShares MSCI Japan ETF",                          "ETF — NIKKEI225"),
    ("EWU",   "iShares MSCI United Kingdom ETF",                 "ETF — FTSE100"),
    ("EWG",   "iShares MSCI Germany ETF",                        "ETF — DAX"),
    ("EWQ",   "iShares MSCI France ETF",                         "ETF — CAC40"),
    ("EWK",   "iShares MSCI Belgium ETF",                        "ETF — BEL20"),
    ("FEZ",   "SPDR EURO STOXX 50 ETF",                          "ETF — ESTOXX50"),
    # One per GICS sector (US Select Sector SPDR funds).
    ("XLE",   "Energy Select Sector SPDR Fund",                  "ETF — Sector: Energy"),
    ("XLB",   "Materials Select Sector SPDR Fund",               "ETF — Sector: Materials"),
    ("XLI",   "Industrial Select Sector SPDR Fund",              "ETF — Sector: Industrials"),
    ("XLY",   "Consumer Discretionary Select Sector SPDR Fund",  "ETF — Sector: Consumer Discretionary"),
    ("XLP",   "Consumer Staples Select Sector SPDR Fund",        "ETF — Sector: Consumer Staples"),
    ("XLV",   "Health Care Select Sector SPDR Fund",             "ETF — Sector: Healthcare"),
    ("XLF",   "Financial Select Sector SPDR Fund",               "ETF — Sector: Financials"),
    ("XLK",   "Technology Select Sector SPDR Fund",              "ETF — Sector: Technology"),
    ("XLC",   "Communication Services Select Sector SPDR Fund",  "ETF — Sector: Communication Services"),
    ("XLU",   "Utilities Select Sector SPDR Fund",               "ETF — Sector: Utilities"),
    ("XLRE",  "Real Estate Select Sector SPDR Fund",             "ETF — Sector: Real Estate"),
    # One per semi-large country (iShares MSCI single-country funds).
    ("EWN",   "iShares MSCI Netherlands ETF",                    "ETF — Country: Netherlands"),
    ("EWI",   "iShares MSCI Italy ETF",                          "ETF — Country: Italy"),
    ("EWP",   "iShares MSCI Spain ETF",                          "ETF — Country: Spain"),
    ("EWL",   "iShares MSCI Switzerland ETF",                    "ETF — Country: Switzerland"),
    ("EWD",   "iShares MSCI Sweden ETF",                         "ETF — Country: Sweden"),
    ("ENOR",  "iShares MSCI Norway ETF",                         "ETF — Country: Norway"),
    ("EDEN",  "iShares MSCI Denmark ETF",                        "ETF — Country: Denmark"),
    ("EFNL",  "iShares MSCI Finland ETF",                        "ETF — Country: Finland"),
    ("EIRL",  "iShares MSCI Ireland ETF",                        "ETF — Country: Ireland"),
    ("EWO",   "iShares MSCI Austria ETF",                        "ETF — Country: Austria"),
    ("EPOL",  "iShares MSCI Poland ETF",                         "ETF — Country: Poland"),
    ("EWA",   "iShares MSCI Australia ETF",                      "ETF — Country: Australia"),
    ("EWC",   "iShares MSCI Canada ETF",                         "ETF — Country: Canada"),
    ("EWZ",   "iShares MSCI Brazil ETF",                         "ETF — Country: Brazil"),
    ("INDA",  "iShares MSCI India ETF",                          "ETF — Country: India"),
    ("EWT",   "iShares MSCI Taiwan ETF",                         "ETF — Country: Taiwan"),
    ("EWY",   "iShares MSCI South Korea ETF",                    "ETF — Country: South Korea"),
    ("EZA",   "iShares MSCI South Africa ETF",                   "ETF — Country: South Africa"),
    ("EWW",   "iShares MSCI Mexico ETF",                         "ETF — Country: Mexico"),
    ("MCHI",  "iShares MSCI China ETF",                          "ETF — Country: China"),
]


def get_etfs() -> pd.DataFrame:
    """Return ETF rows shaped like a get_index_constituents() result.

    Schema: ``[Symbol, Name, Index]`` — Index is the per-ETF label so each
    ETF surfaces a distinct "Indexes" tag in Market (e.g. "ETF — Sector:
    Utilities"). They flow through aggregate_constituents alongside the
    real index constituents.
    """
    log.info(f"Adding {len(ETF_LIST)} curated ETFs")
    df = pd.DataFrame(
        [(sym, name, label) for sym, name, label in ETF_LIST],
        columns=["Symbol", "Name", "Index"],
    )
    return df


# Watchlists: (display label, source, reference).
# Sources:
#   "yahoo_screener"     → ref is a Yahoo predefined screener scrIds value.
#   "dataroma_url"       → ref is a dataroma path (e.g. "/m/g/portfolio.php").
#   "dataroma_activists" → aggregate across DATAROMA_ACTIVIST_CODES; ref ignored.
#
# Yahoo deprecated /u/yahoo-finance/watchlists/ pages in 2026; the original
# 8-list spec from CLAUDE.md is mostly gone. See MEMORY.md "2026-05-06" entries
# for the migration path. This is the slim 6-list replacement (Option B).
WATCHLISTS: list[tuple[str, str, str]] = [
    ("Recent 52-Week Highs",                  "yahoo_screener",     "recent_52_week_highs"),
    ("Berkshire Hathaway Portfolio",          "yahoo_screener",     "top_stocks_owned_by_warren_buffet"),
    ("Top Quarterly Buys (Super Investors)",  "dataroma_url",       "/m/g/portfolio_b.php?q=q&o=c"),
    ("Top Quarterly Sells (Super Investors)", "dataroma_url",       "/m/g/portfolio_s.php?q=q&o=c"),
    ("Most-Held by Super Investors",          "dataroma_url",       "/m/g/portfolio.php"),
    ("Activist Hedge Fund Positions",         "dataroma_activists", ""),
]

# Dataroma manager codes for the 7 activist funds we aggregate. Codes are
# case-sensitive in the URL ?m=... param.
DATAROMA_ACTIVIST_CODES = [
    "ic",   # Carl Icahn / Icahn Capital
    "psc",  # Bill Ackman / Pershing Square
    "VA",   # ValueAct Capital
    "TF",   # Nelson Peltz / Trian Fund Management
    "ENG",  # Glenn Welling / Engaged Capital
    "tp",   # Daniel Loeb / Third Point
    "tci",  # Christopher Hohn / TCI Fund Management
]
DATAROMA_BASE = "https://www.dataroma.com"

# FX: 1 EUR = X foreign currency, so foreign_in_eur = foreign_price / rate
FX_PAIRS = {"USD": "EURUSD=X", "JPY": "EURJPY=X", "GBP": "EURGBP=X", "CHF": "EURCHF=X"}

LOOKBACKS = {
    "Today":   dt.timedelta(days=0),
    "1D ago":  dt.timedelta(days=1),
    "1W ago":  dt.timedelta(days=7),
    "1M ago":  dt.timedelta(days=30),
    "6M ago":  dt.timedelta(days=182),
    "1Y ago":  dt.timedelta(days=365),
    "5Y ago":  dt.timedelta(days=int(5 * 365.25)),
}

# ---------------------------------------------------------------------------
# Ticker normalization
# ---------------------------------------------------------------------------

# Yahoo exchange suffixes we recognize. If a Wikipedia table already lists a
# ticker in fully-qualified Yahoo form, normalize_ticker passes it through
# instead of appending its own index's default suffix — prevents double-
# suffixing when the same stock appears in multiple indexes with different
# conventions (e.g., Airbus is "AIR.PA" in ESTOXX50/CAC40 and DAX's
# fully-qualified form is also "AIR.PA", not "AIR.PA.DE").
KNOWN_EXCHANGE_SUFFIXES = (
    ".PA", ".DE", ".AS", ".BR", ".MI", ".MC", ".HE", ".IR", ".L", ".T",
)


def normalize_ticker(symbol: str, index_name: str) -> str:
    """Convert a Wikipedia-listed symbol into Yahoo's ticker format."""
    s = str(symbol).strip().upper()
    s = s.split()[0]  # strip footnote markers like "AAA[1]"
    s = re.sub(r"\[.*?\]", "", s)

    # Already fully-qualified — pass through. This is the cross-index
    # de-dup guard: every per-index rule below would otherwise blindly
    # append its own suffix.
    if any(s.endswith(sfx) for sfx in KNOWN_EXCHANGE_SUFFIXES):
        return s

    if index_name == "SP500":
        return s.replace(".", "-")             # BRK.B -> BRK-B
    if index_name == "NIKKEI225":
        return s + ".T"
    if index_name == "FTSE100":
        s = s.rstrip(".")                      # "RR." -> "RR"
        # Sub-class shares like "BT.A" use a hyphen on Yahoo: BT-A.L.
        # rstrip("") above only strips a trailing dot; an interior "." (like
        # in BT.A) survives and must be converted to "-".
        if "." in s:
            s = s.replace(".", "-")
        return s + ".L"
    if index_name == "DAX":
        return s + ".DE"
    if index_name == "CAC40":
        return s + ".PA"
    if index_name == "BEL20":
        return s + ".BR"
    if index_name == "ESTOXX50":
        # Wikipedia's ESTOXX 50 table always lists fully-qualified tickers,
        # which the top-of-function suffix check has already handled. Anything
        # reaching here is malformed; pass through and let yfinance 404.
        return s
    return s

# ---------------------------------------------------------------------------
# Wikipedia: index constituents
# ---------------------------------------------------------------------------

_SYMBOL_HINTS = ("symbol", "ticker", "code", "epic")
_NAME_HINTS   = ("company", "security", "name", "issuer")

# BEL 20 Wikipedia formats tickers as "Euronext Brussels:\xa0ABI" — exchange is
# embedded in the cell. Map exchange word → Yahoo suffix.
_BEL20_EXCHANGE_SUFFIX = {"brussels": ".BR", "amsterdam": ".AS"}
_BEL20_TICKER_RE = re.compile(
    r"Euronext\s+(\w+)[\s:\xa0]+([A-Z][A-Z0-9]{0,9})", re.IGNORECASE
)

_NIKKEI_TICKER_RE = re.compile(r"\(\s*TYO\s*:\s*([0-9A-Z]{4,5})\s*\)")


_NIKKEI225_MIN_COUNT = 150  # safety floor; real value is ~225


def _parse_nikkei225_components(html: str, min_count: int = _NIKKEI225_MIN_COUNT) -> pd.DataFrame:
    """Extract (Symbol, Name, Index) rows from the Components section of the
    English Wikipedia Nikkei 225 page.

    Wikipedia stopped publishing a constituent table for Nikkei 225 sometime
    in 2025. The 225 stocks now live in bullet lists under sector
    subheadings, each entry formatted "Company Name (TYO: 9202)" with
    occasional trailing parenthetical annotations like
    "(Holding company for X)". Tickers can be 4 digits or 4 chars including
    letters (e.g. "543A").

    Raises RuntimeError if the result is implausibly small — protects
    against a future Wikipedia restructure silently degrading the workbook
    instead of surfacing as a visible error.
    """
    start = html.find('id="Components"')
    end = html.find('id="Statistics"')
    if start < 0 or end < 0:
        raise RuntimeError("Could not locate Components section on Nikkei 225 Wikipedia page")
    body = html[start:end]
    rows: list[tuple[str, str]] = []
    for raw in re.findall(r"<li[^>]*>(.+?)</li>", body, re.DOTALL):
        text = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", "", raw)).strip()
        m = _NIKKEI_TICKER_RE.search(text)
        if not m:
            continue
        ticker = m.group(1)
        # Name is the prefix before "(TYO:" — strip stray trailing punctuation
        name = text[: m.start()].rstrip().rstrip("(").strip()
        if name and ticker:
            rows.append((ticker, name))
    df = pd.DataFrame(rows, columns=["Symbol", "Name"])
    df["Symbol"] = df["Symbol"].apply(lambda s: normalize_ticker(s, "NIKKEI225"))
    df["Index"] = "NIKKEI225"
    df = df.drop_duplicates(subset=["Symbol"]).reset_index(drop=True)
    if len(df) < min_count:
        raise RuntimeError(
            f"Nikkei 225 parser yielded only {len(df)} constituents "
            f"(expected ~225, floor {min_count}). Wikipedia layout likely changed."
        )
    return df


def _parse_bel20_ticker(cell: str) -> str:
    """Extract 'SYMBOL.BR' or 'SYMBOL.AS' from a BEL 20 Wikipedia ticker cell.

    Returns "" if the cell doesn't match the expected format.
    """
    m = _BEL20_TICKER_RE.search(str(cell))
    if not m:
        return ""
    exchange = m.group(1).lower()
    sym = m.group(2).upper()
    suffix = _BEL20_EXCHANGE_SUFFIX.get(exchange, ".BR")
    return sym + suffix

def _find_constituent_table(tables: list[pd.DataFrame], min_rows: int) -> Optional[tuple[pd.DataFrame, str, str]]:
    """Heuristic: pick the table that has a symbol-like and a name-like column."""
    for tbl in tables:
        cols = [str(c) for c in tbl.columns]
        sym_col = next((c for c in cols if any(h in c.lower() for h in _SYMBOL_HINTS)), None)
        name_col = next((c for c in cols if any(h in c.lower() for h in _NAME_HINTS)), None)
        if sym_col and name_col and len(tbl) >= min_rows:
            return tbl, sym_col, name_col
    return None

def get_index_constituents(index_name: str) -> pd.DataFrame:
    """Returns DataFrame with columns [Symbol, Name, Index]. Symbol is Yahoo-formatted."""
    url = INDEX_WIKI[index_name]
    log.info(f"Fetching {index_name} constituents from Wikipedia")
    # We need a custom UA for Wikipedia too on some networks
    resp = requests.get(url, headers=HTTP_HEADERS, timeout=30)
    resp.raise_for_status()

    if index_name == "NIKKEI225":
        # Wikipedia restructured the page; no constituent table any more.
        df = _parse_nikkei225_components(resp.text)
        log.info(f"  {index_name}: {len(df)} constituents")
        return df

    tables = pd.read_html(io.StringIO(resp.text))
    expected_min = {"BEL20": 15, "CAC40": 30, "DAX": 30, "FTSE100": 80,
                    "NIKKEI225": 150, "SP500": 400, "ESTOXX50": 45}.get(index_name, 10)
    found = _find_constituent_table(tables, expected_min)
    if not found:
        # Fallback: lower threshold
        found = _find_constituent_table(tables, 10)
    if not found:
        raise RuntimeError(f"Could not locate constituent table for {index_name}")
    tbl, sym_col, name_col = found
    df = tbl[[sym_col, name_col]].copy()
    df.columns = ["Symbol", "Name"]
    df = df.dropna()
    df["Symbol"] = df["Symbol"].astype(str).str.strip()
    df["Name"] = df["Name"].astype(str).str.strip()
    if index_name == "BEL20":
        # Wikipedia formats BEL 20 tickers as "Euronext Brussels:\xa0SYMBOL"
        # or "Euronext Amsterdam:\xa0SYMBOL". Extract symbol + correct Yahoo suffix.
        df["Symbol"] = df["Symbol"].apply(_parse_bel20_ticker)
        df = df[df["Symbol"].astype(bool)]
    df = df[df["Symbol"].str.len().between(1, 12)]
    df["Symbol"] = df["Symbol"].apply(lambda s: normalize_ticker(s, index_name))
    df["Index"] = index_name
    df = df.drop_duplicates(subset=["Symbol"]).reset_index(drop=True)
    log.info(f"  {index_name}: {len(df)} constituents")
    return df

# ---------------------------------------------------------------------------
# Watchlists — Yahoo predefined screener API
# ---------------------------------------------------------------------------
#
# Yahoo's legacy watchlist pages (/u/yahoo-finance/watchlists/) are deprecated
# and don't render content. The same data (where it still exists) is now served
# by the predefined-screener API. Auth requires:
#   1. A session cookie (acquired from any Yahoo Finance page or fc.yahoo.com).
#   2. A "crumb" token from /v1/test/getcrumb — passed back as ?crumb=...
# Without both, /v1/finance/screener/predefined/saved returns total>0 but 0 records.
#
# The endpoint caps responses at 5 records per call regardless of count param,
# so we paginate via &start=N.

_SCREENER_API = "https://query1.finance.yahoo.com/v1/finance/screener/predefined/saved"
_SCREENER_PAGE_STEP = 5    # observed per-call cap
_SCREENER_MAX_RECORDS = 250  # safety stop per screener


# Note: we previously tried passing a shared requests.Session into yfinance
# calls to fix HTTP 401 "Invalid Crumb" bursts under parallel load. That
# approach FAILS — yfinance 0.2.x raises YFDataException demanding a
# curl_cffi session (TLS fingerprinting + browser-like behaviour). The
# fix the yfinance maintainers recommend is the OPPOSITE of what we did:
# "stop setting session, let YF handle." Yfinance manages its own
# curl_cffi session internally with crumb caching that's threadsafe
# enough — the 401s we observed are mostly retried by yfinance itself
# (visible as ERROR logs but invisible at the result level: 982 ok / 0
# failed). What does help is retry-on-empty in fetch_ticker_info to
# rescue the rare case where yfinance does give up.


def _yahoo_screener_session() -> Optional[tuple[requests.Session, str]]:
    """Acquire a session + crumb for the Yahoo screener API.

    Returns (session, crumb) on success, or None if auth fails (caller should
    skip Yahoo screeners and continue).
    """
    s = requests.Session()
    s.headers.update(HTTP_HEADERS)
    try:
        # fc.yahoo.com seeds the A1/A3 consent cookie used by yfinance too.
        s.get("https://fc.yahoo.com", timeout=10)
        s.get("https://finance.yahoo.com/research-hub/screener/", timeout=20)
        crumb = s.get(
            "https://query1.finance.yahoo.com/v1/test/getcrumb?lang=en-US&region=US",
            timeout=10,
        ).text.strip()
    except Exception as e:
        log.warning(f"Yahoo screener auth failed: {e}")
        return None
    if not crumb or len(crumb) > 64:
        log.warning(f"Yahoo screener auth: unexpected crumb {crumb!r}")
        return None
    return s, crumb

def fetch_yahoo_screener(scrid: str, sess: requests.Session, crumb: str) -> set[str]:
    """Fetch all tickers for a Yahoo predefined screener id (paginated).

    Returns a set of uppercase Yahoo symbols. Failures log a warning and return
    a partial / empty set rather than raising.
    """
    tickers: set[str] = set()
    start = 0
    while start < _SCREENER_MAX_RECORDS:
        try:
            r = sess.get(_SCREENER_API, params={
                "count": 50, "formatted": "true", "scrIds": scrid,
                "sortField": "", "sortType": "", "start": start,
                "useRecordsResponse": "true", "betaFeatureFlag": "true",
                "lang": "en-US", "region": "US", "crumb": crumb,
            }, timeout=20)
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            log.warning(f"Yahoo screener '{scrid}' page start={start} failed: {e}")
            break
        err = (data.get("finance") or {}).get("error")
        if err:
            log.warning(f"Yahoo screener '{scrid}' error: {err.get('description', err)}")
            break
        result = (data.get("finance") or {}).get("result") or []
        if not result:
            break
        records = result[0].get("records") or result[0].get("quotes") or []
        if not records:
            break
        for rec in records:
            sym = (rec.get("ticker") or rec.get("symbol") or "").strip().upper()
            if sym and not sym.startswith("^") and not sym.endswith("=F") and not sym.endswith("=X"):
                tickers.add(sym)
        if len(records) < _SCREENER_PAGE_STEP:
            break
        start += _SCREENER_PAGE_STEP
        time.sleep(0.25)
    return tickers

# ---------------------------------------------------------------------------
# Watchlists — Dataroma ("super investor" 13F aggregator)
# ---------------------------------------------------------------------------
#
# Dataroma serves plain HTML; ticker links have the form
#     <a href="/m/stock.php?sym=AAPL">AAPL</a>
# regardless of which list page (portfolio.php, portfolio_b.php, holdings.php).
# A single regex covers all of them.

_DATAROMA_SYM_RE = re.compile(r'/m/stock\.php\?sym=([A-Z][A-Z0-9\.]{0,11})', re.IGNORECASE)

def _extract_dataroma_tickers(html: str) -> set[str]:
    """Parse ticker symbols out of a Dataroma page's HTML.

    Dataroma uses dot-style for SP500 sub-classes (BRK.B, BF.B); Yahoo uses
    hyphen-style (BRK-B, BF-B). We emit BOTH so the membership dict matches
    whichever form the index constituents table contains.
    """
    out: set[str] = set()
    for raw in _DATAROMA_SYM_RE.findall(html):
        sym = raw.upper()
        out.add(sym)
        if "." in sym:
            out.add(sym.replace(".", "-"))
    return out

def fetch_dataroma_tickers(path: str) -> set[str]:
    """Fetch tickers from a Dataroma list page. ``path`` is the URL path or full URL."""
    url = DATAROMA_BASE + path if path.startswith("/") else path
    try:
        r = requests.get(url, headers=HTTP_HEADERS, timeout=20)
        r.raise_for_status()
    except Exception as e:
        log.warning(f"Dataroma fetch '{path}' failed: {e}")
        return set()
    return _extract_dataroma_tickers(r.text)

def fetch_dataroma_activist_aggregate() -> set[str]:
    """Aggregate holdings across the 7 tracked activist hedge fund managers."""
    out: set[str] = set()
    for code in DATAROMA_ACTIVIST_CODES:
        out |= fetch_dataroma_tickers(f"/m/holdings.php?m={code}")
        time.sleep(0.25)
    return out

# ---------------------------------------------------------------------------

def fetch_all_watchlists(max_workers: int = 6) -> dict[str, list[str]]:
    """Returns ticker -> [list of watchlist labels it belongs to].

    Speedup: the 6 watchlists are scraped in parallel via ThreadPoolExecutor.
    They have no inter-dependencies. The Yahoo screener entries share ONE
    auth session (acquired once up front); the dataroma entries hit a
    separate host. Total wall-clock drops from ~24s serial to ~5s.
    """
    log.info(f"Fetching watchlists (parallel, workers={max_workers})")
    membership: dict[str, list[str]] = {}
    membership_lock = threading.Lock()

    def _record(label: str, tickers: set[str]) -> None:
        log.info(f"  {label}: {len(tickers)} tickers")
        with membership_lock:
            for t in tickers:
                membership.setdefault(t, []).append(label)

    # Yahoo screener auth ONCE for all yahoo entries — workers share it.
    yahoo_entries = [(label, ref) for label, src, ref in WATCHLISTS if src == "yahoo_screener"]
    auth = _yahoo_screener_session() if yahoo_entries else None
    if yahoo_entries and auth is None:
        log.warning("Skipping Yahoo screener watchlists (auth unavailable)")

    # Build the per-watchlist task list. Each task is a no-arg callable
    # so the executor can fire them concurrently.
    tasks: list[tuple[str, Callable[[], set[str]]]] = []
    for label, src, ref in WATCHLISTS:
        if src == "yahoo_screener":
            if auth is None:
                continue
            sess, crumb = auth
            tasks.append((label, lambda r=ref, s=sess, c=crumb: fetch_yahoo_screener(r, s, c)))
        elif src == "dataroma_url":
            tasks.append((label, lambda r=ref: fetch_dataroma_tickers(r)))
        elif src == "dataroma_activists":
            tasks.append((label, fetch_dataroma_activist_aggregate))

    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        fut_to_label = {ex.submit(fn): label for label, fn in tasks}
        for fut in as_completed(fut_to_label):
            label = fut_to_label[fut]
            try:
                tickers = fut.result()
            except Exception as e:
                log.warning(f"  {label} failed: {e}")
                continue
            _record(label, tickers)

    return membership

# ---------------------------------------------------------------------------
# FX rates
# ---------------------------------------------------------------------------

def get_fx_rates() -> dict[str, float]:
    """Returns {currency_code: units of currency per 1 EUR}. EUR maps to 1.0."""
    rates: dict[str, float] = {"EUR": 1.0}
    log.info("Fetching FX rates (today)")
    for ccy, symbol in FX_PAIRS.items():
        try:
            hist = yf.Ticker(symbol).history(period="5d", auto_adjust=False)
            if hist.empty:
                raise RuntimeError("empty FX history")
            rate = float(hist["Close"].dropna().iloc[-1])
            rates[ccy] = rate
            log.info(f"  EUR/{ccy} = {rate:.4f}")
        except Exception as e:
            log.error(f"  EUR/{ccy} fetch failed: {e}")
            rates[ccy] = float("nan")
    # GBp (UK pence) shares same rate as GBP, but values must be /100 first
    rates["GBp"] = rates.get("GBP", float("nan"))
    return rates


def get_fx_history() -> dict[str, pd.Series]:
    """Returns {currency_code: Close-price Series indexed by date} for the
    full 5-year+ lookback window. Used to populate the Currencies sheet at
    each LOOKBACKS offset.

    Failed currencies map to an empty Series; price_at_or_before then
    returns None for those rows. The job never crashes on FX failure.
    """
    today = dt.date.today()
    start = today - dt.timedelta(days=int(5.2 * 365))
    end = today + dt.timedelta(days=1)
    out: dict[str, pd.Series] = {}
    log.info("Fetching FX history (5y)")
    for ccy, symbol in FX_PAIRS.items():
        try:
            hist = yf.Ticker(symbol).history(start=start, end=end, auto_adjust=False)
            if hist.empty:
                raise RuntimeError("empty FX history")
            out[ccy] = hist["Close"].dropna()
        except Exception as e:
            log.error(f"  EUR/{ccy} history fetch failed: {e}")
            out[ccy] = pd.Series(dtype=float)
    return out

def to_eur(price: float, currency: str, fx: dict[str, float]) -> Optional[float]:
    if price is None or pd.isna(price):
        return None
    if currency == "EUR":
        return float(price)
    if currency == "GBp":
        rate = fx.get("GBP")
        if rate is None or pd.isna(rate):
            return None
        return (float(price) / 100.0) / rate
    rate = fx.get(currency)
    if rate is None or pd.isna(rate):
        return None
    return float(price) / rate

# ---------------------------------------------------------------------------
# Price history & ticker metadata
# ---------------------------------------------------------------------------

def fetch_close_prices(tickers: list[str], start: dt.date, end: dt.date,
                       chunk_size: int = 80,
                       max_workers: int = 6) -> pd.DataFrame:
    """Returns a wide DataFrame of close prices: date index, ticker columns.

    Speedup: chunks are downloaded in parallel via ``ThreadPoolExecutor``.
    The previous 1.0s inter-chunk sleep is removed; the worker count is the
    new concurrency cap. ``yf.download`` is called with ``threads=False``
    because we're doing our own threading at the chunk level — yfinance's
    own per-ticker threading would oversubscribe.

    No ``session`` argument: yfinance 0.2.x requires its own curl_cffi
    session (TLS fingerprinting), so we let yfinance manage it.
    """
    log.info(f"Downloading price history for {len(tickers)} tickers ({start} → {end}, "
             f"parallel workers={max_workers})")
    chunks = [tickers[i:i + chunk_size] for i in range(0, len(tickers), chunk_size)]

    def _fetch_chunk(idx: int, chunk: list[str]) -> Optional[pd.DataFrame]:
        try:
            df = yf.download(
                chunk, start=start, end=end,
                progress=False, auto_adjust=False,
                group_by="column", threads=False,
            )
        except Exception as e:
            log.warning(f"  Price chunk {idx} failed: {e}")
            return None
        if df.empty:
            return None
        if isinstance(df.columns, pd.MultiIndex):
            if "Close" in df.columns.get_level_values(0):
                return df["Close"]
            return None
        if "Close" not in df.columns:
            return None
        return df[["Close"]].rename(columns={"Close": chunk[0]})

    all_closes: list[pd.DataFrame] = []
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        fut_to_idx = {ex.submit(_fetch_chunk, i, c): i
                      for i, c in enumerate(chunks, 1)}
        for fut in as_completed(fut_to_idx):
            idx = fut_to_idx[fut]
            closes = fut.result()
            log.info(f"  Price chunk {idx}/{len(chunks)} done"
                     f"{' (empty)' if closes is None else ''}")
            if closes is not None:
                all_closes.append(closes)

    if not all_closes:
        return pd.DataFrame()
    combined = pd.concat(all_closes, axis=1)
    combined = combined.loc[:, ~combined.columns.duplicated()]
    return combined

def fetch_ticker_info(ticker: str) -> dict:
    """Returns currency, P/E, name, sector, industry, and business description. Best-effort.

    All fields come from a single ``yf.Ticker(t).info`` call.

    Retry-on-empty: if the first call returns ``{}`` (the symptom of a
    transient 401 Invalid Crumb that yfinance gave up on), we sleep briefly
    and retry once. Truly delisted tickers also return empty; the extra
    retry costs ~0.3s per such case but rescues the much more common case
    of an auth blip.
    """
    def _one_call() -> dict:
        try:
            return yf.Ticker(ticker).info or {}
        except Exception as e:
            log.debug(f"info('{ticker}') raised: {e}")
            return {}

    info = _one_call()
    if not info:
        time.sleep(0.3)
        info = _one_call()
    return {
        "currency":    info.get("currency") or "",
        "trailingPE":  info.get("trailingPE"),
        "forwardPE":   info.get("forwardPE"),
        "longName":    info.get("longName") or info.get("shortName") or "",
        "sector":      info.get("sector") or "",
        "industry":    info.get("industry") or "",
        "description": info.get("longBusinessSummary") or "",
    }

def fetch_all_info(tickers: list[str], delay: float = 0.0,
                   progress: Optional[Callable[[int, int, str], None]] = None,
                   should_stop: Optional[Callable[[], bool]] = None,
                   stop_poll_every: int = 25,
                   max_workers: int = 8) -> dict[str, dict]:
    """Fetch yfinance .info for each ticker, in parallel.

    Speedup A: uses ``ThreadPoolExecutor`` with ``max_workers`` concurrent
    workers. yfinance .info is HTTP-bound so threads parallelize well — a
    ~1000-ticker run that took ~4 min sequential drops to ~30–60s.

    The ``delay`` arg is kept for backward compatibility with tests but is
    no longer applied between requests; the concurrency cap is the new
    rate-limiter.

    Optional callbacks:
      - ``progress(done, total, current_symbol)`` invoked as each ticker
        *completes* (not in submit order). Fires from the main thread.
      - ``should_stop()`` polled every ``stop_poll_every`` completions —
        when it returns True, pending futures are cancelled and the
        function returns the partial dict gathered so far. Caller can
        detect this via ``len(out) < len(tickers)``.
    """
    del delay  # kept in signature for callers/tests; no longer used
    total = len(tickers)
    log.info(f"Fetching .info for {total} tickers (parallel, workers={max_workers})")
    out: dict[str, dict] = {}
    stopped = False
    with ThreadPoolExecutor(max_workers=max_workers) as ex:
        future_to_ticker = {ex.submit(fetch_ticker_info, t): t for t in tickers}
        done = 0
        for fut in as_completed(future_to_ticker):
            t = future_to_ticker[fut]
            try:
                out[t] = fut.result()
            except Exception as e:
                log.debug(f"info('{t}') raised in worker: {e}")
                out[t] = {
                    "currency": "", "trailingPE": None, "forwardPE": None,
                    "longName": "", "sector": "", "industry": "", "description": "",
                }
            done += 1
            if done % 50 == 0 or done == 1:
                log.info(f"  info {done}/{total}")
            if progress is not None:
                try:
                    progress(done, total, t)
                except Exception as e:
                    log.debug(f"progress callback failed: {e}")
            if should_stop is not None and (done == 1 or done % stop_poll_every == 0):
                try:
                    if should_stop():
                        log.info(f"  STOP requested at {done}/{total} — cancelling pending .info futures")
                        stopped = True
                        break
                except Exception as e:
                    log.debug(f"should_stop callback raised: {e}")
        if stopped:
            for f in future_to_ticker:
                if not f.done():
                    f.cancel()
    return out

class InfoCache:
    """Speedup F: persistent SQLite cache for yfinance .info results.

    Used by rebuild_inventory to skip the slow per-ticker .info fetch when
    we already have a recent result. The cached fields (currency, sector,
    longName, description) change rarely; a 7-day TTL refreshes naturally
    against renames or sector reclassifications without paying the full
    multi-minute scan on every rebuild.

    Not used by get_quotes — P/E moves daily and refreshing it is the
    whole point of running get_quotes.
    """

    def __init__(self, path: Path):
        self.path = Path(path)
        self._conn = sqlite3.connect(str(self.path))
        self._conn.execute(
            "CREATE TABLE IF NOT EXISTS info_cache ("
            "ticker TEXT PRIMARY KEY, currency TEXT, trailing_pe REAL, "
            "forward_pe REAL, long_name TEXT, sector TEXT, industry TEXT, "
            "description TEXT, fetched_at REAL NOT NULL)"
        )
        # Lightweight migration: older caches predate the industry column.
        try:
            self._conn.execute("ALTER TABLE info_cache ADD COLUMN industry TEXT")
            self._conn.commit()
        except sqlite3.OperationalError:
            pass  # column already exists
        self._conn.commit()

    def get_fresh(self, tickers: list[str], ttl_seconds: float) -> dict[str, dict]:
        if not tickers:
            return {}
        cutoff = time.time() - ttl_seconds
        placeholders = ",".join("?" * len(tickers))
        cur = self._conn.execute(
            f"SELECT ticker, currency, trailing_pe, forward_pe, long_name, sector, industry, description "
            f"FROM info_cache WHERE fetched_at >= ? AND ticker IN ({placeholders})",
            [cutoff, *tickers],
        )
        out: dict[str, dict] = {}
        for t, ccy, tpe, fpe, ln, sec, ind, desc in cur.fetchall():
            out[t] = {
                "currency": ccy or "",
                "trailingPE": tpe,
                "forwardPE": fpe,
                "longName": ln or "",
                "sector": sec or "",
                "industry": ind or "",
                "description": desc or "",
            }
        return out

    def put_many(self, items: dict[str, dict]) -> None:
        if not items:
            return
        now = time.time()
        rows = [
            (
                t,
                (d.get("currency") or ""),
                d.get("trailingPE"),
                d.get("forwardPE"),
                (d.get("longName") or ""),
                (d.get("sector") or ""),
                (d.get("industry") or ""),
                (d.get("description") or ""),
                now,
            )
            for t, d in items.items()
        ]
        self._conn.executemany(
            "INSERT OR REPLACE INTO info_cache (ticker, currency, trailing_pe, forward_pe, "
            "long_name, sector, industry, description, fetched_at) VALUES (?,?,?,?,?,?,?,?,?)",
            rows,
        )
        self._conn.commit()

    def close(self) -> None:
        try:
            self._conn.close()
        except Exception:
            pass


def fetch_all_info_with_cache(
    tickers: list[str],
    cache_path: Path,
    ttl_seconds: Optional[float] = None,
    progress: Optional[Callable[[int, int, str], None]] = None,
    should_stop: Optional[Callable[[], bool]] = None,
    max_workers: int = 8,
) -> tuple[dict[str, dict], int]:
    """Speedup F wrapper around ``fetch_all_info`` — returns ``(info_map, cached_count)``.

    Cache hits short-circuit the .info network call; misses go to the
    parallel fetcher. Fresh results are persisted back to the cache. The
    progress callback fires only for the miss subset (the caller can
    surface ``cached_count`` separately in status text).
    """
    if ttl_seconds is None:
        ttl_seconds = INFO_CACHE_TTL_DAYS * 86400
    cache = InfoCache(Path(cache_path))
    try:
        cached = cache.get_fresh(tickers, ttl_seconds)
        misses = [t for t in tickers if t not in cached]
        if misses:
            fresh = fetch_all_info(
                misses, progress=progress, should_stop=should_stop,
                max_workers=max_workers,
            )
            cache.put_many(fresh)
        else:
            fresh = {}
        return {**cached, **fresh}, len(cached)
    finally:
        cache.close()


def _info_cache_path_for(workbook_path: Optional[Path]) -> Path:
    """Pick the SQLite cache location next to the workbook (or cwd if unknown)."""
    base = Path(workbook_path).parent if workbook_path else Path.cwd()
    return base / INFO_CACHE_FILENAME


def is_recently_updated(last_update_iso: Optional[str], threshold_hours: float,
                        now: Optional[dt.datetime] = None) -> bool:
    """Return True if ``last_update_iso`` is within ``threshold_hours`` of ``now``.

    Accepts the ISO-Z form ``YYYY-MM-DDTHH:MM:SSZ`` that get_quotes writes
    into Market!"Last update (UTC)". Returns False on empty / unparseable
    input — i.e. a missing or malformed timestamp means "not fresh, refetch".
    """
    if not last_update_iso:
        return False
    try:
        ts = dt.datetime.fromisoformat(str(last_update_iso).replace("Z", "+00:00"))
    except Exception:
        return False
    if ts.tzinfo is None:
        ts = ts.replace(tzinfo=dt.UTC)
    now = now or dt.datetime.now(dt.UTC)
    return (now - ts).total_seconds() < threshold_hours * 3600

# ---------------------------------------------------------------------------
# Price lookup
# ---------------------------------------------------------------------------

def price_at_or_before(series: pd.Series, target: dt.date) -> Optional[float]:
    """Last available close on or before target date."""
    if series is None or series.empty:
        return None
    s = series.dropna()
    if s.empty:
        return None
    target_ts = pd.Timestamp(target)
    # series may be tz-aware
    if s.index.tz is not None:
        target_ts = target_ts.tz_localize(s.index.tz)
    valid = s[s.index <= target_ts]
    if valid.empty:
        return None
    return float(valid.iloc[-1])

# ---------------------------------------------------------------------------
# Report assembly
# ---------------------------------------------------------------------------

def aggregate_constituents(per_index: list[pd.DataFrame]) -> pd.DataFrame:
    """Combine per-index constituent DataFrames into one row per Symbol.

    Input rows: ``[Symbol, Name, Index]`` (one row per Symbol per Index).
    Output rows: ``[Symbol, Name, Indexes]`` where Indexes is a comma-separated
    list preserving the order in which indexes were fetched. A stock appearing
    in DAX and EuroStoxx 50 yields one row with ``Indexes = "DAX, ESTOXX50"``.
    """
    if not per_index:
        return pd.DataFrame(columns=["Symbol", "Name", "Indexes"])
    combined = pd.concat(per_index, ignore_index=True)
    if combined.empty:
        return pd.DataFrame(columns=["Symbol", "Name", "Indexes"])

    # Preserve first-seen Name for each Symbol.
    name_map = combined.drop_duplicates(subset=["Symbol"], keep="first").set_index("Symbol")["Name"]
    # Aggregate indexes per Symbol, preserving order, deduping.
    def _join_unique(seq):
        seen = []
        for x in seq:
            if x not in seen:
                seen.append(x)
        return ", ".join(seen)
    indexes_map = combined.groupby("Symbol", sort=False)["Index"].apply(_join_unique)

    out = pd.DataFrame({
        "Symbol":  indexes_map.index,
        "Name":    [name_map.get(s, "") for s in indexes_map.index],
        "Indexes": indexes_map.values,
    })
    return out.reset_index(drop=True)


def build_report(indexes: list[str], info_delay: float) -> tuple[pd.DataFrame, dict]:
    # 1. constituents — one row per Symbol with all index memberships joined,
    #    plus the curated ETF set (one row per ETF).
    parts = [get_index_constituents(idx) for idx in indexes]
    parts.append(get_etfs())
    constituents = aggregate_constituents(parts)
    log.info(f"Total unique tickers: {len(constituents)}")

    # 2. watchlist memberships
    wl_membership = fetch_all_watchlists()

    # 3. FX
    fx = get_fx_rates()

    # 4. price history (5y + buffer)
    today = dt.date.today()
    start = today - dt.timedelta(days=int(5.2 * 365))
    end   = today + dt.timedelta(days=1)
    closes = fetch_close_prices(list(constituents["Symbol"]), start, end)

    # 5. metadata
    info_map = fetch_all_info(list(constituents["Symbol"]), delay=info_delay)

    # 6. assemble rows
    log.info("Assembling rows")
    rows = []
    for _, c in constituents.iterrows():
        sym = c["Symbol"]
        info = info_map.get(sym, {})
        # Pick the first listed index as the currency fallback (their default
        # currencies disagree only across regions; intersection cases are EUR-EUR).
        first_index = c["Indexes"].split(",")[0].strip()
        ccy = info.get("currency") or INDEX_DEFAULT_CCY.get(first_index, "")
        series = closes[sym] if sym in closes.columns else pd.Series(dtype=float)

        row = {
            "Indexes":  c["Indexes"],
            "Symbol":   sym,
            "Name":     info.get("longName") or c["Name"],
            "Sector":   info.get("sector") or "",
            "Industry": info.get("industry") or "",
            "Currency": ccy,
        }
        for label, delta in LOOKBACKS.items():
            target = today - delta
            native = price_at_or_before(series, target)
            row[f"{label} (EUR)"] = to_eur(native, ccy, fx)

        row["P/E (TTM)"]   = info.get("trailingPE")
        row["Forward P/E"] = info.get("forwardPE")

        # Watchlist membership: try sym, sym without suffix, raw US ticker
        sym_root = re.sub(r"\.[A-Z]{1,3}$", "", sym)
        labels: list[str] = []
        for key in {sym, sym_root}:
            labels.extend(wl_membership.get(key, []))
        seen = set()
        labels_dedup = [x for x in labels if not (x in seen or seen.add(x))]
        row["Watchlists"]  = ", ".join(labels_dedup)
        row["Description"] = info.get("description") or ""

        rows.append(row)

    df = pd.DataFrame(rows)
    return df, fx

# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

def write_excel(df: pd.DataFrame, fx: dict, output_path: Path) -> None:
    log.info(f"Writing {output_path}")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="All", index=False)
        meta = pd.DataFrame({
            "Field": [
                "Run timestamp (UTC)", "Total tickers",
                "EUR/USD (1 EUR = X USD)", "EUR/JPY (1 EUR = X JPY)", "EUR/GBP (1 EUR = X GBP)",
                "Note",
            ],
            "Value": [
                dt.datetime.now(dt.UTC).strftime("%Y-%m-%dT%H:%M:%SZ"),
                len(df),
                fx.get("USD"), fx.get("JPY"), fx.get("GBP"),
                "Historical prices converted using TODAY's FX rate.",
            ],
        })
        meta.to_excel(writer, sheet_name="Metadata", index=False)

        ws = writer.sheets["All"]
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.font = Font(bold=True)

        # Auto-size columns
        for col_idx, col in enumerate(df.columns, 1):
            sample = df[col].head(200)
            max_len = max([len(str(col))] + [len(str(s)) for s in sample])
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

        # Number formats
        numeric_cols = [c for c in df.columns if "(EUR)" in c or c in ("P/E (TTM)", "Forward P/E")]
        for col_name in numeric_cols:
            col_idx = df.columns.get_loc(col_name) + 1
            for row_idx in range(2, len(df) + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

# ---------------------------------------------------------------------------
# Persistent workbook — Main + Market sheets
# ---------------------------------------------------------------------------
#
# The "two jobs" architecture stores everything in a single .xlsx (later .xlsm
# with xlwings buttons). Two sheets:
#
#   Main   — user-maintained portfolio plus job controls, status, metadata.
#            Survives every job run; rebuild_inventory and get_quotes never
#            modify rows the user owns (portfolio area), only the named cells.
#   Market — full inventory of stocks across all tracked indexes. Job 1
#            (rebuild_inventory) overwrites this in full; Job 2 (get_quotes)
#            updates the quote/PE/timestamp columns by Symbol key.

# Schema version. Bumped on major changes that add sheets, columns, or change
# the data contract. The filename embeds this so the user can see at a glance
# which generation of the script their workbook matches; the Help sheet renders
# the changelog entries from VERSION_HISTORY below.
SCHEMA_VERSION = "v04"

# Append-only changelog. init-workbook appends any missing versions to the
# Help sheet without touching user edits. Date is when the version was minted.
VERSION_HISTORY: list[tuple[str, str, str]] = [
    ("v01", "2026-05-13",
     "Initial release. 7 indexes + 38 ETFs in Market. Watchlists from Yahoo "
     "screener + dataroma. Symbol hyperlinks. % change columns. xlwings "
     "buttons. Sheets: Main / Market / Help / xlwings.conf."),
    ("v02", "2026-05-13",
     "Help sheet now auto-populated from VERSION_HISTORY (append-only). New "
     "Currencies sheet listing EUR / USD / JPY / GBP / CHF rates today + at "
     "each lookback. New Monthly movers sheet ranking the biggest 1M gainers "
     "that have NOT weakened over 1D or 1W. CHF added to FX_PAIRS. Market "
     "AutoFilter is re-applied to the new data extent on rebuild."),
    ("v03", "2026-05-13",
     "Renamed Monthly movers → Monthly winners. New Monthly losers sheet "
     "(stocks with sustained 1M decline, also negative on 1D and 1W). Both "
     "ranking sheets gain an Owned? column. Ranking is now computed from the "
     "full Market sheet on every get_quotes — previously test mode left it "
     "empty because only the refreshed row was considered. AutoFilter is "
     "now on Monthly winners + Monthly losers; the Losers sheet has the "
     "Owned?=Yes filter pre-selected. Existing 'Monthly movers' sheets are "
     "auto-renamed on next init-workbook or get-quotes."),
    ("v04", "2026-05-14",
     "New Industry column in Market (between Sector and Watchlists) — "
     "sourced from yfinance .info['industry']. AutoFilter now applied on "
     "every result sheet (Market + Monthly winners + Monthly losers + "
     "Currencies). Portfolio auto-adoption: unresolved Main!Portfolio "
     "symbols are looked up on Yahoo and added to the universe with "
     "Indexes='Portfolio'; truly unresolvable ones surface as an error "
     "in Main col C of the offending row. Speedups: parallel .info "
     "(8 workers), parallel price chunks (6 workers), parallel watchlists "
     "(6 workers), SQLite cache for slow-changing .info fields (7-day TTL), "
     "freshness skip in get_quotes (4h). The session-passing experiment "
     "from a draft of v04 was reverted — yfinance 0.2.x requires its own "
     "curl_cffi session. NOTE: existing v03 workbooks need a full "
     "rebuild-inventory after upgrading so Market headers align with the "
     "new Industry column position."),
]

DEFAULT_WORKBOOK_PATH = Path(f"stocks_picker_{SCHEMA_VERSION}.xlsm")

# Test mode trims rebuild_inventory aggressively so the smoke completes in
# a few seconds — top N BEL 20 stocks, no ETFs, no watchlist HTTP. The
# .info loop is still exercised so any Yahoo regression surfaces.
TEST_MODE_TICKER_LIMIT = 5

# Test-mode get_quotes refreshes the top N Market rows (not just 1). Without
# this, the post-loop ranking computation finds nothing useful: it reads all
# Market rows, but if only 1 row has fresh quote data, the other rows have
# None for the % columns and get filtered out → empty Monthly winners/losers.
# 20 is a reasonable balance: ~5 sec for the .info loop, enough rows for
# the ranking to find candidates even in a freshly-rebuilt Market.
TEST_MODE_QUOTE_REFRESH_LIMIT = 20

# Speedup C: if Market!"Last update (UTC)" is within this many hours, skip
# the .info refresh for that ticker on the next get_quotes run. P/E + name
# values from the prior run are kept in place. Prices are still refreshed
# every run (cheap, batched). 4h is short enough that intraday re-runs
# don't go stale and long enough to make a sub-hour re-run feel snappy.
QUOTE_FRESHNESS_HOURS = 4.0

# Speedup F: SQLite cache for slow-changing .info fields used by
# rebuild_inventory (currency, sector, longName, description). P/E is
# excluded — it changes daily and gets refreshed by get_quotes anyway.
INFO_CACHE_FILENAME = "stocks_info_cache.sqlite"
INFO_CACHE_TTL_DAYS = 7

# Where button-triggered errors land. Lives next to the workbook so the user
# can find it without leaving Excel.
ERROR_LOG_FILENAME = "stocks_errors.log"
ERRORS_SHEET_NAME = "Errors"
HELP_SHEET_NAME = "Help"
CURRENCIES_SHEET_NAME = "Currencies"
MONTHLY_WINNERS_SHEET_NAME = "Monthly winners"   # renamed from "Monthly movers" in v03
MONTHLY_LOSERS_SHEET_NAME = "Monthly losers"
MONTHLY_MOVERS_LEGACY_NAME = "Monthly movers"    # migrated from any existing workbook


def _append_error_log(log_path: Path, exc_type: str, message: str, traceback_text: str) -> None:
    """Append a timestamped block to the error log next to the workbook."""
    timestamp = dt.datetime.now(dt.UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    block = (
        "\n" + "=" * 70 + "\n"
        f"{timestamp}  {exc_type}: {message}\n"
        + "=" * 70 + "\n"
        f"{traceback_text}\n"
    )
    try:
        with open(log_path, "a", encoding="utf-8") as f:
            f.write(block)
    except Exception:
        pass  # never let the error handler itself crash

# Yahoo Finance quote URL template. Used to hyperlink each Market!Symbol cell.
YAHOO_QUOTE_URL = "https://finance.yahoo.com/quote/{symbol}"


def yahoo_quote_url(symbol: str) -> str:
    """Build the Yahoo Finance quote URL for a Yahoo-formatted ticker.

    Yahoo's URL handles every suffix we use (.BR, .AS, .DE, .PA, .L, .T,
    .MI, etc.) as-is — no special encoding needed beyond %5E for ^-prefixed
    indexes, which we don't surface.
    """
    return YAHOO_QUOTE_URL.format(symbol=symbol)


_VERSIONED_FILENAME_RE = re.compile(r"^stocks_picker_v(\d+)\.xlsm$", re.IGNORECASE)


def _resolve_workbook(arg_path: str) -> Path:
    """Smart default: when the caller passed our default path, prefer an
    existing workbook over creating a new one.

    Resolution order:
      1. Highest-version stocks_picker_v{NN}.xlsm in the current directory.
      2. stocks.xlsm (pre-versioning filename).
      3. stocks.xlsx (the openpyxl-only filename before setup-buttons).
      4. The caller's path verbatim.

    A non-default explicit path always wins.
    """
    p = Path(arg_path)
    if p != DEFAULT_WORKBOOK_PATH:
        return p
    candidates: list[tuple[int, Path]] = []
    for f in Path(".").iterdir():
        m = _VERSIONED_FILENAME_RE.match(f.name)
        if m:
            candidates.append((int(m.group(1)), f))
    if candidates:
        return max(candidates, key=lambda x: x[0])[1]
    if Path("stocks.xlsm").exists():
        return Path("stocks.xlsm")
    if Path("stocks.xlsx").exists():
        return Path("stocks.xlsx")
    return p

# Market sheet column order. Fixed by index — every part of the code that
# reads/writes Market locates columns by name via this list. % change
# columns are interleaved with each price lookback so a row scan reads
# "price | % change | price | % change | ..." across.
MARKET_COLUMNS = [
    "Symbol", "Name", "Owned?", "Indexes", "Sector", "Industry", "Watchlists", "Currency",
    "Today (EUR)",
    "1D ago (EUR)", "1D %",
    "1W ago (EUR)", "1W %",
    "1M ago (EUR)", "1M %",
    "6M ago (EUR)", "6M %",
    "1Y ago (EUR)", "1Y %",
    "5Y ago (EUR)", "5Y %",
    "P/E (TTM)", "Forward P/E",
    "Description", "Last update (UTC)", "Last error",
]

# (lookback_label, percent_column_name) pairs derived from LOOKBACKS for the
# get_quotes loop. Excludes "Today" since the % is relative to Today.
def _pct_column_pairs() -> list[tuple[str, str]]:
    return [(lbl, lbl.replace(" ago", "") + " %") for lbl in LOOKBACKS if lbl != "Today"]


def _pct_change(today: Optional[float], past: Optional[float]) -> Optional[float]:
    """Fractional change from ``past`` to ``today``. Returns None on bad inputs;
    callers write the result with a percent number format."""
    if today is None or past is None or past == 0:
        return None
    try:
        return (float(today) - float(past)) / float(past)
    except (TypeError, ValueError):
        return None

# Main sheet — named cells. All controls/metadata live ABOVE the portfolio
# area so the portfolio can grow freely without colliding with them.
# The Test-mode cell uses a ";;;" number format so the TRUE/FALSE the
# checkbox writes is hidden — the checkbox is the visual indicator.
MAIN_CELLS: dict[str, str] = {
    "TestMode":       "B5",   # TRUE / FALSE (display hidden; see init_workbook)
    "Status":         "B6",   # live progress text written during job runs
    "LastRebuildAt":  "B7",
    "LastQuotesAt":   "B8",
    "MarketRowCount": "B9",
    "EurUsd":         "B10",
    "EurJpy":         "B11",
    "EurGbp":         "B12",
    "StopRequested":  "B13",  # checkbox-linked; Python polls during .info loop
    "JobRunning":     "B14",  # VBA-flipped wrapper guard against double-clicks
}

# Portfolio area on Main: header row + many rows for user entries.
PORTFOLIO_HEADER_ROW = 15
PORTFOLIO_FIRST_ROW = 16
PORTFOLIO_LAST_ROW = 999  # plenty of room; never collides with controls above


def _bold(cell, size: int = 11) -> None:
    cell.font = Font(bold=True, size=size)


def _layout_main_sheet(ws: Worksheet, *, overwrite: bool = True) -> None:
    """Write the static layout of the Main sheet — title, controls + metadata
    at the top, then portfolio header. Does NOT touch user data cells.

    With overwrite=False, only fills cells that are currently blank — used on
    re-runs of init-workbook so user cosmetic edits / renames survive.

    Layout:
      Row 1   : title
      Row 3   : "Jobs" section
      Row 4   : (xlwings button hint)
      Row 5   : Test mode label + checkbox cell (B5)
      Row 6   : Status label + live progress cell (B6)
      Row 8   : "Metadata" section
      Row 9-14: timestamps + FX rates
      Row 16  : "Portfolio" section
      Row 17  : Portfolio header (Symbol / Notes)
      Row 18+ : user portfolio entries
    """
    def _set(addr: str, value, *, bold: bool = False, size: int = 11,
             italic: bool = False, color: Optional[str] = None) -> None:
        cell = ws[addr]
        if not overwrite:
            # Re-run path: never touch the cell. This respects three kinds of
            # user edits: renames (value changed), clearances (value cleared
            # to None), and bespoke styling. The trade-off is that new labels
            # we add in future versions of the code won't appear in old
            # workbooks unless the user deletes the file.
            return
        cell.value = value
        font_kwargs = {"size": size}
        if bold: font_kwargs["bold"] = True
        if italic: font_kwargs["italic"] = True
        if color: font_kwargs["color"] = color
        cell.font = Font(**font_kwargs)

    _set("A1", "Stock Picker", bold=True, size=16)

    _set("A3", "Jobs", bold=True, size=12)
    _set("A5", "Test mode (only refresh BEL20 + 1 quote):")
    _set("A6", "Status:")

    _set("A7",  "Last rebuild_inventory:")
    _set("A8",  "Last get_quotes:")
    _set("A9",  "Total rows in Market:")
    _set("A10", "EUR/USD (1 EUR = X USD):")
    _set("A11", "EUR/JPY (1 EUR = X JPY):")
    _set("A12", "EUR/GBP (1 EUR = X GBP):")

    _set("A13", "Stop a running job:")
    _set("A14", "Portfolio (manual — list every Symbol you own)", bold=True, size=12)

    _set(ws.cell(row=PORTFOLIO_HEADER_ROW, column=1).coordinate, "Symbol", bold=True)
    _set(ws.cell(row=PORTFOLIO_HEADER_ROW, column=2).coordinate, "Notes", bold=True)

    # Only on fresh creation: hide the TestMode / StopRequested / JobRunning
    # cell displays (the user shouldn't see TRUE/FALSE next to checkboxes
    # or in tracker cells) and set column widths. On re-runs we don't touch
    # any of these — respects user cosmetic edits.
    if overwrite:
        for hidden_addr in (MAIN_CELLS["TestMode"],
                            MAIN_CELLS["StopRequested"],
                            MAIN_CELLS["JobRunning"]):
            ws[hidden_addr].number_format = ";;;"
        ws.column_dimensions["A"].width = 38
        ws.column_dimensions["B"].width = 60


# Excel's "Comma Style" (the ribbon button) accounting-style format.
# Negative numbers shown with a leading minus + space alignment, zero as "-".
COMMA_STYLE = '_-* #,##0.00_-;-* #,##0.00_-;_-* "-"??_-;_-@_-'
PERCENT_STYLE = "0.00%"


def _layout_market_sheet(ws: Worksheet, *, overwrite: bool = True) -> None:
    """Write Market sheet headers, freeze pane, column widths. Does not touch
    data rows.

    With overwrite=False (re-run path), header text is always re-asserted
    (the script's column-name lookups depend on it) but bold + fill styling
    and column widths are left alone so user cosmetic choices survive.
    """
    for col_idx, name in enumerate(MARKET_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx)
        if overwrite or cell.value is None or cell.value != name:
            cell.value = name
        if overwrite:
            _bold(cell)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
    ws.freeze_panes = "A2"

    if overwrite:
        widths = {
            "Symbol": 10, "Name": 28, "Owned?": 8, "Indexes": 18,
            "Sector": 22, "Industry": 26,
            "Watchlists": 38, "Currency": 10,
            "Today (EUR)": 12, "1D ago (EUR)": 12, "1W ago (EUR)": 12, "1M ago (EUR)": 12,
            "6M ago (EUR)": 12, "1Y ago (EUR)": 12, "5Y ago (EUR)": 12,
            "1D %": 9, "1W %": 9, "1M %": 9, "6M %": 9, "1Y %": 9, "5Y %": 9,
            "P/E (TTM)": 10, "Forward P/E": 10,
            "Description": 60, "Last update (UTC)": 20, "Last error": 30,
        }
        for col_idx, name in enumerate(MARKET_COLUMNS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 12)


def _market_col(name: str) -> int:
    """1-based column index in Market for a header name. Raises if unknown."""
    return MARKET_COLUMNS.index(name) + 1


def _check_market_headers(actual_headers: list) -> Optional[str]:
    """Verify a header row matches MARKET_COLUMNS exactly. Returns None on
    match, else an error message. Used to detect a stale-schema workbook
    before get_quotes writes quote data into wrong columns. v04 added
    Industry after Sector — an older v03 workbook would have its quote
    columns shifted by 1 unless rebuild_inventory has been re-run since
    the upgrade.
    """
    for col_idx, expected in enumerate(MARKET_COLUMNS, 1):
        actual = actual_headers[col_idx - 1] if col_idx <= len(actual_headers) else None
        if actual != expected:
            return (f"Market column {col_idx} header is {actual!r} but should be "
                    f"{expected!r}. The workbook was built on a previous schema "
                    f"(current = {SCHEMA_VERSION}). Run `rebuild-inventory` first "
                    "to refresh the Market layout.")
    return None


def _validate_market_layout(market_ws: Worksheet) -> Optional[str]:
    """openpyxl version of _check_market_headers."""
    headers = [market_ws.cell(row=1, column=i).value
               for i in range(1, len(MARKET_COLUMNS) + 1)]
    return _check_market_headers(headers)


_TEST_MODE_TRUTHY = {"TRUE", "T", "YES", "Y", "1"}


def read_stop_requested(main_ws: Worksheet) -> bool:
    """Interpret the Main!StopRequested cell (openpyxl). The STOP checkbox
    set up by setup-buttons flips this cell to TRUE; Python's .info loop
    polls it and breaks out cleanly when set."""
    raw = main_ws[MAIN_CELLS["StopRequested"]].value
    if raw is None:
        return False
    if isinstance(raw, bool):
        return raw
    return str(raw).strip().upper() in _TEST_MODE_TRUTHY


def read_test_mode(main_ws: Worksheet) -> bool:
    """Interpret the Main!TestMode cell as a boolean.

    Tolerates the cell being TRUE/FALSE strings (default seed), a literal
    Python bool (some xlwings paths return that), 1/0, "Yes"/"No", or blank
    (treated as False).
    """
    raw = main_ws[MAIN_CELLS["TestMode"]].value
    if raw is None:
        return False
    if isinstance(raw, bool):
        return raw
    return str(raw).strip().upper() in _TEST_MODE_TRUTHY


def read_portfolio_symbols(main_ws: Worksheet) -> set[str]:
    """Read the Main sheet's manual portfolio area, returning normalized symbols."""
    out: set[str] = set()
    for r in range(PORTFOLIO_FIRST_ROW, PORTFOLIO_LAST_ROW + 1):
        v = main_ws.cell(row=r, column=1).value
        if v is None:
            continue
        sym = str(v).strip().upper()
        if sym:
            out.add(sym)
    return out


def _owned_for(symbol: str, portfolio: set[str]) -> str:
    """Return "Yes" if the Market symbol matches a portfolio entry, else "No".

    Tolerates suffix mismatch: portfolio entry "KBC" matches Market "KBC.BR".
    """
    if symbol in portfolio:
        return "Yes"
    root = re.sub(r"\.[A-Z]{1,3}$", "", symbol)
    if root in portfolio:
        return "Yes"
    return "No"


def _load_xl(path: Path):
    """load_workbook with keep_vba=True for .xlsm (preserves user-imported VBA)."""
    if path.suffix.lower() == ".xlsm":
        return load_workbook(path, keep_vba=True)
    return load_workbook(path)


def init_workbook(path: Path) -> Path:
    """Create or refresh the persistent workbook layout.

    Idempotent: if ``path`` exists, this function preserves all user data and
    only re-asserts headers, section labels, and column widths. It will never
    blow away the portfolio area, the Market data rows, or any VBA the user
    has imported into the workbook.

    Note: openpyxl can't author a macro-enabled .xlsm from scratch — Excel
    rejects the content-type mismatch. If the caller asked for a .xlsm path
    that doesn't yet exist, we redirect to the .xlsx counterpart; the
    subsequent ``setup-buttons`` run does the .xlsm conversion via Excel COM.
    """
    path = Path(path)
    if not path.exists() and path.suffix.lower() == ".xlsm":
        redirected = path.with_suffix(".xlsx")
        log.info(f"openpyxl can't author .xlsm from scratch — writing "
                 f"{redirected.name} instead (setup-buttons will convert).")
        path = redirected
    if path.exists():
        log.info(f"Refreshing workbook layout in {path} (data + cosmetics preserved)")
        wb = _load_xl(path)
        main_ws = wb["Main"] if "Main" in wb.sheetnames else wb.create_sheet("Main", 0)
        market_ws = wb["Market"] if "Market" in wb.sheetnames else wb.create_sheet("Market", 1)
        overwrite = False
    else:
        log.info(f"Creating workbook {path}")
        wb = Workbook()
        # Default sheet "Sheet" → rename to Main; add Market.
        main_ws = wb.active
        main_ws.title = "Main"
        market_ws = wb.create_sheet("Market")
        overwrite = True

    _layout_main_sheet(main_ws, overwrite=overwrite)
    _layout_market_sheet(market_ws, overwrite=overwrite)

    # Auto-rename any legacy "Monthly movers" sheet to "Monthly winners".
    _migrate_monthly_movers_to_winners_openpyxl(wb)

    # Ensure the Help sheet exists and contains a row for every entry in
    # VERSION_HISTORY. Append-only — never modifies rows the user has edited.
    if HELP_SHEET_NAME not in wb.sheetnames:
        help_ws = wb.create_sheet(HELP_SHEET_NAME)
    else:
        help_ws = wb[HELP_SHEET_NAME]
    _ensure_help_sheet_versions(help_ws, fresh=overwrite)

    # Test-mode cell is intentionally left blank on fresh workbooks — the
    # checkbox added by setup-buttons writes TRUE/FALSE into it. read_test_mode()
    # treats blank as False, so the workbook is usable even before setup-buttons.
    if main_ws[MAIN_CELLS["Status"]].value is None:
        main_ws[MAIN_CELLS["Status"]] = "Idle. Click a button to refresh."

    wb.save(path)
    log.info(f"Workbook ready: {path.resolve()}")
    return path


def _ensure_help_sheet_versions(ws: Worksheet, *, fresh: bool) -> None:
    """Make sure every entry in VERSION_HISTORY has a row on the Help sheet.

    Append-only: if a version is already mentioned anywhere in column A
    (e.g. the user manually rewrote its row), we leave it alone. New
    version entries get appended below the sheet's last non-empty row.

    On fresh creation (``fresh=True``) we also set a sensible column width
    for the summary so it's readable without wrap. Existing workbooks keep
    whatever widths the user has chosen.
    """
    # Determine which versions are already on the sheet (case-insensitive
    # match anywhere in column A).
    mentioned: set[str] = set()
    last_used_row = 0
    max_row = ws.max_row if ws.max_row else 0
    for r in range(1, max_row + 1):
        v = ws.cell(row=r, column=1).value
        if v is not None:
            last_used_row = r
            if isinstance(v, str):
                token = v.strip().lower()
                for ver, _date, _summary in VERSION_HISTORY:
                    if token == ver.lower():
                        mentioned.add(ver)
    # Append missing entries one row past the last used row, with a blank
    # spacer between user content and our changelog if appropriate.
    next_row = last_used_row + 1 if last_used_row else 1
    if mentioned:
        # Some history exists already; append right after the last used row.
        pass
    elif last_used_row > 0:
        # User has prior content (e.g. their credit line) but no version
        # entries yet — leave a blank spacer for visual separation.
        next_row = last_used_row + 2

    appended = 0
    for ver, date, summary in VERSION_HISTORY:
        if ver in mentioned:
            continue
        ws.cell(row=next_row, column=1, value=ver).font = Font(bold=True)
        ws.cell(row=next_row, column=2, value=date)
        c = ws.cell(row=next_row, column=3, value=summary)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        next_row += 1
        appended += 1

    if fresh and appended > 0:
        # Sensible widths only on a brand-new sheet; respect user widths after.
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 120
    if appended:
        log.info(f"  Help sheet: appended {appended} version entr{'y' if appended == 1 else 'ies'}")


# ---------------------------------------------------------------------------
# Job 1: rebuild_inventory — refresh Market structural data
# ---------------------------------------------------------------------------
#
# Fetches the universe of stocks (constituents + watchlist memberships +
# company info) and overwrites the Market sheet in place. Quote columns are
# left blank for get_quotes to fill.

def _write_market_structural_row(
    market_ws, row: int, c: pd.Series, info: dict,
    wl_membership: dict, portfolio: set[str], cols: dict[str, int],
) -> None:
    """Write the structural (non-quote) columns for one Market data row.

    Shared by the full-rebuild path (write to fresh rows 2..N+1) and the
    test-mode incremental-update path (write to existing rows discovered by
    Symbol lookup). Quote / Last update / Last error columns are intentionally
    left alone — those belong to get_quotes.
    """
    sym = c["Symbol"]

    # Watchlist membership: try sym + suffix-stripped root
    sym_root = re.sub(r"\.[A-Z]{1,3}$", "", sym)
    labels: list[str] = []
    for key in {sym, sym_root}:
        labels.extend(wl_membership.get(key, []))
    seen: set[str] = set()
    labels_dedup = [x for x in labels if not (x in seen or seen.add(x))]

    first_index = c["Indexes"].split(",")[0].strip()
    ccy = info.get("currency") or INDEX_DEFAULT_CCY.get(first_index, "")

    sym_cell = market_ws.cell(row=row, column=cols["Symbol"], value=sym)
    sym_cell.hyperlink = yahoo_quote_url(sym)
    sym_cell.style = "Hyperlink"
    market_ws.cell(row=row, column=cols["Name"],        value=info.get("longName") or c["Name"])
    market_ws.cell(row=row, column=cols["Owned?"],      value=_owned_for(sym, portfolio))
    market_ws.cell(row=row, column=cols["Indexes"],     value=c["Indexes"])
    market_ws.cell(row=row, column=cols["Sector"],      value=info.get("sector") or "")
    market_ws.cell(row=row, column=cols["Industry"],    value=info.get("industry") or "")
    market_ws.cell(row=row, column=cols["Watchlists"],  value=", ".join(labels_dedup))
    market_ws.cell(row=row, column=cols["Currency"],    value=ccy)
    desc_cell = market_ws.cell(row=row, column=cols["Description"], value=info.get("description") or "")
    # Force wrap text OFF on the Description cell — Excel sometimes
    # auto-wraps long strings and the resulting tall rows are jarring.
    desc_cell.alignment = Alignment(wrap_text=False)
    # Pre-format the quote/PE cells so when get_quotes writes a value
    # it'll render with thousand separators and 2 decimals out of the box.
    for fmt_col in ("Today (EUR)", "1D ago (EUR)", "1W ago (EUR)", "1M ago (EUR)",
                    "6M ago (EUR)", "1Y ago (EUR)", "5Y ago (EUR)",
                    "P/E (TTM)", "Forward P/E"):
        market_ws.cell(row=row, column=cols[fmt_col]).number_format = COMMA_STYLE
    # Pre-format the % cells with the percent style.
    for _past_label, pct_col in _pct_column_pairs():
        market_ws.cell(row=row, column=cols[pct_col]).number_format = PERCENT_STYLE


def rebuild_inventory(
    workbook_path: Path,
    indexes: Optional[list[str]] = None,
    info_delay: float = 0.25,
    status: Optional[Callable[[str], None]] = None,
    test_mode: Optional[bool] = None,
) -> int:
    """Rebuild the Market sheet from scratch. Preserves Main (portfolio + cells)
    other than the metadata cells this job owns.

    ``indexes``: subset of INDEX_WIKI keys; defaults to all.
    ``test_mode``: True / False / None. When None, reads Main!TestMode cell.
                   When True, restricts to BEL20 regardless of ``indexes``.
    ``status``: optional one-arg callback for live progress text (Phase 8
                wires this to Main!Status via xlwings; CLI mode passes None
                and progress goes to the log instead).
    """
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        log.info(f"Workbook {workbook_path} does not exist — running init-workbook first")
        init_workbook(workbook_path)

    def _say(msg: str) -> None:
        log.info(msg)
        if status is not None:
            try:
                status(msg)
            except Exception as e:
                log.debug(f"status callback raised: {e}")

    t0 = time.time()
    _say("Reading portfolio from Main")
    wb = _load_xl(workbook_path)
    if test_mode is None:
        test_mode = read_test_mode(wb["Main"])
    if test_mode:
        indexes = ["BEL20"]
        _say("Test mode ON — BEL20 only (first 5 stocks, no ETFs, no watchlists)")
    elif indexes is None:
        indexes = list(INDEX_WIKI.keys())
    _say(f"rebuild_inventory: starting ({', '.join(indexes)})")

    portfolio = read_portfolio_symbols(wb["Main"])
    portfolio_entries = read_portfolio_entries(wb["Main"])
    _say(f"  Portfolio entries: {len(portfolio)}")

    _say("Fetching index constituents")
    parts = [get_index_constituents(idx) for idx in indexes]
    if not test_mode:
        parts.append(get_etfs())
    constituents = aggregate_constituents(parts)
    if test_mode:
        constituents = constituents.head(TEST_MODE_TICKER_LIMIT)
    _say(f"  Total unique tickers: {len(constituents)}")

    # Auto-adopt unresolved portfolio entries (look up on Yahoo, append as
    # synthetic "Portfolio"-indexed constituents). Unresolvable symbols get
    # an error string in Main column C of their own row. Skipped in test
    # mode since constituents is intentionally trimmed.
    if not test_mode:
        main_ws_local = wb["Main"]
        def _write_pf_err(row_idx: int, msg: Optional[str]) -> None:
            main_ws_local.cell(row=row_idx, column=PORTFOLIO_ERROR_COL, value=msg)
        constituents = resolve_or_adopt_portfolio(
            portfolio_entries, constituents, _write_pf_err,
        )

    if test_mode:
        wl_membership: dict[str, list[str]] = {}
    else:
        _say("Fetching watchlists")
        wl_membership = fetch_all_watchlists()

    _say(f"Fetching .info for {len(constituents)} tickers — this is the slow part")

    def info_progress(done: int, total: int, sym: str) -> None:
        # Throttle status updates: report every 25 tickers + at boundaries.
        if status is not None and (done == 1 or done == total or done % 25 == 0):
            status(f"info {done}/{total} — {sym}")

    # Reset stale STOP flag at the start of each CLI run; cell-based STOP
    # works via the workbook checkbox same as the button path.
    main_for_stop = wb["Main"]
    if main_for_stop[MAIN_CELLS["StopRequested"]].value is not None:
        main_for_stop[MAIN_CELLS["StopRequested"]] = "FALSE"

    # Speedup F: use the SQLite info cache. 7-day TTL keeps slow-changing
    # fields warm; only true misses pay the .info round-trip.
    cache_path = _info_cache_path_for(workbook_path)
    info_map, cache_hits = fetch_all_info_with_cache(
        list(constituents["Symbol"]),
        cache_path=cache_path,
        progress=info_progress,
        should_stop=lambda: read_stop_requested(wb["Main"]),
    )
    if cache_hits:
        _say(f"  .info cache hits: {cache_hits}/{len(constituents)} "
             f"(fresh fetches: {len(constituents) - cache_hits})")
    if len(info_map) < len(constituents):
        _say(f"STOPPED at {len(info_map)}/{len(constituents)} tickers — Market not modified.")
        return 0

    market_ws = wb["Market"]
    cols = {name: i + 1 for i, name in enumerate(MARKET_COLUMNS)}

    # Re-assert the header row so a SCHEMA_VERSION bump (which may add or
    # rename columns — v04 added Industry at col 6) propagates to existing
    # workbooks. Without this, get_quotes refuses to run because the layout
    # validator sees mismatched headers. overwrite=False preserves any
    # user-set bold/fill/widths on already-correct columns.
    _layout_market_sheet(market_ws, overwrite=False)

    if test_mode:
        # Preserve Market — only update structural columns of rows that
        # already match the test constituents by Symbol. Skip any test row
        # that isn't already in Market (most likely Market hasn't had a
        # full rebuild yet; user will see the warning). This keeps the
        # ~983 rows of stale quote data intact so the post-loop ranking
        # computation has something to rank.
        _say("Test mode: updating matching Market rows (no destructive wipe)")
        existing_by_sym = {}
        for r in range(2, market_ws.max_row + 1):
            v = market_ws.cell(row=r, column=cols["Symbol"]).value
            if v:
                existing_by_sym[str(v).strip().upper()] = r
        updated, skipped = 0, 0
        for _, c in constituents.iterrows():
            sym = c["Symbol"]
            row = existing_by_sym.get(sym.upper())
            if row is None:
                skipped += 1
                continue
            _write_market_structural_row(
                market_ws, row, c, info_map.get(sym, {}),
                wl_membership, portfolio, cols,
            )
            updated += 1
        if skipped and not existing_by_sym:
            _say("  warning: Market is empty — run a full rebuild first to seed it")
        else:
            _say(f"  test mode: updated {updated} existing rows, skipped {skipped} new")
    else:
        _say("Writing Market sheet")
        # Clear existing data rows (preserve row 1 headers).
        if market_ws.max_row > 1:
            market_ws.delete_rows(2, market_ws.max_row - 1)
        for r_offset, (_, c) in enumerate(constituents.iterrows(), 0):
            row = 2 + r_offset
            sym = c["Symbol"]
            _write_market_structural_row(
                market_ws, row, c, info_map.get(sym, {}),
                wl_membership, portfolio, cols,
            )
        # Quote values and Last update/error are left blank for get_quotes.

    # Re-apply AutoFilter to span the current data extent. In test mode we
    # preserved the existing rows; in full-rebuild mode we wrote N+1 rows.
    last_col_letter = get_column_letter(len(MARKET_COLUMNS))
    last_data_row = market_ws.max_row if test_mode else (1 + len(constituents))
    market_ws.auto_filter.ref = f"A1:{last_col_letter}{last_data_row}"

    # Update Main metadata cells — but don't pretend Market shrank to 5 rows
    # in test mode (it still has its full count, we just updated 5 of them).
    main_ws = wb["Main"]
    now_iso = dt.datetime.now(dt.UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    main_ws[MAIN_CELLS["LastRebuildAt"]] = now_iso
    if not test_mode:
        main_ws[MAIN_CELLS["MarketRowCount"]] = len(constituents)

    wb.save(workbook_path)
    duration = time.time() - t0
    if test_mode:
        _say(f"rebuild_inventory: done in {duration:.0f}s (test mode).")
    else:
        _say(f"rebuild_inventory: done in {duration:.0f}s. {len(constituents)} rows.")
    return 0


# ---------------------------------------------------------------------------
# xlwings entry points — called from VBA buttons in stocks.xlsm
# ---------------------------------------------------------------------------
#
# These functions:
#   * Use xlwings to read state from / write cells to the OPEN workbook
#     (openpyxl can't write to a file Excel has open).
#   * Stream "Status:" cell updates so the user sees live progress.
#   * Share the network fetchers (get_index_constituents, fetch_all_watchlists,
#     fetch_all_info) with the CLI rebuild_inventory / get_quotes paths.
#
# Wiring: VBA module stocks_picker.bas (see vba/) calls
#   RunPython "import stocks_report; stocks_report.button_rebuild_inventory()"
# from each macro. The user imports stocks_picker.bas once via Alt+F11, then
# assigns macros to button shapes via right-click → Assign Macro.

def _xw_ensure_market_headers(market_xw, status: Optional[Callable[[str], None]] = None) -> bool:
    """Re-assert Market!row 1 against MARKET_COLUMNS via xlwings.

    Sole purpose: propagate a SCHEMA_VERSION header change (e.g. v04
    added Industry at col 6) to an existing workbook. Only writes if the
    current header row doesn't already match — preserves user-set
    bold/fill on already-correct columns.

    Returns True if headers were rewritten, False if already correct.
    """
    last_col_letter = get_column_letter(len(MARKET_COLUMNS))
    existing = market_xw.range(f"A1:{last_col_letter}1").value
    if not isinstance(existing, list):
        existing = [existing]
    while len(existing) < len(MARKET_COLUMNS):
        existing.append(None)
    if all(existing[i] == name for i, name in enumerate(MARKET_COLUMNS)):
        return False
    if status is not None:
        status("Refreshing Market header row (schema may have changed)")
    market_xw.range("A1").value = [list(MARKET_COLUMNS)]
    return True


def _xw_reapply_market_autofilter(market_xw, wb_xw, last_col_letter: str, last_row: int) -> bool:
    """Re-apply Market's AutoFilter spanning the new data extent.

    Excel COM's `Range.AutoFilter()` (no args) toggles the filter, but is
    finicky about activation state — it raises "AutoFilter method of Range
    class failed" on a non-active sheet. We work around by activating
    Market briefly inside ScreenUpdating=False, restoring the prior view
    in finally.

    Returns True on success, False if the re-apply failed (in which case
    we ensured AutoFilterMode is at least cleanly off so the workbook
    doesn't show a half-broken filter). Never raises.
    """
    try:
        if market_xw.api.AutoFilterMode:
            market_xw.api.AutoFilterMode = False
        app = wb_xw.app
        prev_screen = app.api.ScreenUpdating
        prev_sheet = wb_xw.api.ActiveSheet
        app.api.ScreenUpdating = False
        try:
            market_xw.api.Activate()
            market_xw.api.Range(f"A1:{last_col_letter}{last_row}").AutoFilter()
        finally:
            try:
                prev_sheet.Activate()
            except Exception:
                pass
            app.api.ScreenUpdating = prev_screen
        return True
    except Exception as af_err:
        log.warning(
            f"Couldn't re-apply Market AutoFilter via xlwings: {af_err}. "
            "Rebuild data is intact; re-enable via Data → Filter if you want it."
        )
        try:
            market_xw.api.AutoFilterMode = False
        except Exception:
            pass
        return False


def _xw_last_visible_sheet(wb_xw):
    """Return the last VISIBLE sheet in the xlwings workbook.

    xlwings's `sheets.add(after=...)` translates to Excel COM's
    `Worksheet.Move(After=...)`, which **fails** when the After-target is
    hidden or very-hidden (the case for our `xlwings.conf` config sheet at
    the end of the workbook). Picking the last visible sheet sidesteps that.
    """
    for s in reversed(list(wb_xw.sheets)):
        try:
            if s.api.Visible == -1:  # xlSheetVisible (-1 = visible, 0 hidden, 2 very-hidden)
                return s
        except Exception:
            pass
    return wb_xw.sheets[0]  # fallback to first sheet — should never get here


def _xw_status(main_sheet) -> Callable[[str], None]:
    """Status writer that updates Main!Status live. Truncated to fit one cell."""
    addr = MAIN_CELLS["Status"]
    def _set(msg: str) -> None:
        main_sheet.range(addr).value = msg[:200]
    return _set


def _handle_button_exception(wb, job_name: str, exc: BaseException, traceback_text: str) -> None:
    """Surface a button-triggered exception in three places:

      * stocks_errors.log next to the workbook — full traceback, appended.
      * an "Errors" sheet in the workbook — one row per error (timestamp,
        job, exception, message, traceback).
      * the Main!Status cell — concise summary + hint.

    Never re-raises (caller already caught the exception); this swap of
    the xlwings/VBA truncated MessageBox is the whole point of the wrapper.
    """
    exc_type = type(exc).__name__
    message = str(exc)
    timestamp = dt.datetime.now(dt.UTC).strftime("%Y-%m-%dT%H:%M:%SZ")

    # 1. Log file next to workbook
    try:
        log_path = Path(wb.fullname).parent / ERROR_LOG_FILENAME
        _append_error_log(log_path, exc_type, message, traceback_text)
    except Exception:
        pass

    # 2. Errors sheet (create-if-missing, append a row)
    try:
        if ERRORS_SHEET_NAME not in [s.name for s in wb.sheets]:
            err = wb.sheets.add(ERRORS_SHEET_NAME, after=_xw_last_visible_sheet(wb))
            err.range("A1").value = [[
                "Timestamp (UTC)", "Job", "Exception", "Message", "Traceback",
            ]]
            err.range("A1:E1").api.Font.Bold = True
        err = wb.sheets[ERRORS_SHEET_NAME]
        # Find first empty row in column A
        col_a = err.range("A:A")
        last_used = col_a.api.SpecialCells(11).Row  # xlCellTypeLastCell
        target_row = max(last_used + 1, 2)
        err.range((target_row, 1)).value = [[
            timestamp, job_name, exc_type, message[:500], traceback_text[:8000],
        ]]
        # AutoFilter on the current data extent — keeps the sheet sortable
        # without the user having to enable filters manually after the first
        # error lands.
        try:
            if err.api.AutoFilterMode:
                err.api.AutoFilterMode = False
            err.range(f"A1:E{target_row}").api.AutoFilter()
        except Exception:
            pass
    except Exception:
        pass

    # 3. Status cell — short summary + actionable hint
    try:
        wb.sheets["Main"].range(MAIN_CELLS["Status"]).value = (
            f"ERROR ({timestamp[11:19]} UTC) — {exc_type}: {message[:120]} — "
            f"see {ERROR_LOG_FILENAME} and the Errors sheet for the full traceback."
        )[:240]
    except Exception:
        pass


def _xw_read_portfolio(main_sheet) -> set[str]:
    out: set[str] = set()
    for r in range(PORTFOLIO_FIRST_ROW, PORTFOLIO_LAST_ROW + 1):
        v = main_sheet.range((r, 1)).value
        if v is None:
            continue
        sym = str(v).strip().upper()
        if sym:
            out.add(sym)
    return out


def _xw_read_test_mode(main_sheet) -> bool:
    raw = main_sheet.range(MAIN_CELLS["TestMode"]).value
    if raw is None:
        return False
    if isinstance(raw, bool):
        return raw
    return str(raw).strip().upper() in _TEST_MODE_TRUTHY


def _xw_read_stop_requested(main_sheet) -> bool:
    """Read Main!StopRequested via xlwings. Polled during the .info loop
    every N tickers — when TRUE the loop breaks out and the job exits
    cleanly (no Market write for rebuild, partial-write for get_quotes)."""
    try:
        raw = main_sheet.range(MAIN_CELLS["StopRequested"]).value
    except Exception:
        return False
    if raw is None:
        return False
    if isinstance(raw, bool):
        return raw
    return str(raw).strip().upper() in _TEST_MODE_TRUTHY


def _xw_clear_stop_requested(main_sheet) -> None:
    """Reset Main!StopRequested to FALSE at the start of each job."""
    try:
        main_sheet.range(MAIN_CELLS["StopRequested"]).value = "FALSE"
    except Exception:
        pass


def button_rebuild_inventory() -> None:
    """Entry point for the Excel "Rebuild Inventory" button."""
    import traceback as _tb
    import xlwings as xw
    wb = xw.Book.caller()
    main = wb.sheets["Main"]
    market = wb.sheets["Market"]
    status = _xw_status(main)

    try:
        # Reset Stop flag at the start of each run so a stale TRUE from a
        # previous run doesn't immediately halt this one.
        _xw_clear_stop_requested(main)

        test_mode = _xw_read_test_mode(main)
        indexes = ["BEL20"] if test_mode else list(INDEX_WIKI.keys())
        status(f"rebuild: starting ({'TEST: BEL20 head' if test_mode else 'full'})")

        portfolio = _xw_read_portfolio(main)
        portfolio_entries = _xw_read_portfolio_entries(main)
        status(f"Portfolio entries: {len(portfolio)}")

        status("Fetching index constituents…")
        parts = [get_index_constituents(idx) for idx in indexes]
        if not test_mode:
            parts.append(get_etfs())
        constituents = aggregate_constituents(parts)
        if test_mode:
            constituents = constituents.head(TEST_MODE_TICKER_LIMIT)
        status(f"Total unique tickers: {len(constituents)}")

        # Auto-adopt unresolved portfolio entries. Errors land in Main col C
        # of the offending row. Skipped in test mode (constituents trimmed).
        if not test_mode:
            def _write_pf_err_xw(row_idx: int, msg: Optional[str]) -> None:
                main.range((row_idx, PORTFOLIO_ERROR_COL)).value = msg
            constituents = resolve_or_adopt_portfolio(
                portfolio_entries, constituents, _write_pf_err_xw,
            )

        if test_mode:
            wl_membership: dict[str, list[str]] = {}
        else:
            status("Fetching watchlists…")
            wl_membership = fetch_all_watchlists()

        status(f"Fetching .info for {len(constituents)} tickers (slow step)")

        def info_progress(done: int, total: int, sym: str) -> None:
            if done == 1 or done == total or done % 10 == 0:
                status(f"info {done}/{total} — {sym}")

        # Speedup F: SQLite cache short-circuits repeat .info fetches.
        wb_path = Path(wb.fullname) if wb.fullname else None
        cache_path = _info_cache_path_for(wb_path)
        info_map, cache_hits = fetch_all_info_with_cache(
            list(constituents["Symbol"]),
            cache_path=cache_path,
            progress=info_progress,
            should_stop=lambda: _xw_read_stop_requested(main),
        )
        if cache_hits:
            status(f"  .info cache hits: {cache_hits}/{len(constituents)} "
                   f"(fresh: {len(constituents) - cache_hits})")
        # If the user clicked STOP mid-.info, bail before writing Market.
        # Partial structural data isn't useful — leave Market as it was.
        if len(info_map) < len(constituents):
            status(f"STOPPED at {len(info_map)}/{len(constituents)} tickers — Market not modified.")
            return

        status("Building rows…")
        rows = []
        for _, c in constituents.iterrows():
            sym = c["Symbol"]
            info = info_map.get(sym, {})
            sym_root = re.sub(r"\.[A-Z]{1,3}$", "", sym)
            labels: list[str] = []
            for key in {sym, sym_root}:
                labels.extend(wl_membership.get(key, []))
            seen = set()
            labels_dedup = [x for x in labels if not (x in seen or seen.add(x))]
            first_index = c["Indexes"].split(",")[0].strip()
            ccy = info.get("currency") or INDEX_DEFAULT_CCY.get(first_index, "")
            row = [None] * len(MARKET_COLUMNS)
            row[MARKET_COLUMNS.index("Symbol")]      = sym
            row[MARKET_COLUMNS.index("Name")]        = info.get("longName") or c["Name"]
            row[MARKET_COLUMNS.index("Owned?")]      = _owned_for(sym, portfolio)
            row[MARKET_COLUMNS.index("Indexes")]     = c["Indexes"]
            row[MARKET_COLUMNS.index("Sector")]      = info.get("sector") or ""
            row[MARKET_COLUMNS.index("Industry")]    = info.get("industry") or ""
            row[MARKET_COLUMNS.index("Watchlists")]  = ", ".join(labels_dedup)
            row[MARKET_COLUMNS.index("Currency")]    = ccy
            row[MARKET_COLUMNS.index("Description")] = info.get("description") or ""
            rows.append(row)

        sym_col_idx = MARKET_COLUMNS.index("Symbol") + 1

        # Re-assert Market!row 1 so SCHEMA_VERSION bumps (e.g. v04 Industry)
        # propagate to existing workbooks — same fix as the CLI path.
        _xw_ensure_market_headers(market, status)

        if test_mode:
            # Preserve Market — only update structural columns of rows that
            # already match by Symbol. Keeps the ranking computation (which
            # reads ALL Market rows) populated with real stale quote data
            # instead of seeing just 5 freshly-emptied rows.
            status(f"Test mode: updating {len(rows)} matching rows in place (no wipe)")
            existing_last_row = market.used_range.last_cell.row
            existing_by_sym: dict[str, int] = {}
            if existing_last_row >= 2:
                col_letter = get_column_letter(sym_col_idx)
                syms_in_market = market.range(f"{col_letter}2:{col_letter}{existing_last_row}").value
                if not isinstance(syms_in_market, list):
                    syms_in_market = [syms_in_market]
                for i, v in enumerate(syms_in_market):
                    if v:
                        existing_by_sym[str(v).strip().upper()] = 2 + i
            updated, skipped = 0, 0
            for row_data in rows:
                sym = row_data[sym_col_idx - 1]
                target = existing_by_sym.get(str(sym).upper()) if sym else None
                if target is None:
                    skipped += 1
                    continue
                # Write only the structural columns we care about — leave
                # quote / Last update / Last error cells untouched.
                for col_name in ("Symbol", "Name", "Owned?", "Indexes", "Sector",
                                  "Watchlists", "Currency", "Description"):
                    col_i = MARKET_COLUMNS.index(col_name) + 1
                    market.range((target, col_i)).value = row_data[col_i - 1]
                # Refresh the Symbol hyperlink (in case cell.clear was ever called)
                cell = market.range((target, sym_col_idx))
                market.api.Hyperlinks.Add(
                    Anchor=cell.api,
                    Address=yahoo_quote_url(sym),
                    TextToDisplay=sym,
                )
                updated += 1
            if skipped and not existing_by_sym:
                status("Test mode warning: Market is empty — run a full rebuild first to seed it")
            else:
                status(f"Test mode: updated {updated}, skipped {skipped} new symbols")
        else:
            status(f"Writing {len(rows)} rows to Market…")
            # Clear existing data rows + their hyperlinks. clear_contents leaves
            # hyperlinks attached as ghost references; clear() wipes both.
            last_row = market.used_range.last_cell.row
            if last_row >= 2:
                market.range(f"A2:{get_column_letter(len(MARKET_COLUMNS))}{last_row}").clear()
            # Bulk write — single COM call
            if rows:
                market.range((2, 1)).value = rows
                # Attach Yahoo hyperlink to each Symbol cell. One COM call per row;
                # ~975 calls take ~5 sec total which is dwarfed by the .info loop.
                for offset, row_data in enumerate(rows):
                    sym = row_data[sym_col_idx - 1]
                    if not sym:
                        continue
                    cell = market.range((2 + offset, sym_col_idx))
                    market.api.Hyperlinks.Add(
                        Anchor=cell.api,
                        Address=yahoo_quote_url(sym),
                        TextToDisplay=sym,
                    )

                # Column-level formats — one COM call per range, applied to all
                # data rows in one shot.
                last_row = 1 + len(rows)
                for fmt_col_name in ("Today (EUR)", "1D ago (EUR)", "1W ago (EUR)",
                                      "1M ago (EUR)", "6M ago (EUR)", "1Y ago (EUR)",
                                      "5Y ago (EUR)", "P/E (TTM)", "Forward P/E"):
                    col_letter = get_column_letter(MARKET_COLUMNS.index(fmt_col_name) + 1)
                    market.range(f"{col_letter}2:{col_letter}{last_row}").api.NumberFormat = COMMA_STYLE
                for _past_label, pct_col in _pct_column_pairs():
                    col_letter = get_column_letter(MARKET_COLUMNS.index(pct_col) + 1)
                    market.range(f"{col_letter}2:{col_letter}{last_row}").api.NumberFormat = PERCENT_STYLE
                desc_letter = get_column_letter(MARKET_COLUMNS.index("Description") + 1)
                market.range(f"{desc_letter}2:{desc_letter}{last_row}").api.WrapText = False

                # Re-apply AutoFilter to span the new data extent. See the
                # helper docstring for the activation-state workaround.
                last_col_letter = get_column_letter(len(MARKET_COLUMNS))
                _xw_reapply_market_autofilter(market, wb, last_col_letter, last_row)

        now_iso = dt.datetime.now(dt.UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
        main.range(MAIN_CELLS["LastRebuildAt"]).value = now_iso
        if not test_mode:
            main.range(MAIN_CELLS["MarketRowCount"]).value = len(rows)
        status(f"rebuild: done. {len(rows) if not test_mode else 'test mode update'} at {now_iso}.")
    except Exception as e:
        _handle_button_exception(wb, "rebuild_inventory", e, _tb.format_exc())
        # Do NOT re-raise — xlwings/VBA would otherwise show its truncated
        # popup. The handler has surfaced the full traceback in 3 places.


def button_get_quotes() -> None:
    """Entry point for the Excel "Get Quotes" button."""
    import traceback as _tb
    import xlwings as xw
    wb = xw.Book.caller()
    main = wb.sheets["Main"]
    market = wb.sheets["Market"]
    status = _xw_status(main)

    try:
        _xw_clear_stop_requested(main)  # reset stale STOP flag

        test_mode = _xw_read_test_mode(main)
        status(f"quotes: starting (test_mode={test_mode})")

        # Validate schema: header row must match MARKET_COLUMNS exactly.
        header_vals = market.range((1, 1), (1, len(MARKET_COLUMNS))).value
        if not isinstance(header_vals, list):
            header_vals = [header_vals]
        layout_err = _check_market_headers(header_vals)
        if layout_err is not None:
            status(f"⚠ {layout_err}")
            return

        # Read current Market: pull Symbol + Currency + Indexes into a list
        last_row = market.used_range.last_cell.row
        if last_row < 2:
            status("Market is empty — run Rebuild Inventory first")
            return

        sym_col = MARKET_COLUMNS.index("Symbol") + 1
        ccy_col = MARKET_COLUMNS.index("Currency") + 1
        idx_col = MARKET_COLUMNS.index("Indexes") + 1

        market_rows: list[tuple[int, str, str, str]] = []
        # Bulk-read all three columns in one shot
        syms = market.range((2, sym_col), (last_row, sym_col)).value
        ccys = market.range((2, ccy_col), (last_row, ccy_col)).value
        idxs = market.range((2, idx_col), (last_row, idx_col)).value
        # When the range is a single cell xlwings returns a scalar, not a list
        if not isinstance(syms, list): syms = [syms]
        if not isinstance(ccys, list): ccys = [ccys]
        if not isinstance(idxs, list): idxs = [idxs]
        for offset, (s, c, i) in enumerate(zip(syms, ccys, idxs)):
            if not s:
                continue
            market_rows.append((2 + offset, str(s).strip(), str(c or ""), str(i or "")))

        portfolio = _xw_read_portfolio(main)
        if test_mode:
            # Refresh the top N Market rows so the post-loop ranking has
            # populated data. Portfolio symbols come first so the user sees
            # their own stocks refresh.
            portfolio_first = []
            rest = []
            portfolio_upper = {p.upper() for p in portfolio}
            for row in market_rows:
                sym_u = row[1].upper()
                root_u = re.sub(r"\.[A-Z]{1,3}$", "", sym_u)
                if sym_u in portfolio_upper or root_u in portfolio_upper:
                    portfolio_first.append(row)
                else:
                    rest.append(row)
            ordered = portfolio_first + rest
            targets = ordered[:TEST_MODE_QUOTE_REFRESH_LIMIT]
            if not targets:
                status("Test mode: Market is empty")
                return
            status(f"Test mode: refreshing top {len(targets)} rows "
                   f"({len(portfolio_first)} from portfolio)")
        else:
            targets = market_rows

        status("Fetching FX rates + history")
        fx = get_fx_rates()
        fx_history = get_fx_history()
        _ensure_currencies_sheet_xlwings(wb, fx_history)

        status(f"Downloading price history for {len(targets)} tickers")
        today = dt.date.today()
        start = today - dt.timedelta(days=int(5.2 * 365))
        end   = today + dt.timedelta(days=1)
        closes = fetch_close_prices([t[1] for t in targets], start, end)

        # Speedup C: filter out targets whose Last update is recent.
        lu_col_xw = MARKET_COLUMNS.index("Last update (UTC)") + 1
        stale_targets: list[tuple[int, str, str, str]] = []
        fresh_count = 0
        for row in targets:
            last_iso = market.range((row[0], lu_col_xw)).value
            if isinstance(last_iso, dt.datetime):
                # xlwings may surface ISO timestamps as native datetimes
                if last_iso.tzinfo is None:
                    last_iso_for_check = last_iso.replace(tzinfo=dt.UTC).isoformat()
                else:
                    last_iso_for_check = last_iso.isoformat()
            else:
                last_iso_for_check = last_iso
            if is_recently_updated(last_iso_for_check, QUOTE_FRESHNESS_HOURS):
                fresh_count += 1
            else:
                stale_targets.append(row)
        if fresh_count:
            status(f"Skipping .info for {fresh_count} rows refreshed within {QUOTE_FRESHNESS_HOURS:g}h")
        status(f"Fetching .info for {len(stale_targets)} tickers (P/E)")

        def info_progress(done: int, total: int, sym: str) -> None:
            if done == 1 or done == total or done % 10 == 0:
                status(f"quotes {done}/{total} — {sym}")

        info_map = fetch_all_info(
            [t[1] for t in stale_targets], delay=0.25, progress=info_progress,
            should_stop=lambda: _xw_read_stop_requested(main),
        )
        info_stopped = len(info_map) < len(stale_targets)
        if info_stopped:
            status(f"STOP detected — only {len(info_map)} of {len(stale_targets)} .info fetches done; "
                   "continuing to write rows we have")

        status(f"Writing quote columns for {len(targets)} rows…")
        now_iso = dt.datetime.now(dt.UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
        cols = {n: MARKET_COLUMNS.index(n) + 1 for n in MARKET_COLUMNS}

        ok = 0
        fail = 0
        for (row_idx, sym, ccy, _indexes_csv) in targets:
            # Stop check between rows (in case the per-row loop is itself slow)
            if _xw_read_stop_requested(main):
                status(f"STOP detected — refreshed {ok} of {len(targets)} rows")
                break
            try:
                series = closes[sym] if sym in closes.columns else pd.Series(dtype=float)
                eur_by_label = {}
                for label, delta in LOOKBACKS.items():
                    target_d = today - delta
                    native = price_at_or_before(series, target_d)
                    eur = to_eur(native, ccy, fx)
                    eur_by_label[label] = eur
                    market.range((row_idx, cols[f"{label} (EUR)"])).value = eur
                today_eur = eur_by_label.get("Today")
                for past_label, pct_col in _pct_column_pairs():
                    pct = _pct_change(today_eur, eur_by_label.get(past_label))
                    rng = market.range((row_idx, cols[pct_col]))
                    rng.value = pct
                    rng.api.NumberFormat = PERCENT_STYLE
                if sym in info_map:
                    info = info_map[sym]
                    market.range((row_idx, cols["P/E (TTM)"])).value   = info.get("trailingPE")
                    market.range((row_idx, cols["Forward P/E"])).value = info.get("forwardPE")
                    market.range((row_idx, cols["Last update (UTC)"])).value = now_iso
                # else: skipped by freshness filter — keep prior P/E + Last update
                market.range((row_idx, cols["Last error"])).value = None
                ok += 1
            except Exception as e:
                market.range((row_idx, cols["Last error"])).value = f"{type(e).__name__}: {e}"[:300]
                fail += 1

        # Migrate legacy "Monthly movers" → "Monthly winners" if needed; rank
        # from the full Market sheet so test mode (1 refreshed ticker) still
        # shows meaningful winners/losers.
        _migrate_monthly_movers_to_winners_xlwings(wb)
        status("Computing Monthly winners + losers")
        all_records = _read_market_records_xlwings(market)
        winners = _compute_monthly_winners(all_records)
        losers  = _compute_monthly_losers(all_records)
        if len(all_records) < 50:
            status(f"⚠ Market only has {len(all_records)} rows — ranking sparse. "
                   "Run full rebuild-inventory to seed Market.")
        # Adaptive filter: skip the Owned?=Yes pre-filter when no owned
        # losers exist — otherwise the sheet looks empty to the user.
        has_owned_losers = any(_is_owned(m.get("owned")) for m in losers)
        status(f"Winners: {len(winners)}, Losers: {len(losers)} "
               f"(owned: {sum(1 for m in losers if _is_owned(m.get('owned')))})")
        _ensure_ranking_sheet_xlwings(wb, MONTHLY_WINNERS_SHEET_NAME, winners)
        _ensure_ranking_sheet_xlwings(
            wb, MONTHLY_LOSERS_SHEET_NAME, losers,
            filter_owned_yes=has_owned_losers,
        )

        main.range(MAIN_CELLS["LastQuotesAt"]).value = now_iso
        if not test_mode:
            if fx.get("USD") is not None and not pd.isna(fx["USD"]):
                main.range(MAIN_CELLS["EurUsd"]).value = float(fx["USD"])
            if fx.get("JPY") is not None and not pd.isna(fx["JPY"]):
                main.range(MAIN_CELLS["EurJpy"]).value = float(fx["JPY"])
            if fx.get("GBP") is not None and not pd.isna(fx["GBP"]):
                main.range(MAIN_CELLS["EurGbp"]).value = float(fx["GBP"])

        status(f"quotes: done. {ok} ok, {fail} failed at {now_iso}.")
    except Exception as e:
        _handle_button_exception(wb, "get_quotes", e, _tb.format_exc())


# ---------------------------------------------------------------------------
# Currencies sheet — populated by get_quotes after FX history is fetched
# ---------------------------------------------------------------------------

CURRENCY_HEADERS = ["Pair"] + list(LOOKBACKS.keys())  # "Pair", "Today", "1D ago", ...
FX_RATE_FORMAT = "#,##0.0000"  # FX needs more precision than equities


def _currency_rows(fx_history: dict[str, pd.Series]) -> list[list]:
    """One row per FX pair: ["EUR/USD", rate_today, rate_1d_ago, ...]."""
    today_d = dt.date.today()
    rows: list[list] = []
    for ccy in FX_PAIRS:  # iteration order = display order in the sheet
        hist = fx_history.get(ccy, pd.Series(dtype=float))
        row: list = [f"EUR/{ccy}"]
        for _label, delta in LOOKBACKS.items():
            row.append(price_at_or_before(hist, today_d - delta))
        rows.append(row)
    return rows


def _ensure_currencies_sheet_openpyxl(wb, fx_history: dict[str, pd.Series]) -> None:
    """Write/refresh the Currencies sheet using openpyxl (CLI path)."""
    if CURRENCIES_SHEET_NAME not in wb.sheetnames:
        ws = wb.create_sheet(CURRENCIES_SHEET_NAME)
        is_new = True
    else:
        ws = wb[CURRENCIES_SHEET_NAME]
        is_new = False

    # Header — re-assert text (column-name contract) but apply styling/widths
    # only on first creation so user cosmetics survive.
    for col_idx, name in enumerate(CURRENCY_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        if is_new or cell.value != name:
            cell.value = name
        if is_new:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
    if is_new:
        ws.freeze_panes = "B2"
        ws.column_dimensions["A"].width = 12
        for col_idx in range(2, len(CURRENCY_HEADERS) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 13

    # Data rows
    rows = _currency_rows(fx_history)
    for r_offset, row in enumerate(rows, 0):
        for c_offset, val in enumerate(row, 0):
            cell = ws.cell(row=2 + r_offset, column=1 + c_offset, value=val)
            if c_offset > 0:
                cell.number_format = FX_RATE_FORMAT
    # Trim any data rows from a prior run that exceed the current pair count.
    for r in range(2 + len(rows), ws.max_row + 1):
        for c in range(1, len(CURRENCY_HEADERS) + 1):
            ws.cell(row=r, column=c).value = None

    # AutoFilter spanning the current data extent — keeps the user's
    # sort/filter affordance consistent with Market / Monthly winners / losers.
    last_col = get_column_letter(len(CURRENCY_HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{1 + len(rows)}"


def _ensure_currencies_sheet_xlwings(wb_xw, fx_history: dict[str, pd.Series]) -> None:
    """Write/refresh the Currencies sheet using xlwings (button path)."""
    sheet_names = [s.name for s in wb_xw.sheets]
    if CURRENCIES_SHEET_NAME not in sheet_names:
        cur = wb_xw.sheets.add(CURRENCIES_SHEET_NAME, after=_xw_last_visible_sheet(wb_xw))
        is_new = True
    else:
        cur = wb_xw.sheets[CURRENCIES_SHEET_NAME]
        is_new = False
    # Header
    cur.range("A1").value = [CURRENCY_HEADERS]
    if is_new:
        cur.range(f"A1:{get_column_letter(len(CURRENCY_HEADERS))}1").api.Font.Bold = True
        cur.api.Range("A2").Select()  # ignored; intent is to freeze
        cur.api.Application.ActiveWindow.FreezePanes = False
        # Use COM Window object for freeze pane
    # Data
    rows = _currency_rows(fx_history)
    cur.range((2, 1)).value = rows
    # Number format on numeric columns
    last_col_letter = get_column_letter(len(CURRENCY_HEADERS))
    cur.range(f"B2:{last_col_letter}{1 + len(rows)}").api.NumberFormat = FX_RATE_FORMAT
    if is_new:
        cur.range("A:A").api.ColumnWidth = 12
        for col_idx in range(2, len(CURRENCY_HEADERS) + 1):
            cur.range(f"{get_column_letter(col_idx)}:{get_column_letter(col_idx)}").api.ColumnWidth = 13

    # AutoFilter on the current data extent (toggled cleanly so the user
    # always sees a fresh filter row, even if Excel had one stale from a
    # prior run with a different row count).
    last_col_letter = get_column_letter(len(CURRENCY_HEADERS))
    try:
        if cur.api.AutoFilterMode:
            cur.api.AutoFilterMode = False
        cur.range(f"A1:{last_col_letter}{1 + len(rows)}").api.AutoFilter()
    except Exception as e:
        log.debug(f"Currencies AutoFilter apply failed (non-fatal): {e}")


# ---------------------------------------------------------------------------
# Monthly winners + Monthly losers sheets
# ---------------------------------------------------------------------------
#
# Same 9-column layout, same input record schema, different filter+sort rules.
# Computed from a snapshot of the FULL Market sheet on every get_quotes
# (not just the rows refreshed this run), so test-mode get_quotes (which
# only refreshes 1 ticker) still produces a meaningful ranking from the
# rest of the existing data.

RANKING_HEADERS = ["Symbol", "Name", "Owned?", "Indexes", "Sector",
                   "Today (EUR)", "1D %", "1W %", "1M %"]
MONTHLY_RANKING_TOP_N = 50
OWNED_COL_INDEX_1BASED = RANKING_HEADERS.index("Owned?") + 1  # 3


PORTFOLIO_ADOPTED_INDEX_LABEL = "Portfolio"
PORTFOLIO_ERROR_COL = 3  # Main column C: portfolio error messages


def _portfolio_symbol_resolves(sym: str, market_upper: set[str],
                               market_roots: set[str]) -> bool:
    """True if a portfolio entry matches an existing constituent (with or
    without exchange suffix)."""
    s = sym.upper()
    if s in market_upper or s in market_roots:
        return True
    return any(m.startswith(s + ".") for m in market_upper)


def _yahoo_lookup_for_adoption(sym: str) -> Optional[dict]:
    """Best-effort Yahoo lookup. Returns a dict with at least Name if Yahoo
    knows the ticker, else None.

    A ticker is considered resolved if .info exposes a price-like field
    (regularMarketPrice, currentPrice, previousClose) or a symbol field —
    enough to confirm the ticker is live.
    """
    try:
        info = yf.Ticker(sym).info or {}
    except Exception as e:
        log.debug(f"yahoo lookup '{sym}' raised: {e}")
        return None
    has_price = any(info.get(k) is not None for k in
                    ("regularMarketPrice", "currentPrice", "previousClose"))
    has_symbol = bool(info.get("symbol")) or bool(info.get("shortName")) or bool(info.get("longName"))
    if has_price or has_symbol:
        return {
            "Name": info.get("longName") or info.get("shortName") or sym,
            "currency": info.get("currency") or "",
        }
    return None


def resolve_or_adopt_portfolio(
    portfolio_entries: list[tuple[int, str]],
    constituents: pd.DataFrame,
    write_error: Callable[[int, Optional[str]], None],
) -> pd.DataFrame:
    """Auto-adopt unresolved portfolio entries into the constituents universe.

    For each ``(row_idx, symbol)`` in ``portfolio_entries``:
      * If the symbol already matches a constituent (with or without
        exchange suffix), clear any stale error and move on.
      * Else try ``yf.Ticker(symbol)``. If Yahoo recognises it, append a
        synthetic constituent row with ``Indexes="Portfolio"`` so the
        ticker flows through .info / price / Owned? like any other row,
        and clear the error.
      * Else call ``write_error(row_idx, msg)`` with a one-line reason —
        the rebuild proceeds (no longer raises).

    Returns the (possibly augmented) constituents DataFrame.
    """
    market_upper = {s.upper() for s in constituents["Symbol"]}
    market_roots = {re.sub(r"\.[A-Z]{1,3}$", "", s.upper()) for s in constituents["Symbol"]}
    extra_rows: list[dict] = []
    adopted: list[str] = []
    failed: list[str] = []
    for row_idx, sym in portfolio_entries:
        if _portfolio_symbol_resolves(sym, market_upper, market_roots):
            write_error(row_idx, None)
            continue
        info = _yahoo_lookup_for_adoption(sym)
        if info is not None:
            extra_rows.append({
                "Symbol": sym,
                "Name": info["Name"],
                "Indexes": PORTFOLIO_ADOPTED_INDEX_LABEL,
            })
            # Refresh the lookup sets so two portfolio entries pointing at
            # the same ticker don't both get appended.
            market_upper.add(sym.upper())
            market_roots.add(re.sub(r"\.[A-Z]{1,3}$", "", sym.upper()))
            write_error(row_idx, None)
            adopted.append(sym)
        else:
            write_error(row_idx, f"Symbol not found on Yahoo Finance ({sym})")
            failed.append(sym)
    if adopted:
        log.info(f"  Adopted {len(adopted)} portfolio symbol(s) into universe: {', '.join(adopted)}")
    if failed:
        log.warning(f"  {len(failed)} portfolio symbol(s) unresolvable on Yahoo: "
                    f"{', '.join(failed)} — error written to Main col C")
    if extra_rows:
        return pd.concat([constituents, pd.DataFrame(extra_rows)], ignore_index=True)
    return constituents


def read_portfolio_entries(main_ws: Worksheet) -> list[tuple[int, str]]:
    """Row-aware portfolio reader: ``[(row_idx, SYMBOL), ...]`` for openpyxl.

    Used by the auto-adopt path to write errors back to the specific row
    where the unresolvable symbol lives.
    """
    out: list[tuple[int, str]] = []
    for r in range(PORTFOLIO_FIRST_ROW, PORTFOLIO_LAST_ROW + 1):
        v = main_ws.cell(row=r, column=1).value
        if v is None:
            continue
        sym = str(v).strip().upper()
        if sym:
            out.append((r, sym))
    return out


def _xw_read_portfolio_entries(main_sheet) -> list[tuple[int, str]]:
    """Row-aware portfolio reader for xlwings (live workbook)."""
    out: list[tuple[int, str]] = []
    for r in range(PORTFOLIO_FIRST_ROW, PORTFOLIO_LAST_ROW + 1):
        v = main_sheet.range((r, 1)).value
        if v is None:
            continue
        sym = str(v).strip().upper()
        if sym:
            out.append((r, sym))
    return out


def _is_owned(owned_value) -> bool:
    """Interpret the Market!Owned? cell value as a bool. Tolerates None,
    plain bool, "Yes"/"No" strings, "1"/"0", whitespace."""
    if owned_value is None:
        return False
    if isinstance(owned_value, bool):
        return owned_value
    return str(owned_value).strip().lower() in ("yes", "y", "true", "t", "1")


def _compute_monthly_winners(records: list[dict],
                              top_n: int = MONTHLY_RANKING_TOP_N) -> list[dict]:
    """Top 1M gainers with NO 1D or 1W weakness.

    Filter: pct_1m > 0 AND pct_1d >= 0 AND pct_1w >= 0.
    Sort: pct_1m descending.
    """
    qualified: list[dict] = []
    for r in records:
        d, w, m = r.get("pct_1d"), r.get("pct_1w"), r.get("pct_1m")
        if d is None or w is None or m is None:
            continue
        if m > 0 and d >= 0 and w >= 0:
            qualified.append(r)
    qualified.sort(key=lambda r: r["pct_1m"], reverse=True)
    return qualified[:top_n]


def _compute_monthly_losers(records: list[dict],
                             top_n: int = MONTHLY_RANKING_TOP_N) -> list[dict]:
    """Worst sustained 1M decliners — also negative on 1D and 1W.

    Filter: pct_1m < 0 AND pct_1d <= 0 AND pct_1w <= 0.  (Mirrors the
    winners' "no weakness" logic: we want sustained slides, not bounces.)
    Sort: pct_1m ascending (most negative first).
    """
    qualified: list[dict] = []
    for r in records:
        d, w, m = r.get("pct_1d"), r.get("pct_1w"), r.get("pct_1m")
        if d is None or w is None or m is None:
            continue
        if m < 0 and d <= 0 and w <= 0:
            qualified.append(r)
    qualified.sort(key=lambda r: r["pct_1m"])
    return qualified[:top_n]


def _ranking_row_values(m: dict) -> list:
    """One row of values matching RANKING_HEADERS order."""
    return [
        m["symbol"],
        m.get("name") or "",
        m.get("owned") or "No",
        m.get("indexes") or "",
        m.get("sector") or "",
        m.get("today"),
        m.get("pct_1d"),
        m.get("pct_1w"),
        m.get("pct_1m"),
    ]


def _ensure_ranking_sheet_openpyxl(wb, sheet_name: str, rows: list[dict],
                                    *, filter_owned_yes: bool = False) -> None:
    """Write/refresh a winners-style sheet via openpyxl (CLI path).

    ``filter_owned_yes`` toggles a pre-applied AutoFilter on Owned?=Yes,
    used for the Monthly losers sheet so the user immediately sees decline
    on stocks they own.
    """
    if sheet_name not in wb.sheetnames:
        ws = wb.create_sheet(sheet_name)
        is_new = True
    else:
        ws = wb[sheet_name]
        is_new = False

    for col_idx, name in enumerate(RANKING_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx)
        if is_new or cell.value != name:
            cell.value = name
        if is_new:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="F2F2F2")
    if is_new:
        ws.freeze_panes = "A2"
        widths = {"Symbol": 10, "Name": 28, "Owned?": 8, "Indexes": 18,
                  "Sector": 22, "Today (EUR)": 12,
                  "1D %": 9, "1W %": 9, "1M %": 9}
        for col_idx, name in enumerate(RANKING_HEADERS, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = widths.get(name, 12)

    # Clear prior data rows; write new ones.
    for r in range(2, ws.max_row + 1):
        for c in range(1, len(RANKING_HEADERS) + 1):
            ws.cell(row=r, column=c).value = None
    for r_offset, m in enumerate(rows, 0):
        row = 2 + r_offset
        sym = m["symbol"]
        sym_cell = ws.cell(row=row, column=1, value=sym)
        sym_cell.hyperlink = yahoo_quote_url(sym)
        sym_cell.style = "Hyperlink"
        values = _ranking_row_values(m)
        ws.cell(row=row, column=2, value=values[1])
        ws.cell(row=row, column=3, value=values[2])
        ws.cell(row=row, column=4, value=values[3])
        ws.cell(row=row, column=5, value=values[4])
        c6 = ws.cell(row=row, column=6, value=values[5])
        c6.number_format = COMMA_STYLE
        for pct_col_idx, val in ((7, values[6]), (8, values[7]), (9, values[8])):
            cell = ws.cell(row=row, column=pct_col_idx, value=val)
            cell.number_format = PERCENT_STYLE

    # AutoFilter on the data range, with optional pre-set Owned?=Yes filter.
    last_col_letter = get_column_letter(len(RANKING_HEADERS))
    last_row = 1 + len(rows) if rows else 1
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"
    # Clear any prior filterColumn objects
    ws.auto_filter.filterColumn = []
    if filter_owned_yes and rows:
        from openpyxl.worksheet.filters import FilterColumn, Filters
        fc = FilterColumn(colId=OWNED_COL_INDEX_1BASED - 1)  # 0-based for FilterColumn
        fc.filters = Filters(filter=["Yes"])
        ws.auto_filter.filterColumn.append(fc)


def _ensure_ranking_sheet_xlwings(wb_xw, sheet_name: str, rows: list[dict],
                                   *, filter_owned_yes: bool = False) -> None:
    """Write/refresh a winners-style sheet via xlwings (button path)."""
    sheet_names = [s.name for s in wb_xw.sheets]
    if sheet_name not in sheet_names:
        sh = wb_xw.sheets.add(sheet_name, after=_xw_last_visible_sheet(wb_xw))
        is_new = True
    else:
        sh = wb_xw.sheets[sheet_name]
        is_new = False

    sh.range("A1").value = [RANKING_HEADERS]
    if is_new:
        sh.range(f"A1:{get_column_letter(len(RANKING_HEADERS))}1").api.Font.Bold = True

    last_row_now = sh.used_range.last_cell.row
    if last_row_now >= 2:
        sh.range(f"A2:{get_column_letter(len(RANKING_HEADERS))}{last_row_now}").clear()

    if rows:
        bulk = [_ranking_row_values(m) for m in rows]
        sh.range((2, 1)).value = bulk
        last_row = 1 + len(rows)
        sh.range(f"F2:F{last_row}").api.NumberFormat = COMMA_STYLE
        sh.range(f"G2:I{last_row}").api.NumberFormat = PERCENT_STYLE
        # Hyperlink each Symbol cell
        for offset, m in enumerate(rows):
            cell = sh.range((2 + offset, 1))
            sh.api.Hyperlinks.Add(
                Anchor=cell.api,
                Address=yahoo_quote_url(m["symbol"]),
                TextToDisplay=m["symbol"],
            )

        # AutoFilter (best-effort via the same helper Market uses).
        last_col_letter = get_column_letter(len(RANKING_HEADERS))
        try:
            if sh.api.AutoFilterMode:
                sh.api.AutoFilterMode = False
            app = wb_xw.app
            prev_screen = app.api.ScreenUpdating
            prev_sheet = wb_xw.api.ActiveSheet
            app.api.ScreenUpdating = False
            try:
                sh.api.Activate()
                if filter_owned_yes:
                    sh.api.Range(f"A1:{last_col_letter}{last_row}").AutoFilter(
                        Field=OWNED_COL_INDEX_1BASED, Criteria1="Yes",
                    )
                else:
                    sh.api.Range(f"A1:{last_col_letter}{last_row}").AutoFilter()
            finally:
                try:
                    prev_sheet.Activate()
                except Exception:
                    pass
                app.api.ScreenUpdating = prev_screen
        except Exception as e:
            log.warning(f"Couldn't apply AutoFilter on {sheet_name}: {e}")
            try:
                sh.api.AutoFilterMode = False
            except Exception:
                pass


def _read_market_records_openpyxl(market_ws) -> list[dict]:
    """Snapshot every Market data row into the shape _compute_monthly_*
    expects. Used by get_quotes so the ranking sees the whole market,
    not just the rows we refreshed this run."""
    cols = {n: MARKET_COLUMNS.index(n) + 1 for n in MARKET_COLUMNS}
    out: list[dict] = []
    for r in range(2, market_ws.max_row + 1):
        sym = market_ws.cell(row=r, column=cols["Symbol"]).value
        if not sym:
            continue
        out.append({
            "symbol":  str(sym),
            "name":    market_ws.cell(row=r, column=cols["Name"]).value,
            "owned":   market_ws.cell(row=r, column=cols["Owned?"]).value,
            "indexes": market_ws.cell(row=r, column=cols["Indexes"]).value,
            "sector":  market_ws.cell(row=r, column=cols["Sector"]).value,
            "today":   market_ws.cell(row=r, column=cols["Today (EUR)"]).value,
            "pct_1d":  market_ws.cell(row=r, column=cols["1D %"]).value,
            "pct_1w":  market_ws.cell(row=r, column=cols["1W %"]).value,
            "pct_1m":  market_ws.cell(row=r, column=cols["1M %"]).value,
        })
    return out


def _read_market_records_xlwings(market_xw) -> list[dict]:
    """Same as the openpyxl version but bulk-reads via xlwings COM."""
    cols = {n: MARKET_COLUMNS.index(n) + 1 for n in MARKET_COLUMNS}
    last_row = market_xw.used_range.last_cell.row
    if last_row < 2:
        return []
    last_col = len(MARKET_COLUMNS)
    last_col_letter = get_column_letter(last_col)
    # Bulk read — one COM call instead of N×M cell reads.
    matrix = market_xw.range(f"A2:{last_col_letter}{last_row}").value
    # Single row → xlwings returns a flat list; normalize to list-of-lists.
    if matrix and not isinstance(matrix[0], list):
        matrix = [matrix]
    out: list[dict] = []
    for row_vals in matrix:
        sym = row_vals[cols["Symbol"] - 1]
        if not sym:
            continue
        out.append({
            "symbol":  str(sym),
            "name":    row_vals[cols["Name"] - 1],
            "owned":   row_vals[cols["Owned?"] - 1],
            "indexes": row_vals[cols["Indexes"] - 1],
            "sector":  row_vals[cols["Sector"] - 1],
            "today":   row_vals[cols["Today (EUR)"] - 1],
            "pct_1d":  row_vals[cols["1D %"] - 1],
            "pct_1w":  row_vals[cols["1W %"] - 1],
            "pct_1m":  row_vals[cols["1M %"] - 1],
        })
    return out


def _migrate_monthly_movers_to_winners_openpyxl(wb) -> None:
    """If the workbook still has a legacy 'Monthly movers' sheet, rename it."""
    if MONTHLY_MOVERS_LEGACY_NAME in wb.sheetnames and MONTHLY_WINNERS_SHEET_NAME not in wb.sheetnames:
        wb[MONTHLY_MOVERS_LEGACY_NAME].title = MONTHLY_WINNERS_SHEET_NAME
        log.info(f"  migrated sheet '{MONTHLY_MOVERS_LEGACY_NAME}' → '{MONTHLY_WINNERS_SHEET_NAME}'")


def _migrate_monthly_movers_to_winners_xlwings(wb_xw) -> None:
    """xlwings counterpart of the migration."""
    names = [s.name for s in wb_xw.sheets]
    if MONTHLY_MOVERS_LEGACY_NAME in names and MONTHLY_WINNERS_SHEET_NAME not in names:
        wb_xw.sheets[MONTHLY_MOVERS_LEGACY_NAME].name = MONTHLY_WINNERS_SHEET_NAME
        log.info(f"  migrated sheet '{MONTHLY_MOVERS_LEGACY_NAME}' → '{MONTHLY_WINNERS_SHEET_NAME}'")


# ---------------------------------------------------------------------------
# Job 2: get_quotes — refresh price and P/E columns daily
# ---------------------------------------------------------------------------

def _read_market_symbols(market_ws: Worksheet) -> list[tuple[int, str, str, str]]:
    """Iterate Market data rows. Returns ``[(row_idx, symbol, currency, indexes), ...]``."""
    sym_c = _market_col("Symbol")
    ccy_c = _market_col("Currency")
    idx_c = _market_col("Indexes")
    out: list[tuple[int, str, str, str]] = []
    for r in range(2, market_ws.max_row + 1):
        sym = market_ws.cell(row=r, column=sym_c).value
        if not sym:
            continue
        ccy = market_ws.cell(row=r, column=ccy_c).value or ""
        idxs = market_ws.cell(row=r, column=idx_c).value or ""
        out.append((r, str(sym).strip(), str(ccy), str(idxs)))
    return out


def _pick_test_target(
    market_rows: list[tuple[int, str, str, str]],
    portfolio: set[str],
) -> Optional[tuple[int, str, str, str]]:
    """Test-mode target: top portfolio entry that exists in Market, else first
    BEL20 row, else first Market row."""
    if portfolio:
        # User may have typed "KBC" (root) or "KBC.BR" (full Yahoo) in Main.
        # Market always stores the full Yahoo form. Match either direction.
        for sym in sorted(portfolio):
            for row in market_rows:
                market_sym = row[1].upper()
                market_root = re.sub(r"\.[A-Z]{1,3}$", "", market_sym)
                if market_sym == sym or market_root == sym:
                    return row
    for row in market_rows:
        if "BEL20" in row[3]:
            return row
    return market_rows[0] if market_rows else None


def get_quotes(
    workbook_path: Path,
    test_mode: Optional[bool] = None,
    info_delay: float = 0.25,
    status: Optional[Callable[[str], None]] = None,
) -> int:
    """Refresh quote columns in Market for every Symbol (or one in test mode).

    Sets per-row "Last update (UTC)" on success and "Last error" on failure;
    successful refresh clears any previous error. ``test_mode=None`` reads
    Main!TestMode from the workbook.
    """
    workbook_path = Path(workbook_path)
    if not workbook_path.exists():
        log.error(f"Workbook {workbook_path} does not exist. Run init-workbook + rebuild-inventory first.")
        return 1

    def _say(msg: str) -> None:
        log.info(msg)
        if status is not None:
            try:
                status(msg)
            except Exception as e:
                log.debug(f"status callback raised: {e}")

    t0 = time.time()
    wb = _load_xl(workbook_path)
    main_ws = wb["Main"]
    market_ws = wb["Market"]
    if test_mode is None:
        test_mode = read_test_mode(main_ws)
    _say(f"get_quotes: starting (test_mode={test_mode})")

    layout_err = _validate_market_layout(market_ws)
    if layout_err is not None:
        log.error(layout_err)
        return 1

    portfolio = read_portfolio_symbols(main_ws)
    market_rows = _read_market_symbols(market_ws)

    if not market_rows:
        log.error("Market sheet is empty. Run rebuild-inventory first.")
        return 1

    if test_mode:
        # Refresh the top N Market rows (not just 1). The post-loop ranking
        # computation reads ALL Market rows; with only 1 refreshed it had
        # nothing populated to rank on, producing empty Monthly winners/losers.
        # Bring the user's portfolio symbol(s) to the front of the slice so
        # the user sees their stock refresh first when they're watching.
        portfolio_first = []
        rest = []
        portfolio_upper = {p.upper() for p in portfolio}
        for row in market_rows:
            sym_u = row[1].upper()
            root_u = re.sub(r"\.[A-Z]{1,3}$", "", sym_u)
            if sym_u in portfolio_upper or root_u in portfolio_upper:
                portfolio_first.append(row)
            else:
                rest.append(row)
        ordered = portfolio_first + rest
        targets = ordered[:TEST_MODE_QUOTE_REFRESH_LIMIT]
        _say(f"Test mode: refreshing top {len(targets)} rows "
             f"({len(portfolio_first)} from portfolio)")
    else:
        targets = market_rows

    _say("Fetching FX rates + history")
    fx = get_fx_rates()
    fx_history = get_fx_history()
    _ensure_currencies_sheet_openpyxl(wb, fx_history)

    _say(f"Downloading price history for {len(targets)} tickers")
    today = dt.date.today()
    start = today - dt.timedelta(days=int(5.2 * 365))
    end   = today + dt.timedelta(days=1)
    closes = fetch_close_prices([t[1] for t in targets], start, end)

    # Speedup C: skip .info for any row whose Last update is recent. We
    # still refresh prices for every target (cheap, batched) — only the
    # per-ticker .info call is dropped. P/E + Last update on skipped rows
    # stays at its previous value.
    lu_col = _market_col("Last update (UTC)")
    stale_targets: list[tuple[int, str, str, str]] = []
    fresh_count = 0
    for row in targets:
        last_iso = market_ws.cell(row=row[0], column=lu_col).value
        if is_recently_updated(last_iso, QUOTE_FRESHNESS_HOURS):
            fresh_count += 1
        else:
            stale_targets.append(row)
    if fresh_count:
        _say(f"Skipping .info for {fresh_count} rows refreshed within {QUOTE_FRESHNESS_HOURS:g}h")
    _say(f"Fetching .info for {len(stale_targets)} tickers (P/E)")

    def info_progress(done: int, total: int, sym: str) -> None:
        if status is not None and (done == 1 or done == total or done % 25 == 0):
            status(f"quotes {done}/{total} — {sym}")

    # Reset stale STOP flag at the start
    if main_ws[MAIN_CELLS["StopRequested"]].value is not None:
        main_ws[MAIN_CELLS["StopRequested"]] = "FALSE"

    info_map = fetch_all_info(
        [t[1] for t in stale_targets], delay=info_delay, progress=info_progress,
        should_stop=lambda: read_stop_requested(main_ws),
    )

    _say("Writing quote columns")
    now_iso = dt.datetime.now(dt.UTC).strftime("%Y-%m-%dT%H:%M:%SZ")
    cols = {name: _market_col(name) for name in MARKET_COLUMNS}

    successes = 0
    failures = 0
    for (row_idx, sym, ccy, indexes_csv) in targets:
        if read_stop_requested(main_ws):
            _say(f"STOP detected — refreshed {successes} of {len(targets)} rows")
            break
        try:
            series = closes[sym] if sym in closes.columns else pd.Series(dtype=float)
            eur_by_label: dict[str, Optional[float]] = {}
            for label, delta in LOOKBACKS.items():
                target = today - delta
                native = price_at_or_before(series, target)
                eur = to_eur(native, ccy, fx)
                eur_by_label[label] = eur
                market_ws.cell(row=row_idx, column=cols[f"{label} (EUR)"], value=eur)
            # % change vs Today for each non-Today lookback
            today_eur = eur_by_label.get("Today")
            for past_label, pct_col in _pct_column_pairs():
                pct = _pct_change(today_eur, eur_by_label.get(past_label))
                cell = market_ws.cell(row=row_idx, column=cols[pct_col], value=pct)
                cell.number_format = PERCENT_STYLE
            if sym in info_map:
                info = info_map[sym]
                market_ws.cell(row=row_idx, column=cols["P/E (TTM)"],   value=info.get("trailingPE"))
                market_ws.cell(row=row_idx, column=cols["Forward P/E"], value=info.get("forwardPE"))
                market_ws.cell(row=row_idx, column=cols["Last update (UTC)"], value=now_iso)
            # else: skipped by freshness filter — leave P/E + Last update at prior values
            market_ws.cell(row=row_idx, column=cols["Last error"], value=None)  # clear previous
            successes += 1
        except Exception as e:
            market_ws.cell(row=row_idx, column=cols["Last error"], value=f"{type(e).__name__}: {e}"[:300])
            failures += 1
            log.warning(f"  {sym}: {e}")

    # Migrate legacy "Monthly movers" sheet name if present, then compute
    # winners + losers from the FULL Market (not just refreshed targets) so
    # test mode with 1 refreshed ticker still shows a meaningful ranking.
    _migrate_monthly_movers_to_winners_openpyxl(wb)
    all_records = _read_market_records_openpyxl(market_ws)
    winners = _compute_monthly_winners(all_records)
    losers  = _compute_monthly_losers(all_records)
    if len(all_records) < 50:
        _say(f"  ⚠ Market only has {len(all_records)} rows — "
             "ranking will be sparse. Run a full rebuild-inventory to seed Market properly.")
    # Adaptive filter: only pre-apply Owned?=Yes if at least one owned loser
    # exists. Otherwise the filter would hide every row and the sheet looks
    # empty (the 2026-05-13 user bug).
    has_owned_losers = any(_is_owned(m.get("owned")) for m in losers)
    _say(f"Monthly winners: {len(winners)} qualify; losers: {len(losers)} "
         f"(owned among losers: {sum(1 for m in losers if _is_owned(m.get('owned')))})")
    _ensure_ranking_sheet_openpyxl(wb, MONTHLY_WINNERS_SHEET_NAME, winners)
    _ensure_ranking_sheet_openpyxl(
        wb, MONTHLY_LOSERS_SHEET_NAME, losers,
        filter_owned_yes=has_owned_losers,
    )

    # Update Main metadata
    main_ws[MAIN_CELLS["LastQuotesAt"]] = now_iso
    if not test_mode:
        # Don't overwrite the headline FX values in test mode — they only describe
        # a 1-symbol run which is rarely interesting.
        if fx.get("USD") is not None and not pd.isna(fx["USD"]):
            main_ws[MAIN_CELLS["EurUsd"]] = float(fx["USD"])
        if fx.get("JPY") is not None and not pd.isna(fx["JPY"]):
            main_ws[MAIN_CELLS["EurJpy"]] = float(fx["JPY"])
        if fx.get("GBP") is not None and not pd.isna(fx["GBP"]):
            main_ws[MAIN_CELLS["EurGbp"]] = float(fx["GBP"])

    wb.save(workbook_path)
    duration = time.time() - t0
    _say(f"get_quotes: done in {duration:.0f}s. {successes} ok, {failures} failed.")
    return 0


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------

def _cmd_legacy_run(args) -> int:
    """Legacy single-shot mode: builds a fresh stocks_report_YYYY-MM-DD.xlsx."""
    if args.indexes.upper() == "ALL":
        indexes = list(INDEX_WIKI.keys())
    else:
        indexes = [s.strip().upper() for s in args.indexes.split(",")]
        unknown = [i for i in indexes if i not in INDEX_WIKI]
        if unknown:
            log.error(f"Unknown index(es): {unknown}")
            log.error(f"Valid: {list(INDEX_WIKI.keys())}")
            return 2

    output = Path(args.output) if args.output else Path(f"stocks_report_{dt.date.today()}.xlsx")

    t0 = time.time()
    df, fx = build_report(indexes, info_delay=args.info_delay)
    write_excel(df, fx, output)
    log.info(f"Done in {time.time() - t0:.0f}s. Rows: {len(df)}. Output: {output}")
    return 0


def _cmd_init_workbook(args) -> int:
    """Create or refresh stocks.xlsx with Main + Market sheet templates."""
    init_workbook(Path(args.workbook))
    return 0


def _cmd_setup_buttons(args) -> int:
    """One-time: convert the workbook to .xlsm and add buttons + xlwings.conf sheet.

    Requires Excel + xlwings. The user still needs to import vba/stocks_picker.bas
    once via Alt+F11 — programmatic VBA injection needs Excel's "Trust access to
    the VBA project object model" setting, which we can't toggle here.

    Flow:
      - Source: whichever workbook --workbook points at (default stocks.xlsx).
      - Target: same basename with .xlsm extension.
      - On success, the source .xlsx is removed (unless source IS already .xlsm).
    """
    import xlwings as xw

    src = Path(args.workbook).resolve()
    if not src.exists():
        # init-workbook redirects .xlsm → .xlsx since openpyxl can't author
        # a valid .xlsm. Look for that fallback before erroring out.
        fallback = src.with_suffix(".xlsx")
        if fallback.exists():
            log.info(f"{src.name} doesn't exist; using {fallback.name} as the "
                     "init-workbook output to convert.")
            src = fallback
        else:
            log.error(f"Workbook {src} does not exist. Run init-workbook first.")
            return 1
    target = src.with_suffix(".xlsm")

    log.info(f"Opening {src} in Excel (headless)")
    app = xw.App(visible=False, add_book=False)
    try:
        wb = app.books.open(str(src))
        main = wb.sheets["Main"]

        # 0. Import the VBA module so the buttons' OnAction macros (RebuildInventory,
        #    GetQuotes) actually resolve when the user clicks them. Requires Excel's
        #    "Trust access to the VBA project object model" setting to be enabled —
        #    falls back to a logged warning + the old manual-import instructions
        #    if it isn't, so the script still does everything else useful.
        bas_path = (Path(__file__).resolve().parent / "vba" / "stocks_picker.bas")
        vba_imported = False
        if bas_path.exists():
            try:
                vbproject = wb.api.VBProject
                # Remove every existing std-module that defines our macros —
                # catches not only "stocks_picker" but also stale duplicates
                # under names Excel made up on prior conflicts (Module1,
                # stocks_picker1, etc.). Without this, re-running this command
                # produces the "Ambiguous name detected: RebuildInventory"
                # error when the user clicks a button.
                signatures = ("Sub RebuildInventory", "Sub GetQuotes")
                doomed: list = []
                for i in range(1, vbproject.VBComponents.Count + 1):
                    comp = vbproject.VBComponents.Item(i)
                    if comp.Type != 1:  # 1 = vbext_ct_StdModule
                        continue
                    try:
                        n_lines = comp.CodeModule.CountOfLines
                        if n_lines == 0:
                            continue
                        code = comp.CodeModule.Lines(1, n_lines)
                        if any(sig in code for sig in signatures):
                            doomed.append(comp)
                    except Exception:
                        pass
                for comp in doomed:
                    log.info(f"  removing prior VBA module '{comp.Name}' "
                             "(contained RebuildInventory / GetQuotes)")
                    try:
                        vbproject.VBComponents.Remove(comp)
                    except Exception as e:
                        log.warning(f"    couldn't remove {comp.Name}: {e}")
                vbproject.VBComponents.Import(str(bas_path))
                vba_imported = True
                log.info(f"Imported VBA module from {bas_path.name}")
            except Exception as e:
                msg = str(e)
                if "trust" in msg.lower() or "Programmatic access" in msg:
                    log.warning(
                        "Could not import VBA programmatically — Excel's "
                        "\"Trust access to the VBA project object model\" "
                        "setting is OFF. The .bas file at "
                        f"{bas_path} must be imported manually via Alt+F11."
                    )
                else:
                    log.warning(f"VBA import failed ({type(e).__name__}: {msg}); "
                                "will need manual Alt+F11 import.")
        else:
            log.warning(f"VBA source not found at {bas_path}; skipping import.")

        # 1. Add xlwings.conf sheet (used by xlwings to find Python interpreter).
        if "xlwings.conf" not in [s.name for s in wb.sheets]:
            conf = wb.sheets.add("xlwings.conf", after=wb.sheets[wb.sheets.count - 1])
        else:
            conf = wb.sheets["xlwings.conf"]
        interpreter = (Path.cwd() / ".venv" / "Scripts" / "python.exe").resolve()
        pythonpath = Path.cwd().resolve()
        conf_rows = [
            ["INTERPRETER_WIN", str(interpreter)],
            ["PYTHONPATH",      str(pythonpath)],
            ["SHOW CONSOLE",    "False"],
        ]
        conf.range("A1").value = conf_rows
        try:
            conf.api.Visible = 2  # xlSheetVeryHidden (only visible via VBA editor)
        except Exception:
            pass

        # 2. Remove any prior StockPicker_* shapes we added (idempotent —
        #    covers buttons + the test-mode checkbox).
        for shp in list(main.shapes):
            if shp.name.startswith("StockPicker_"):
                shp.delete()

        # 3. Add two buttons + a Test-mode checkbox in the Jobs area to the
        #    right of column B. Column A (~38 chars ≈ 270 px) and column B
        #    (~60 chars ≈ 430 px) together reach left=700, so controls live
        #    at left=720+ to avoid hiding either.
        #    Shapes.AddFormControl(type, left, top, w, h):
        #      0 = msoFormControlButton
        #      1 = msoFormControlCheckBox
        sheet_api = main.api
        btn1 = sheet_api.Shapes.AddFormControl(0, 720, 10, 130, 28)
        btn1.Name = "StockPicker_Rebuild"
        btn1.TextFrame.Characters().Text = "Rebuild Inventory"
        btn1.OnAction = "RebuildInventory"
        btn2 = sheet_api.Shapes.AddFormControl(0, 720, 42, 130, 28)
        btn2.Name = "StockPicker_GetQuotes"
        btn2.TextFrame.Characters().Text = "Get Quotes"
        btn2.OnAction = "GetQuotes"

        # 4. Add a Test-mode checkbox, linked to Main!B5 (the existing
        #    MAIN_CELLS["TestMode"] address). Excel writes TRUE/FALSE into
        #    the linked cell automatically when the user clicks it, and
        #    read_test_mode() already understands those values.
        chk = sheet_api.Shapes.AddFormControl(1, 720, 78, 160, 24)
        chk.Name = "StockPicker_TestMode"
        chk.TextFrame.Characters().Text = "Test mode (BEL20 + 1 quote)"
        chk.ControlFormat.LinkedCell = f"Main!{MAIN_CELLS['TestMode']}"
        # Default state: unchecked (xlOff = -4146). If the user previously
        # set the cell to TRUE manually, mirror that into the checkbox.
        existing = main.range(MAIN_CELLS["TestMode"]).value
        if isinstance(existing, str) and existing.strip().upper() in _TEST_MODE_TRUTHY:
            chk.ControlFormat.Value = 1   # xlOn
        elif existing is True:
            chk.ControlFormat.Value = 1
        else:
            chk.ControlFormat.Value = -4146  # xlOff

        # 5. Add the STOP checkbox, linked to Main!StopRequested. Excel
        #    Form Controls write their state to LinkedCell via the internal
        #    cell engine — instantly, even while xlwings' RunPython has the
        #    VBA thread blocked. That's the property that makes a checkbox
        #    the only Form Control type that works for a "stop running job"
        #    signal (regular buttons queue their macro until VBA frees).
        #
        #    Add an A13 label inline since the cosmetic-preservation rule
        #    blocks init-workbook from doing it on existing workbooks.
        if not main.range("A13").value:
            main.range("A13").value = "Stop a running job:"
        # Hide the value display in B13 so the user sees only the checkbox.
        try:
            main.range(MAIN_CELLS["StopRequested"]).api.NumberFormat = ";;;"
            main.range(MAIN_CELLS["JobRunning"]).api.NumberFormat = ";;;"
        except Exception:
            pass
        stop_chk = sheet_api.Shapes.AddFormControl(1, 720, 110, 200, 28)
        stop_chk.Name = "StockPicker_Stop"
        stop_chk.TextFrame.Characters().Text = "⛔ STOP running job"
        # Make the STOP label visually distinctive (red bold) so the user
        # never confuses it with the Test-mode checkbox above.
        try:
            stop_chk.TextFrame.Characters().Font.Bold = True
            stop_chk.TextFrame.Characters().Font.Color = 255  # red (BGR=0x0000FF)
        except Exception:
            pass
        stop_chk.ControlFormat.LinkedCell = f"Main!{MAIN_CELLS['StopRequested']}"
        stop_chk.ControlFormat.Value = -4146  # xlOff at install

        # 6. SaveAs .xlsm (FileFormat 52 = xlOpenXMLWorkbookMacroEnabled).
        if target.exists() and target != src:
            target.unlink()  # Excel SaveAs refuses to overwrite without prompts in headless mode
        wb.api.SaveAs(str(target), FileFormat=52)
        wb.close()

        # Remove the source .xlsx if it's distinct from the .xlsm target.
        if src != target and src.exists():
            src.unlink()

        log.info(f"setup-buttons: workbook saved as {target}")
        log.info("")
        if vba_imported:
            log.info("VBA module imported — the buttons should respond to clicks as")
            log.info(f"soon as you open {target.name} and click 'Enable Content'.")
        else:
            log.info("Manual VBA-import step still needed (couldn't do it auto-)")
            log.info(f"  1. Open {target.name} in Excel.")
            log.info("  2. Press Alt+F11 -> File -> Import File -> select vba/stocks_picker.bas.")
            log.info("  3. Save.")
            log.info("Enable 'Trust access to the VBA project object model' in Excel's")
            log.info("Trust Center to skip this step on the next setup-buttons run.")
        log.info("")
        log.info("First-time-per-machine step (skip if already done):")
        log.info(r"  .\.venv\Scripts\xlwings.exe addin install")
        log.info(r"  Workbook VBA editor: Tools -> References -> tick 'xlwings'")
        return 0
    finally:
        app.quit()


def _cmd_get_quotes(args) -> int:
    """Job 2 — refresh quote columns (prices + P/E) for every Market row."""
    # --test forces test mode; without it, the workbook's TestMode cell decides.
    return get_quotes(
        workbook_path=_resolve_workbook(args.workbook),
        test_mode=True if args.test else None,
        info_delay=args.info_delay,
    )


def _cmd_rebuild_inventory(args) -> int:
    """Job 1 — refresh the Market sheet structural data (no quotes)."""
    # --test forces BEL20 only; without it, the workbook's TestMode cell decides.
    indexes = None
    if not args.test and args.indexes and args.indexes.upper() != "ALL":
        indexes = [s.strip().upper() for s in args.indexes.split(",")]
        unknown = [i for i in indexes if i not in INDEX_WIKI]
        if unknown:
            log.error(f"Unknown index(es): {unknown}")
            log.error(f"Valid: {list(INDEX_WIKI.keys())}")
            return 2
    return rebuild_inventory(
        workbook_path=_resolve_workbook(args.workbook),
        indexes=indexes,
        info_delay=args.info_delay,
        test_mode=True if args.test else None,
    )


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__.split("\n\n")[0])
    sub = parser.add_subparsers(dest="cmd")

    # init-workbook — create or refresh the persistent workbook
    p_init = sub.add_parser(
        "init-workbook",
        help="Create or refresh stocks.xlsm with Main + Market sheet templates.",
    )
    p_init.add_argument(
        "--workbook", default=str(DEFAULT_WORKBOOK_PATH),
        help=f"Workbook path (default: {DEFAULT_WORKBOOK_PATH})",
    )

    # setup-buttons — one-time: add Excel buttons + xlwings.conf sheet
    p_setup = sub.add_parser(
        "setup-buttons",
        help="One-time: add Excel buttons + xlwings.conf sheet to the workbook.",
    )
    p_setup.add_argument(
        "--workbook", default=str(DEFAULT_WORKBOOK_PATH),
        help=f"Workbook path (default: {DEFAULT_WORKBOOK_PATH})",
    )

    # get-quotes — Job 2 (refresh prices + P/E + per-row last update/error)
    p_quotes = sub.add_parser(
        "get-quotes",
        help="Job 2: refresh quote columns in the Market sheet (run daily).",
    )
    p_quotes.add_argument(
        "--workbook", default=str(DEFAULT_WORKBOOK_PATH),
        help=f"Workbook path (default: {DEFAULT_WORKBOOK_PATH})",
    )
    p_quotes.add_argument(
        "--info-delay", type=float, default=0.25,
        help="Seconds between yfinance .info calls (default: 0.25)",
    )
    p_quotes.add_argument(
        "--test", action="store_true",
        help="Test mode: refresh 1 symbol only (top of Main portfolio or first BEL20).",
    )

    # rebuild-inventory — Job 1 (no quotes; populates Market structural data)
    p_rebuild = sub.add_parser(
        "rebuild-inventory",
        help="Job 1: refresh the Market sheet's structural data (no quote fetch).",
    )
    p_rebuild.add_argument(
        "--workbook", default=str(DEFAULT_WORKBOOK_PATH),
        help=f"Workbook path (default: {DEFAULT_WORKBOOK_PATH})",
    )
    p_rebuild.add_argument(
        "--indexes", default="ALL",
        help="Subset of indexes to refresh (default: ALL).",
    )
    p_rebuild.add_argument(
        "--info-delay", type=float, default=0.25,
        help="Seconds between yfinance .info calls (default: 0.25)",
    )
    p_rebuild.add_argument(
        "--test", action="store_true",
        help="Test mode: BEL20 only.",
    )

    # run — legacy single-shot all-in-one report (default when no subcommand)
    p_run = sub.add_parser(
        "run",
        help="Legacy single-shot: write a fresh dated xlsx with everything.",
    )
    p_run.add_argument(
        "--indexes", default="ALL",
        help="Comma-separated: SP500,NIKKEI225,FTSE100,DAX,CAC40,BEL20,ESTOXX50 (default: ALL)",
    )
    p_run.add_argument(
        "--output", default=None,
        help="Output xlsx path (default: stocks_report_YYYY-MM-DD.xlsx in current dir)",
    )
    p_run.add_argument(
        "--info-delay", type=float, default=0.25,
        help="Seconds between yfinance .info calls (default: 0.25)",
    )

    # Legacy compat: when called as `stocks_report.py --indexes BEL20` with
    # no subcommand, fall back to the run subparser. Detect by looking for
    # a known top-level flag.
    known_subcommands = {"init-workbook", "setup-buttons", "rebuild-inventory", "get-quotes", "run", "-h", "--help"}
    raw_args = sys.argv[1:]
    if raw_args and raw_args[0] not in known_subcommands:
        raw_args = ["run", *raw_args]
    args = parser.parse_args(raw_args)

    if args.cmd is None or args.cmd == "run":
        # Provide defaults for legacy invocation (`stocks_report.py` with no args).
        for attr, default in (("indexes", "ALL"), ("output", None), ("info_delay", 0.25)):
            if not hasattr(args, attr):
                setattr(args, attr, default)
        return _cmd_legacy_run(args)


    if args.cmd == "init-workbook":
        return _cmd_init_workbook(args)
    if args.cmd == "setup-buttons":
        return _cmd_setup_buttons(args)
    if args.cmd == "rebuild-inventory":
        return _cmd_rebuild_inventory(args)
    if args.cmd == "get-quotes":
        return _cmd_get_quotes(args)

    parser.print_help()
    return 2


if __name__ == "__main__":
    sys.exit(main())
