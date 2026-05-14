"""Offline unit tests for stocks_report. Run with: python -m pytest tests/ -v"""

import datetime as dt
import sys
from pathlib import Path

import pandas as pd
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))
from stocks_report import (  # noqa: E402
    CURRENCY_HEADERS,
    ETF_LIST,
    FX_PAIRS,
    MAIN_CELLS,
    MARKET_COLUMNS,
    MONTHLY_LOSERS_SHEET_NAME,
    MONTHLY_MOVERS_LEGACY_NAME,
    MONTHLY_WINNERS_SHEET_NAME,
    OWNED_COL_INDEX_1BASED,
    PORTFOLIO_FIRST_ROW,
    RANKING_HEADERS,
    SCHEMA_VERSION,
    VERSION_HISTORY,
    _append_error_log,
    _compute_monthly_losers,
    _compute_monthly_winners,
    _currency_rows,
    _ensure_help_sheet_versions,
    _extract_dataroma_tickers,
    _migrate_monthly_movers_to_winners_openpyxl,
    _owned_for,
    _parse_bel20_ticker,
    _parse_nikkei225_components,
    _pct_change,
    _pct_column_pairs,
    _pick_test_target,
    _resolve_workbook,
    aggregate_constituents,
    get_etfs,
    init_workbook,
    normalize_ticker,
    price_at_or_before,
    read_portfolio_symbols,
    read_test_mode,
    to_eur,
    yahoo_quote_url,
)


# ---------------------------------------------------------------------------
# normalize_ticker
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("raw,index_name,expected", [
    # S&P 500: dot becomes hyphen
    ("AAPL",     "SP500",     "AAPL"),
    ("BRK.B",    "SP500",     "BRK-B"),
    ("BF.B",     "SP500",     "BF-B"),
    # DAX
    ("SAP",      "DAX",       "SAP.DE"),
    ("SAP.DE",   "DAX",       "SAP.DE"),  # idempotent
    # CAC 40
    ("AIR",      "CAC40",     "AIR.PA"),
    # BEL 20
    ("KBC",      "BEL20",     "KBC.BR"),
    ("APAM.AS",  "BEL20",     "APAM.AS"),  # Amsterdam-listed, must not be re-suffixed
    # Cross-index dedup — regression for the 2026-05-13 batch of bogus rows:
    ("AIR.PA",   "DAX",       "AIR.PA"),   # Airbus in DAX; was → AIR.PA.DE
    ("MT.AS",    "CAC40",     "MT.AS"),    # ArcelorMittal in CAC40; was → MT.AS.PA
    ("ADS.DE",   "ESTOXX50",  "ADS.DE"),   # Adidas in ESTOXX50; suffix already present
    # FTSE 100 sub-class shares use a hyphen on Yahoo (like SP500's BRK-B):
    ("BT.A",     "FTSE100",   "BT-A.L"),   # was → BT.A.L
    ("BT-A.L",   "FTSE100",   "BT-A.L"),   # idempotent
    # EuroStoxx 50: Wikipedia provides fully-qualified Yahoo tickers — pass through
    ("ADS.DE",   "ESTOXX50",  "ADS.DE"),
    ("ADYEN.AS", "ESTOXX50",  "ADYEN.AS"),
    ("AI.PA",    "ESTOXX50",  "AI.PA"),
    ("ISP.MI",   "ESTOXX50",  "ISP.MI"),  # Milan
    ("IBE.MC",   "ESTOXX50",  "IBE.MC"),  # Madrid
    # FTSE 100: trailing dot stripped, .L appended
    ("ULVR",     "FTSE100",   "ULVR.L"),
    ("RR.",      "FTSE100",   "RR.L"),
    ("BHP.",     "FTSE100",   "BHP.L"),
    # Nikkei 225 — alphanumeric tickers like 543A.T (Archion) are valid
    ("7203",     "NIKKEI225", "7203.T"),
    ("543A",     "NIKKEI225", "543A.T"),
    ("285A",     "NIKKEI225", "285A.T"),
    # Robustness: footnote markers, lowercase
    ("AAPL[1]",  "SP500",     "AAPL"),
    ("aapl",     "SP500",     "AAPL"),
    ("  AAPL ",  "SP500",     "AAPL"),
])
def test_normalize_ticker(raw, index_name, expected):
    assert normalize_ticker(raw, index_name) == expected


# ---------------------------------------------------------------------------
# _parse_bel20_ticker — Wikipedia BEL 20 cell format is "Euronext X:\xa0SYMBOL"
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("cell,expected", [
    ("Euronext Brussels:\xa0ABI",     "ABI.BR"),
    ("Euronext Brussels:\xa0ARGX",    "ARGX.BR"),
    ("Euronext Amsterdam:\xa0APAM",   "APAM.AS"),
    ("Euronext Brussels: KBC",        "KBC.BR"),  # plain space
    ("euronext brussels:\xa0umi",     "UMI.BR"),  # case-insensitive
    ("garbage value",                 ""),         # no match
    ("",                              ""),
])
def test_parse_bel20_ticker(cell, expected):
    assert _parse_bel20_ticker(cell) == expected


# ---------------------------------------------------------------------------
# _parse_nikkei225_components — bullet-list scraping (no table on the page)
# ---------------------------------------------------------------------------

_NIKKEI_FIXTURE = """
<h2 id="Components">Components</h2>
<ul>
  <li>Energy (0.30%)</li>
  <li>Materials (7.00%)</li>
</ul>
<h3>Air transport</h3>
<ul>
  <li>ANA Holdings Inc. (TYO: 9202)</li>
  <li>Japan Airlines Co., Ltd. (TYO: 9201)</li>
</ul>
<h3>Automotive</h3>
<ul>
  <li>Archion Corp. (TYO: 543A)</li>
  <li>Honda Motor Co., Ltd. (TYO: 7267)</li>
  <li>M3 Inc. (TYO: 2413) (Parent company for MDLinx)</li>
</ul>
<h2 id="Statistics">Statistics</h2>
<li>This must not match (TYO: 9999)</li>
"""


def test_parse_nikkei225_components_extracts_pairs():
    # min_count=0 bypasses the production safety floor since the fixture is tiny.
    df = _parse_nikkei225_components(_NIKKEI_FIXTURE, min_count=0)
    pairs = list(zip(df["Symbol"], df["Name"]))
    # 5 stocks; sector summary lines and the post-Statistics li are excluded.
    assert pairs == [
        ("9202.T", "ANA Holdings Inc."),
        ("9201.T", "Japan Airlines Co., Ltd."),
        ("543A.T", "Archion Corp."),
        ("7267.T", "Honda Motor Co., Ltd."),
        ("2413.T", "M3 Inc."),
    ]
    assert (df["Index"] == "NIKKEI225").all()


def test_parse_nikkei225_components_ignores_trailing_annotations():
    """Entries like 'Foo (TYO: 9147)(Holding company for Foo)' must still parse."""
    html = '<h2 id="Components">x</h2><li>Foo Corp. (TYO: 9147)(Holding company for Foo)</li><h2 id="Statistics">x</h2>'
    df = _parse_nikkei225_components(html, min_count=0)
    assert list(df["Symbol"]) == ["9147.T"]
    assert list(df["Name"]) == ["Foo Corp."]


def test_parse_nikkei225_components_raises_when_no_components_section():
    with pytest.raises(RuntimeError, match="Components section"):
        _parse_nikkei225_components("<html>no components header</html>")


def test_parse_nikkei225_components_raises_when_too_few_results():
    """Safety net: if a future page change makes the parser yield only a handful
    of entries, surface as an error rather than silently writing 5 rows."""
    html = '<h2 id="Components">x</h2><li>Foo (TYO: 9001)</li><h2 id="Statistics">x</h2>'
    with pytest.raises(RuntimeError, match="Wikipedia layout likely changed"):
        _parse_nikkei225_components(html)  # uses default min_count=150


# ---------------------------------------------------------------------------
# _extract_dataroma_tickers — pulls /m/stock.php?sym=X links from HTML
# ---------------------------------------------------------------------------

def test_extract_dataroma_tickers_basic():
    html = (
        '<tr><td class="sym"><a href="/m/stock.php?sym=AMZN">AMZN</a></td>'
        '<td class="stock"><a href="/m/stock.php?sym=AMZN">Amazon.com Inc.</a></td></tr>'
        '<tr><td class="sym"><a href="/m/stock.php?sym=MSFT">MSFT</a></td></tr>'
    )
    assert _extract_dataroma_tickers(html) == {"AMZN", "MSFT"}


def test_extract_dataroma_tickers_dual_form_for_dot_classes():
    """Dataroma writes BRK.B; SP500 in this report uses BRK-B. Both forms must come back."""
    html = '<a href="/m/stock.php?sym=BRK.B">BRK.B</a> <a href="/m/stock.php?sym=BF.B">BF.B</a>'
    out = _extract_dataroma_tickers(html)
    assert {"BRK.B", "BRK-B", "BF.B", "BF-B"} <= out


def test_extract_dataroma_tickers_empty_on_no_match():
    assert _extract_dataroma_tickers("<html>no links here</html>") == set()


def test_extract_dataroma_tickers_ignores_unrelated_links():
    html = '<a href="/m/managers.php">Managers</a> <a href="/m/stock.php?sym=AAPL">AAPL</a>'
    assert _extract_dataroma_tickers(html) == {"AAPL"}


# ---------------------------------------------------------------------------
# aggregate_constituents — dedups by Symbol, joins Indexes
# ---------------------------------------------------------------------------

def _df(rows):
    return pd.DataFrame(rows, columns=["Symbol", "Name", "Index"])


def test_aggregate_constituents_dedups_by_symbol():
    """A stock in two indexes yields one row with both indexes listed."""
    dax = _df([("SAP.DE", "SAP", "DAX"), ("ALV.DE", "Allianz", "DAX")])
    estoxx = _df([("SAP.DE", "SAP SE", "ESTOXX50"), ("AIR.PA", "Airbus", "ESTOXX50")])
    out = aggregate_constituents([dax, estoxx])
    rows = {r["Symbol"]: r for _, r in out.iterrows()}
    assert rows["SAP.DE"]["Indexes"] == "DAX, ESTOXX50"
    assert rows["ALV.DE"]["Indexes"] == "DAX"
    assert rows["AIR.PA"]["Indexes"] == "ESTOXX50"
    assert len(out) == 3  # SAP collapsed into one row


def test_aggregate_constituents_preserves_first_seen_name():
    dax = _df([("SAP.DE", "SAP", "DAX")])
    estoxx = _df([("SAP.DE", "SAP SE", "ESTOXX50")])  # different Name spelling
    out = aggregate_constituents([dax, estoxx])
    assert out.iloc[0]["Name"] == "SAP"  # DAX listed first → its Name wins


def test_aggregate_constituents_dedups_repeated_index():
    """Defensive: if the same Symbol+Index appears twice, Indexes lists it once."""
    a = _df([("AAPL", "Apple", "SP500"), ("AAPL", "Apple", "SP500")])
    out = aggregate_constituents([a])
    assert out.iloc[0]["Indexes"] == "SP500"


def test_aggregate_constituents_empty_input():
    out = aggregate_constituents([])
    assert list(out.columns) == ["Symbol", "Name", "Indexes"]
    assert len(out) == 0


# ---------------------------------------------------------------------------
# init_workbook — creates Main + Market layout, idempotent
# ---------------------------------------------------------------------------

def test_init_workbook_creates_main_and_market(tmp_path):
    """Fresh-create writes both sheets with the expected headers and named cells."""
    from openpyxl import load_workbook
    path = tmp_path / "stocks.xlsx"
    init_workbook(path)
    assert path.exists()
    wb = load_workbook(path)
    assert wb.sheetnames == ["Main", "Market", "Help"]
    # Market headers match MARKET_COLUMNS
    market = wb["Market"]
    headers = [market.cell(row=1, column=i).value for i in range(1, len(MARKET_COLUMNS) + 1)]
    assert headers == MARKET_COLUMNS
    # Main has the title; TestMode cell is intentionally blank on fresh init
    # (the checkbox added by setup-buttons writes TRUE/FALSE into it).
    main = wb["Main"]
    assert main["A1"].value == "Stock Picker"
    assert main[MAIN_CELLS["TestMode"]].value is None


def test_init_workbook_is_idempotent(tmp_path):
    """Running init twice does not clobber data the user added."""
    from openpyxl import load_workbook
    path = tmp_path / "stocks.xlsx"
    init_workbook(path)

    # User adds a portfolio entry and toggles test mode.
    wb = load_workbook(path)
    main = wb["Main"]
    main.cell(row=PORTFOLIO_FIRST_ROW, column=1, value="KBC.BR")
    main.cell(row=PORTFOLIO_FIRST_ROW, column=2, value=10)
    main[MAIN_CELLS["TestMode"]] = "TRUE"
    market = wb["Market"]
    market["A2"] = "FAKE.XX"   # pretend Market has data already
    wb.save(path)

    # Run init again — user data MUST survive.
    init_workbook(path)
    wb2 = load_workbook(path)
    assert wb2["Main"].cell(row=PORTFOLIO_FIRST_ROW, column=1).value == "KBC.BR"
    assert wb2["Main"].cell(row=PORTFOLIO_FIRST_ROW, column=2).value == 10
    assert wb2["Main"][MAIN_CELLS["TestMode"]].value == "TRUE"
    assert wb2["Market"]["A2"].value == "FAKE.XX"


def test_init_workbook_market_columns_unique():
    """No duplicate column names — otherwise lookups by name would be ambiguous."""
    assert len(MARKET_COLUMNS) == len(set(MARKET_COLUMNS))


def test_init_workbook_preserves_user_cosmetic_edits(tmp_path):
    """User edits to label text and column widths must survive a re-run."""
    from openpyxl import load_workbook
    path = tmp_path / "stocks.xlsx"
    init_workbook(path)

    # Pretend the user renamed the title and widened column A.
    wb = load_workbook(path)
    wb["Main"]["A1"] = "Patrick's Stock Picker"
    wb["Main"].column_dimensions["A"].width = 100
    wb["Main"]["A3"] = "My Custom Jobs Section"  # custom section label
    wb.save(path)

    init_workbook(path)  # re-run on existing file
    wb2 = load_workbook(path)
    assert wb2["Main"]["A1"].value == "Patrick's Stock Picker"
    assert wb2["Main"]["A3"].value == "My Custom Jobs Section"
    assert wb2["Main"].column_dimensions["A"].width == 100


def test_init_workbook_preserves_cleared_labels(tmp_path):
    """If the user CLEARED a label (set to None), a re-run must not restore it.

    This is the trap that caught us on 2026-05-13 — the prior _set helper
    rewrote any cell whose value was None, which silently undid the user's
    deliberate clearances of A3 and A5.
    """
    from openpyxl import load_workbook
    path = tmp_path / "stocks.xlsx"
    init_workbook(path)

    # User clears two section labels.
    wb = load_workbook(path)
    wb["Main"]["A3"] = None
    wb["Main"]["A5"] = None
    wb.save(path)

    init_workbook(path)  # re-run on existing file
    wb2 = load_workbook(path)
    assert wb2["Main"]["A3"].value is None, "A3 was restored after re-run; user cleared it"
    assert wb2["Main"]["A5"].value is None, "A5 was restored after re-run; user cleared it"


# ---------------------------------------------------------------------------
# ETF list + get_etfs()
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# _append_error_log — appends a timestamped block to the log file
# ---------------------------------------------------------------------------

def test_append_error_log_creates_file(tmp_path):
    log = tmp_path / "stocks_errors.log"
    _append_error_log(log, "RuntimeError", "boom", "Traceback line 1\nline 2\n")
    text = log.read_text(encoding="utf-8")
    assert "RuntimeError: boom" in text
    assert "Traceback line 1" in text
    assert "line 2" in text
    # ISO timestamp prefix should be present
    assert "T" in text and "Z" in text


def test_append_error_log_appends_to_existing(tmp_path):
    log = tmp_path / "stocks_errors.log"
    _append_error_log(log, "ValueError", "first", "tb1\n")
    _append_error_log(log, "TypeError", "second", "tb2\n")
    text = log.read_text(encoding="utf-8")
    # Both entries present
    assert "ValueError: first" in text
    assert "TypeError: second" in text
    # Second is after first
    assert text.index("first") < text.index("second")


def test_append_error_log_swallows_io_errors(tmp_path):
    """Log helper must never raise — its caller is already handling an
    exception and a follow-on crash would mask the original error."""
    # Pass a nonexistent parent directory; the call must NOT raise.
    bogus = tmp_path / "no" / "such" / "dir" / "errors.log"
    _append_error_log(bogus, "RuntimeError", "boom", "tb\n")  # should not raise


def test_etf_list_has_one_per_index_plus_sector_and_country():
    """Sanity: 7 index ETFs + 11 sector ETFs + 19+ country ETFs."""
    assert len(ETF_LIST) >= 30, f"ETF_LIST shrank to {len(ETF_LIST)}; was 37+"
    labels = [label for _, _, label in ETF_LIST]
    # Must cover every tracked index
    for idx in ("SP500", "NIKKEI225", "FTSE100", "DAX", "CAC40", "BEL20", "ESTOXX50"):
        assert any(f"ETF — {idx}" == lbl for lbl in labels), f"No ETF for {idx}"
    # Must have at least one sector ETF labelled "Utilities" (user explicitly named it)
    assert any("Sector: Utilities" in lbl for lbl in labels)
    # Must have several country ETFs
    country_count = sum(1 for lbl in labels if lbl.startswith("ETF — Country:"))
    assert country_count >= 15


def test_etf_tickers_are_unique():
    tickers = [sym for sym, _, _ in ETF_LIST]
    assert len(tickers) == len(set(tickers)), "Duplicate ETF tickers"


def test_get_etfs_returns_constituent_shape():
    df = get_etfs()
    assert list(df.columns) == ["Symbol", "Name", "Index"]
    assert len(df) == len(ETF_LIST)
    # Every row's Index label starts with "ETF — "
    assert df["Index"].str.startswith("ETF —").all()


def test_get_etfs_aggregates_with_index_data():
    """ETFs flow through aggregate_constituents alongside index DataFrames."""
    import pandas as pd
    index_df = pd.DataFrame(
        [("AAPL", "Apple", "SP500"), ("KBC.BR", "KBC", "BEL20")],
        columns=["Symbol", "Name", "Index"],
    )
    combined = aggregate_constituents([index_df, get_etfs()])
    syms = set(combined["Symbol"])
    assert "AAPL" in syms
    assert "SPY" in syms  # known ETF
    assert "XLU" in syms  # the Utilities sector ETF specifically


@pytest.mark.parametrize("symbol,expected", [
    ("AAPL",    "https://finance.yahoo.com/quote/AAPL"),
    ("KBC.BR",  "https://finance.yahoo.com/quote/KBC.BR"),
    ("BRK-B",   "https://finance.yahoo.com/quote/BRK-B"),
    ("7203.T",  "https://finance.yahoo.com/quote/7203.T"),
    ("543A.T",  "https://finance.yahoo.com/quote/543A.T"),
])
def test_yahoo_quote_url(symbol, expected):
    assert yahoo_quote_url(symbol) == expected


# ---------------------------------------------------------------------------
# % change helpers
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("today,past,expected", [
    (110, 100, 0.10),       # +10%
    (90,  100, -0.10),      # -10%
    (100, 100, 0.0),
    (100.5, 100, 0.005),
    (None, 100, None),       # missing today
    (100, None, None),       # missing past
    (100, 0,   None),        # divide-by-zero
    (100, "x", None),        # bad type
])
def test_pct_change(today, past, expected):
    out = _pct_change(today, past)
    if expected is None:
        assert out is None
    else:
        assert out == pytest.approx(expected)


def test_pct_column_pairs_covers_every_non_today_lookback():
    """Exactly one (label, column_name) pair per lookback except Today."""
    pairs = _pct_column_pairs()
    expected_labels = ["1D ago", "1W ago", "1M ago", "6M ago", "1Y ago", "5Y ago"]
    assert [lbl for lbl, _ in pairs] == expected_labels
    # Each column name must exist in MARKET_COLUMNS so writes don't KeyError.
    for _label, col in pairs:
        assert col in MARKET_COLUMNS, f"{col!r} missing from MARKET_COLUMNS"


def test_market_columns_includes_pct_columns_interleaved():
    """After every '<X> ago (EUR)' column the next column must be '<X> %'."""
    for i, name in enumerate(MARKET_COLUMNS):
        if name.endswith(" ago (EUR)"):
            label_short = name.replace(" ago (EUR)", "")
            assert MARKET_COLUMNS[i + 1] == f"{label_short} %", (
                f"After {name!r}, expected {label_short!r} %, got "
                f"{MARKET_COLUMNS[i+1]!r}"
            )


# ---------------------------------------------------------------------------
# Schema version, VERSION_HISTORY, _resolve_workbook
# ---------------------------------------------------------------------------

def test_schema_version_string_format():
    """v01, v02, ... — letter v then digits."""
    assert SCHEMA_VERSION.startswith("v")
    assert SCHEMA_VERSION[1:].isdigit()


def test_version_history_includes_current_schema():
    """Every released SCHEMA_VERSION must have a VERSION_HISTORY entry."""
    versions = {v for v, _date, _msg in VERSION_HISTORY}
    assert SCHEMA_VERSION in versions, (
        f"SCHEMA_VERSION={SCHEMA_VERSION!r} missing from VERSION_HISTORY"
    )


def test_version_history_unique_keys():
    versions = [v for v, _date, _msg in VERSION_HISTORY]
    assert len(versions) == len(set(versions)), "Duplicate version in VERSION_HISTORY"


def test_resolve_workbook_uses_explicit_path_when_not_default(tmp_path):
    """A non-default --workbook always wins, even if a versioned file exists."""
    user_path = tmp_path / "my_custom.xlsm"
    assert _resolve_workbook(str(user_path)) == user_path


def test_resolve_workbook_finds_highest_version(tmp_path, monkeypatch):
    """When the caller passes our DEFAULT_WORKBOOK_PATH, pick the existing
    stocks_picker_vNN.xlsm with the highest NN."""
    from stocks_report import DEFAULT_WORKBOOK_PATH
    monkeypatch.chdir(tmp_path)
    (tmp_path / "stocks_picker_v01.xlsm").touch()
    (tmp_path / "stocks_picker_v03.xlsm").touch()
    (tmp_path / "stocks_picker_v02.xlsm").touch()
    out = _resolve_workbook(str(DEFAULT_WORKBOOK_PATH))
    assert out.name == "stocks_picker_v03.xlsm"


def test_resolve_workbook_falls_back_to_legacy_xlsm(tmp_path, monkeypatch):
    from stocks_report import DEFAULT_WORKBOOK_PATH
    monkeypatch.chdir(tmp_path)
    (tmp_path / "stocks.xlsm").touch()
    out = _resolve_workbook(str(DEFAULT_WORKBOOK_PATH))
    assert out.name == "stocks.xlsm"


# ---------------------------------------------------------------------------
# Help-sheet population
# ---------------------------------------------------------------------------

def test_ensure_help_sheet_versions_appends_missing(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Help"
    _ensure_help_sheet_versions(ws, fresh=True)
    # Every entry in VERSION_HISTORY should appear in column A
    col_a = [ws.cell(row=r, column=1).value for r in range(1, ws.max_row + 1)]
    for ver, _date, _msg in VERSION_HISTORY:
        assert ver in col_a, f"{ver} missing from Help sheet"


def test_ensure_help_sheet_versions_preserves_user_credit_line(tmp_path):
    """User has a credit line at A1; our entries get appended below."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "Developed by Claude"
    _ensure_help_sheet_versions(ws, fresh=False)
    assert ws["A1"].value == "Developed by Claude", "user credit clobbered"
    col_a_below = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
    for ver, _date, _msg in VERSION_HISTORY:
        assert ver in col_a_below


def test_ensure_help_sheet_versions_does_not_duplicate(tmp_path):
    """Running twice does not re-append already-mentioned versions."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    _ensure_help_sheet_versions(ws, fresh=True)
    rows_after_first = ws.max_row
    _ensure_help_sheet_versions(ws, fresh=False)
    assert ws.max_row == rows_after_first, "second call appended duplicate entries"


# ---------------------------------------------------------------------------
# Currencies sheet helpers
# ---------------------------------------------------------------------------

def test_chf_is_in_fx_pairs():
    assert "CHF" in FX_PAIRS
    assert FX_PAIRS["CHF"] == "EURCHF=X"


def test_currency_headers_match_lookbacks():
    """Pair + every LOOKBACKS entry."""
    from stocks_report import LOOKBACKS
    assert CURRENCY_HEADERS == ["Pair"] + list(LOOKBACKS.keys())


def test_currency_rows_one_per_pair(tmp_path):
    """_currency_rows produces one row per FX_PAIRS entry, each starting with
    "EUR/{ccy}" and having a value per LOOKBACKS column."""
    from stocks_report import LOOKBACKS
    fake_history = {ccy: pd.Series(dtype=float) for ccy in FX_PAIRS}
    rows = _currency_rows(fake_history)
    assert len(rows) == len(FX_PAIRS)
    for r in rows:
        assert len(r) == 1 + len(LOOKBACKS)
        assert r[0].startswith("EUR/")


# ---------------------------------------------------------------------------
# Monthly movers
# ---------------------------------------------------------------------------

def test_compute_monthly_winners_filters_and_sorts():
    """Filter: 1D >= 0 AND 1W >= 0 AND 1M > 0. Sort: 1M desc."""
    records = [
        {"symbol": "A", "pct_1d": 0.01,  "pct_1w": 0.02,  "pct_1m": 0.10},  # qualifies
        {"symbol": "B", "pct_1d": 0.01,  "pct_1w": 0.02,  "pct_1m": 0.05},  # qualifies
        {"symbol": "C", "pct_1d": -0.01, "pct_1w": 0.05,  "pct_1m": 0.20},  # 1D negative — skip
        {"symbol": "D", "pct_1d": 0.01,  "pct_1w": -0.01, "pct_1m": 0.30},  # 1W negative — skip
        {"symbol": "E", "pct_1d": 0.01,  "pct_1w": 0.01,  "pct_1m": -0.05}, # 1M negative — skip
        {"symbol": "F", "pct_1d": None,  "pct_1w": 0.01,  "pct_1m": 0.10},  # missing — skip
        {"symbol": "G", "pct_1d": 0.0,   "pct_1w": 0.0,   "pct_1m": 0.07},  # zero 1D/1W is ok (>=0)
    ]
    out = _compute_monthly_winners(records, top_n=10)
    # A (10%), G (7%), B (5%) — order by 1M desc.
    assert [r["symbol"] for r in out] == ["A", "G", "B"]


def test_compute_monthly_winners_respects_top_n():
    records = [
        {"symbol": f"X{i}", "pct_1d": 0.01, "pct_1w": 0.01, "pct_1m": i / 100.0}
        for i in range(1, 11)
    ]
    out = _compute_monthly_winners(records, top_n=3)
    assert [r["symbol"] for r in out] == ["X10", "X9", "X8"]


def test_compute_monthly_losers_filters_and_sorts():
    """Filter: 1M < 0 AND 1D <= 0 AND 1W <= 0. Sort: 1M asc (worst first)."""
    records = [
        {"symbol": "A", "pct_1d": -0.01, "pct_1w": -0.02, "pct_1m": -0.10},  # qualifies
        {"symbol": "B", "pct_1d": -0.01, "pct_1w": -0.02, "pct_1m": -0.05},  # qualifies, less bad
        {"symbol": "C", "pct_1d": 0.01,  "pct_1w": -0.05, "pct_1m": -0.20},  # 1D positive — skip (bounce)
        {"symbol": "D", "pct_1d": -0.01, "pct_1w": 0.01,  "pct_1m": -0.30},  # 1W positive — skip
        {"symbol": "E", "pct_1d": -0.01, "pct_1w": -0.01, "pct_1m": 0.05},   # 1M positive — skip
        {"symbol": "F", "pct_1d": None,  "pct_1w": -0.01, "pct_1m": -0.10},  # missing — skip
        {"symbol": "G", "pct_1d": 0.0,   "pct_1w": 0.0,   "pct_1m": -0.07},  # zero 1D/1W ok (<=0)
    ]
    out = _compute_monthly_losers(records, top_n=10)
    # A (-10%), G (-7%), B (-5%) — ascending = most negative first.
    assert [r["symbol"] for r in out] == ["A", "G", "B"]


def test_ranking_headers_includes_owned_and_required_columns():
    for col in ("Symbol", "Owned?", "Today (EUR)", "1D %", "1W %", "1M %"):
        assert col in RANKING_HEADERS


def test_migrate_monthly_movers_renames_legacy_sheet(tmp_path):
    """Workbook with old 'Monthly movers' sheet gets it renamed in place,
    its content preserved."""
    from openpyxl import Workbook, load_workbook
    path = tmp_path / "legacy.xlsx"
    wb = Workbook()
    legacy = wb.active
    legacy.title = MONTHLY_MOVERS_LEGACY_NAME
    legacy["A1"] = "had data"
    wb.save(path)
    wb = load_workbook(path)
    _migrate_monthly_movers_to_winners_openpyxl(wb)
    wb.save(path)
    wb2 = load_workbook(path)
    assert MONTHLY_MOVERS_LEGACY_NAME not in wb2.sheetnames
    assert MONTHLY_WINNERS_SHEET_NAME in wb2.sheetnames
    assert wb2[MONTHLY_WINNERS_SHEET_NAME]["A1"].value == "had data"


def test_fetch_all_info_should_stop_breaks_early(monkeypatch):
    """should_stop callback short-circuits the loop. Returns partial dict —
    len(out) < len(tickers) signals the caller that the run was interrupted."""
    import stocks_report as sr
    monkeypatch.setattr(sr, "fetch_ticker_info", lambda t, session=None: {"longName": t})
    monkeypatch.setattr(sr.time, "sleep", lambda s: None)
    tickers = [f"T{i}" for i in range(100)]
    # Stop after the first poll (at i==1).
    out = sr.fetch_all_info(tickers, delay=0.0, should_stop=lambda: True, stop_poll_every=25)
    assert len(out) < len(tickers), "fetch_all_info should break early on stop"


def test_fetch_all_info_should_stop_lets_loop_finish_when_false(monkeypatch):
    import stocks_report as sr
    monkeypatch.setattr(sr, "fetch_ticker_info", lambda t, session=None: {"longName": t})
    monkeypatch.setattr(sr.time, "sleep", lambda s: None)
    tickers = [f"T{i}" for i in range(50)]
    out = sr.fetch_all_info(tickers, delay=0.0, should_stop=lambda: False)
    assert len(out) == len(tickers)


def test_portfolio_symbol_resolves_exact_and_suffix():
    """The match logic accepts exact, suffix-stripped, and root-with-suffix forms."""
    import stocks_report as sr
    market_upper = {"AAPL", "KBC.BR", "SAP.DE"}
    market_roots = {"AAPL", "KBC", "SAP"}
    assert sr._portfolio_symbol_resolves("AAPL", market_upper, market_roots)
    assert sr._portfolio_symbol_resolves("KBC", market_upper, market_roots)
    assert sr._portfolio_symbol_resolves("KBC.BR", market_upper, market_roots)
    assert not sr._portfolio_symbol_resolves("TYPO", market_upper, market_roots)


def test_resolve_or_adopt_clears_error_when_symbol_already_in_market(monkeypatch):
    """If a portfolio entry is already a constituent, write_error(row, None)
    clears any stale error from a previous run."""
    import stocks_report as sr
    monkeypatch.setattr(sr, "_yahoo_lookup_for_adoption",
                        lambda s, session=None: pytest.fail("must not query Yahoo for resolved symbol"))
    constituents = pd.DataFrame([{"Symbol": "AAPL", "Name": "Apple", "Indexes": "SP500"}])
    written: list[tuple[int, object]] = []
    out = sr.resolve_or_adopt_portfolio(
        [(16, "AAPL")], constituents, lambda r, m: written.append((r, m)),
    )
    assert len(out) == 1  # no synthetic row added
    assert written == [(16, None)]


def test_resolve_or_adopt_adopts_when_yahoo_knows_ticker(monkeypatch):
    """An unresolved symbol that Yahoo knows becomes a synthetic row with
    Indexes='Portfolio'; error is cleared."""
    import stocks_report as sr
    monkeypatch.setattr(sr, "_yahoo_lookup_for_adoption",
                        lambda s, session=None: {"Name": f"Long {s}", "currency": "EUR"})
    constituents = pd.DataFrame([{"Symbol": "AAPL", "Name": "Apple", "Indexes": "SP500"}])
    written: list[tuple[int, object]] = []
    out = sr.resolve_or_adopt_portfolio(
        [(17, "CSKRL.XC")], constituents, lambda r, m: written.append((r, m)),
    )
    assert "CSKRL.XC" in set(out["Symbol"])
    adopted = out[out["Symbol"] == "CSKRL.XC"].iloc[0]
    assert adopted["Indexes"] == "Portfolio"
    assert adopted["Name"] == "Long CSKRL.XC"
    assert written == [(17, None)]


def test_resolve_or_adopt_writes_error_when_yahoo_fails(monkeypatch):
    """An unresolved symbol that Yahoo can't find: error written to row,
    constituents unchanged, no raise."""
    import stocks_report as sr
    monkeypatch.setattr(sr, "_yahoo_lookup_for_adoption", lambda s, session=None: None)
    constituents = pd.DataFrame([{"Symbol": "AAPL", "Name": "Apple", "Indexes": "SP500"}])
    written: list[tuple[int, object]] = []
    out = sr.resolve_or_adopt_portfolio(
        [(18, "TYPO")], constituents, lambda r, m: written.append((r, m)),
    )
    assert "TYPO" not in set(out["Symbol"])
    assert len(written) == 1 and written[0][0] == 18
    assert "Yahoo" in written[0][1]
    assert "TYPO" in written[0][1]


def test_resolve_or_adopt_empty_portfolio_ok():
    """Empty portfolio is a no-op on an empty market."""
    import stocks_report as sr
    constituents = pd.DataFrame(columns=["Symbol", "Name", "Indexes"])
    out = sr.resolve_or_adopt_portfolio([], constituents, lambda r, m: None)
    assert len(out) == 0


def test_resolve_or_adopt_does_not_double_adopt_same_ticker(monkeypatch):
    """Two portfolio rows pointing at the same unresolved ticker yield one
    synthetic constituent, not two."""
    import stocks_report as sr
    monkeypatch.setattr(sr, "_yahoo_lookup_for_adoption",
                        lambda s, session=None: {"Name": s, "currency": ""})
    constituents = pd.DataFrame([{"Symbol": "AAPL", "Name": "Apple", "Indexes": "SP500"}])
    out = sr.resolve_or_adopt_portfolio(
        [(16, "FOO.X"), (17, "FOO.X")], constituents, lambda r, m: None,
    )
    foo_rows = out[out["Symbol"] == "FOO.X"]
    assert len(foo_rows) == 1


def test_read_portfolio_entries_preserves_row_indices():
    """read_portfolio_entries returns (row_idx, SYMBOL) tuples for error reporting."""
    from openpyxl import Workbook
    import stocks_report as sr
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws.cell(row=16, column=1, value="AAPL")
    ws.cell(row=17, column=1, value=None)  # gap
    ws.cell(row=18, column=1, value="kbc.br")  # mixed case → uppercased
    out = sr.read_portfolio_entries(ws)
    assert out == [(16, "AAPL"), (18, "KBC.BR")]


def test_read_stop_requested_interprets_cell(tmp_path):
    """read_stop_requested mirrors read_test_mode's tolerance to common
    truthy/falsy values."""
    from openpyxl import Workbook
    import stocks_report as sr
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws[MAIN_CELLS["StopRequested"]] = None
    assert sr.read_stop_requested(ws) is False
    ws[MAIN_CELLS["StopRequested"]] = "FALSE"
    assert sr.read_stop_requested(ws) is False
    ws[MAIN_CELLS["StopRequested"]] = "TRUE"
    assert sr.read_stop_requested(ws) is True
    ws[MAIN_CELLS["StopRequested"]] = True
    assert sr.read_stop_requested(ws) is True


def test_get_quotes_test_mode_populates_ranking_on_unfilled_market(tmp_path, monkeypatch):
    """Regression for 2026-05-13 (round 2): Monthly winners + losers were STILL
    empty in test mode because:
      - rebuild_inventory --test had previously wiped Market down to 5 rows
        with NO quote data (only structural).
      - get_quotes --test refreshed only 1 of those 5 → 4 rows with None
        for the % columns → filtered out of the ranking → empty winners/losers.

    The fix: test-mode get_quotes refreshes TOP N (=20) rows, not just 1.
    Combined with a Market of at least a few rows (BEL20 + ETFs gives ~58),
    the ranking has enough populated %s to produce a meaningful report.

    This test deliberately seeds Market with rows that have NO %s (matching
    the realistic state after a test-mode rebuild), runs get_quotes test
    mode against it, and asserts the ranking ends up populated.
    """
    import datetime as dt
    import pandas as pd
    import stocks_report as sr
    from openpyxl import load_workbook

    path = tmp_path / "ranking_smoke.xlsx"
    init_workbook(path)
    wb = load_workbook(path)
    market = wb["Market"]
    cols = {n: MARKET_COLUMNS.index(n) + 1 for n in MARKET_COLUMNS}

    # Seed 30 Market rows with structural data only — no % columns
    # populated. This mimics the state right after a test-mode rebuild
    # before any get_quotes has run.
    seed_count = 30
    for r in range(2, 2 + seed_count):
        sym = f"FAKE{r-2:02d}.X"
        market.cell(row=r, column=cols["Symbol"], value=sym)
        market.cell(row=r, column=cols["Name"], value=f"Fake {r-2}")
        market.cell(row=r, column=cols["Currency"], value="EUR")
        market.cell(row=r, column=cols["Owned?"], value="No")
        market.cell(row=r, column=cols["Indexes"], value="BEL20")
        # NOTE: deliberately no Today (EUR) / 1D % / etc. — they're None.
    wb.save(path)

    # Mock the network layer. fetch_close_prices returns deterministic
    # series that produce a 50/50 mix of winners and losers under to_eur.
    today = dt.date.today()
    dates = pd.date_range(end=today, periods=400).normalize()

    def mock_close_prices(tickers, *args, **kwargs):
        data = {}
        for i, t in enumerate(tickers):
            base = 100.0
            today_price = base * (1.15 if i % 2 == 0 else 0.85)  # winners vs losers
            series = pd.Series(
                [base] * (len(dates) - 1) + [today_price],
                index=dates,
            )
            data[t] = series
        return pd.concat(data, axis=1)

    monkeypatch.setattr(sr, "fetch_close_prices", mock_close_prices)
    monkeypatch.setattr(sr, "fetch_all_info", lambda tickers, **kw:
                        {t: {"trailingPE": 15.0, "forwardPE": 14.0} for t in tickers})
    monkeypatch.setattr(sr, "get_fx_rates", lambda: {
        "EUR": 1.0, "USD": 1.1, "JPY": 165.0, "GBP": 0.85, "CHF": 1.0,
        "GBp": 0.85,
    })
    monkeypatch.setattr(sr, "get_fx_history",
                        lambda: {ccy: pd.Series(dtype=float)
                                 for ccy in ("USD", "JPY", "GBP", "CHF")})

    # Run get_quotes in test mode
    result = sr.get_quotes(workbook_path=path, test_mode=True, info_delay=0.0)
    assert result == 0, "get_quotes test mode returned non-zero"

    wb2 = load_workbook(path)

    # The first TEST_MODE_QUOTE_REFRESH_LIMIT rows must now have populated %s
    refreshed_pct_count = 0
    for r in range(2, 2 + min(seed_count, sr.TEST_MODE_QUOTE_REFRESH_LIMIT)):
        if wb2["Market"].cell(row=r, column=cols["1M %"]).value is not None:
            refreshed_pct_count += 1
    assert refreshed_pct_count >= sr.TEST_MODE_QUOTE_REFRESH_LIMIT - 2, (
        f"Only {refreshed_pct_count} rows have 1M %; expected ~{sr.TEST_MODE_QUOTE_REFRESH_LIMIT}. "
        "Test mode is refreshing too few rows."
    )

    # The ranking sheets must exist AND have at least one data row each.
    from stocks_report import MONTHLY_WINNERS_SHEET_NAME, MONTHLY_LOSERS_SHEET_NAME
    assert MONTHLY_WINNERS_SHEET_NAME in wb2.sheetnames
    assert MONTHLY_LOSERS_SHEET_NAME in wb2.sheetnames
    winners = wb2[MONTHLY_WINNERS_SHEET_NAME]
    losers = wb2[MONTHLY_LOSERS_SHEET_NAME]
    assert winners.cell(row=2, column=1).value is not None, (
        "Monthly winners is EMPTY after get_quotes test mode — this is the user's bug"
    )
    assert losers.cell(row=2, column=1).value is not None, (
        "Monthly losers is EMPTY after get_quotes test mode — this is the user's bug"
    )

    # In this seed scenario, NO row is marked Owned?=Yes, so the losers
    # sheet's pre-filter must be UNAPPLIED (otherwise Excel would hide
    # every loser row and the user would see the sheet as empty — the
    # actual 2026-05-13 bug).
    los_filter = losers.auto_filter
    assert not los_filter.filterColumn, (
        "Owned?=Yes pre-filter is active but no row is Owned?=Yes — "
        "Excel will hide every loser and the sheet will look empty"
    )


def test_get_quotes_test_mode_skips_owned_filter_when_no_owned_losers(tmp_path, monkeypatch):
    """The Owned?=Yes pre-filter on Monthly losers must adapt: if no owned
    losers exist, skip the pre-filter so the user actually sees the rows.
    """
    import datetime as dt
    import pandas as pd
    import stocks_report as sr
    from openpyxl import load_workbook
    from stocks_report import MONTHLY_LOSERS_SHEET_NAME

    path = tmp_path / "adaptive_filter.xlsx"
    init_workbook(path)
    wb = load_workbook(path)
    market = wb["Market"]
    cols = {n: MARKET_COLUMNS.index(n) + 1 for n in MARKET_COLUMNS}
    # 10 rows, all Owned?=No, all losers.
    for r in range(2, 12):
        sym = f"LOSER{r-2}.X"
        market.cell(row=r, column=cols["Symbol"], value=sym)
        market.cell(row=r, column=cols["Name"], value=f"Loser {r-2}")
        market.cell(row=r, column=cols["Currency"], value="EUR")
        market.cell(row=r, column=cols["Owned?"], value="No")
        market.cell(row=r, column=cols["Indexes"], value="X")
        market.cell(row=r, column=cols["1D %"], value=-0.01)
        market.cell(row=r, column=cols["1W %"], value=-0.02)
        market.cell(row=r, column=cols["1M %"], value=-0.05 * (r - 1))
        market.cell(row=r, column=cols["Today (EUR)"], value=100.0)
    wb.save(path)

    # Mock network
    dates = pd.date_range(end=dt.date.today(), periods=400).normalize()
    monkeypatch.setattr(sr, "fetch_close_prices",
        lambda tickers, *a, **kw: pd.concat({t: pd.Series([100]*399 + [95], index=dates) for t in tickers}, axis=1))
    monkeypatch.setattr(sr, "fetch_all_info", lambda tickers, **kw: {t: {} for t in tickers})
    monkeypatch.setattr(sr, "get_fx_rates", lambda: {"EUR": 1.0, "GBp": 0.85, "USD": 1.1, "JPY": 165, "GBP": 0.85, "CHF": 1.0})
    monkeypatch.setattr(sr, "get_fx_history",
        lambda: {ccy: pd.Series(dtype=float) for ccy in ("USD","JPY","GBP","CHF")})

    sr.get_quotes(workbook_path=path, test_mode=True, info_delay=0.0)

    wb2 = load_workbook(path)
    losers = wb2[MONTHLY_LOSERS_SHEET_NAME]
    # losers ARE present (refresh produced declining stocks)
    assert losers.cell(row=2, column=1).value is not None
    # but pre-filter must be UNAPPLIED since no losers are owned.
    assert not losers.auto_filter.filterColumn, (
        "Pre-filter must NOT be applied when no owned losers exist; "
        "otherwise Excel hides every row and the user sees empty"
    )


def test_get_quotes_test_mode_applies_owned_filter_when_owned_losers_present(tmp_path, monkeypatch):
    """Conversely: WHEN at least one loser is owned, the pre-filter SHOULD
    apply so the user immediately sees their declining holdings first."""
    import datetime as dt
    import pandas as pd
    import stocks_report as sr
    from openpyxl import load_workbook
    from stocks_report import MONTHLY_LOSERS_SHEET_NAME, OWNED_COL_INDEX_1BASED

    path = tmp_path / "applied_filter.xlsx"
    init_workbook(path)
    wb = load_workbook(path)
    market = wb["Market"]
    cols = {n: MARKET_COLUMNS.index(n) + 1 for n in MARKET_COLUMNS}
    for r in range(2, 12):
        sym = f"LOSER{r-2}.X"
        market.cell(row=r, column=cols["Symbol"], value=sym)
        market.cell(row=r, column=cols["Name"], value=f"Loser {r-2}")
        market.cell(row=r, column=cols["Currency"], value="EUR")
        # First two rows are owned
        market.cell(row=r, column=cols["Owned?"], value=("Yes" if r <= 3 else "No"))
        market.cell(row=r, column=cols["Indexes"], value="X")
        market.cell(row=r, column=cols["1D %"], value=-0.01)
        market.cell(row=r, column=cols["1W %"], value=-0.02)
        market.cell(row=r, column=cols["1M %"], value=-0.05 * (r - 1))
        market.cell(row=r, column=cols["Today (EUR)"], value=100.0)
    wb.save(path)

    dates = pd.date_range(end=dt.date.today(), periods=400).normalize()
    monkeypatch.setattr(sr, "fetch_close_prices",
        lambda tickers, *a, **kw: pd.concat({t: pd.Series([100]*399 + [95], index=dates) for t in tickers}, axis=1))
    monkeypatch.setattr(sr, "fetch_all_info", lambda tickers, **kw: {t: {} for t in tickers})
    monkeypatch.setattr(sr, "get_fx_rates", lambda: {"EUR": 1.0, "GBp": 0.85, "USD": 1.1, "JPY": 165, "GBP": 0.85, "CHF": 1.0})
    monkeypatch.setattr(sr, "get_fx_history",
        lambda: {ccy: pd.Series(dtype=float) for ccy in ("USD","JPY","GBP","CHF")})

    sr.get_quotes(workbook_path=path, test_mode=True, info_delay=0.0)

    wb2 = load_workbook(path)
    losers = wb2[MONTHLY_LOSERS_SHEET_NAME]
    # Pre-filter must be applied on Owned? column
    assert losers.auto_filter.filterColumn, (
        "Pre-filter must be applied when at least one owned loser exists"
    )
    fc = losers.auto_filter.filterColumn[0]
    assert fc.colId == OWNED_COL_INDEX_1BASED - 1  # 0-based
    assert "Yes" in fc.filters.filter


def test_rebuild_inventory_test_mode_preserves_market(tmp_path, monkeypatch):
    """Regression for 2026-05-13 'Monthly winners/losers empty in test mode'.

    rebuild_inventory --test used to wipe Market down to 5 rows, which left
    the subsequent ranking computation (which reads ALL Market rows) with
    nothing to rank. The fix: in test mode, preserve Market and only update
    structural columns of rows that match the test constituents by Symbol.
    """
    import pandas as pd
    import stocks_report as sr
    from openpyxl import load_workbook

    # Build a workbook with a Market sheet that has 50 rows of stale-but-
    # populated data, including the BEL20 rows the test mode will try to update.
    path = tmp_path / "preserve_test.xlsx"
    init_workbook(path)
    wb = load_workbook(path)
    market = wb["Market"]
    cols = {n: MARKET_COLUMNS.index(n) + 1 for n in MARKET_COLUMNS}
    seed_symbols = ["ABI.BR", "ACKB.BR", "AED.BR", "AGS.BR", "ARGX.BR"] + [
        f"FAKE{i}" for i in range(45)
    ]
    for i, sym in enumerate(seed_symbols, 2):
        market.cell(row=i, column=cols["Symbol"], value=sym)
        market.cell(row=i, column=cols["Name"], value=f"stale name for {sym}")
        market.cell(row=i, column=cols["1D %"], value=0.01)
        market.cell(row=i, column=cols["1W %"], value=0.02)
        market.cell(row=i, column=cols["1M %"], value=0.05)
    wb.save(path)

    # Stub the network-touching helpers so the test doesn't hit Wikipedia / Yahoo.
    fake_constituents = pd.DataFrame(
        [(s, f"refreshed {s}", "BEL20") for s in seed_symbols[:5]],
        columns=["Symbol", "Name", "Index"],
    )
    monkeypatch.setattr(sr, "get_index_constituents", lambda idx: fake_constituents)
    monkeypatch.setattr(sr, "get_etfs", lambda: pd.DataFrame(columns=["Symbol", "Name", "Index"]))
    monkeypatch.setattr(sr, "fetch_all_watchlists", lambda: {})
    monkeypatch.setattr(sr, "fetch_all_info", lambda tickers, **kwargs: {
        t: {"currency": "EUR", "longName": f"refreshed {t}", "sector": "TestSector"} for t in tickers
    })

    # Run rebuild in test mode.
    sr.rebuild_inventory(workbook_path=path, test_mode=True, info_delay=0.0)

    # Market must still have all 50 seed rows.
    wb2 = load_workbook(path)
    market2 = wb2["Market"]
    syms_after = [market2.cell(row=r, column=cols["Symbol"]).value
                  for r in range(2, market2.max_row + 1)]
    syms_after = [s for s in syms_after if s]  # drop blanks if any
    assert len(syms_after) == 50, (
        f"test mode wiped Market down to {len(syms_after)} rows — must preserve all 50"
    )

    # The 5 BEL 20 rows must have refreshed Names.
    for sym in seed_symbols[:5]:
        row = syms_after.index(sym) + 2
        assert market2.cell(row=row, column=cols["Name"]).value == f"refreshed {sym}", (
            f"test mode didn't update structural data for {sym}"
        )

    # The 45 non-BEL20 stale rows must still have their stale data — % columns
    # untouched so the ranking computation has something to work with.
    fake_row = syms_after.index("FAKE0") + 2
    assert market2.cell(row=fake_row, column=cols["1M %"]).value == 0.05
    assert market2.cell(row=fake_row, column=cols["Name"]).value == "stale name for FAKE0"


def test_migrate_monthly_movers_no_op_when_winners_already_exists(tmp_path):
    """If both legacy and new names exist, migration must NOT clobber the
    new sheet — leave the legacy one alone for the user to delete."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.active.title = MONTHLY_MOVERS_LEGACY_NAME
    wb.create_sheet(MONTHLY_WINNERS_SHEET_NAME)
    _migrate_monthly_movers_to_winners_openpyxl(wb)
    assert MONTHLY_MOVERS_LEGACY_NAME in wb.sheetnames
    assert MONTHLY_WINNERS_SHEET_NAME in wb.sheetnames


def test_init_workbook_portfolio_header_has_no_quantity(tmp_path):
    """Portfolio header should be just Symbol / Notes (Quantity removed)."""
    from openpyxl import load_workbook
    from stocks_report import PORTFOLIO_HEADER_ROW
    path = tmp_path / "stocks.xlsx"
    init_workbook(path)
    wb = load_workbook(path)
    main = wb["Main"]
    assert main.cell(row=PORTFOLIO_HEADER_ROW, column=1).value == "Symbol"
    assert main.cell(row=PORTFOLIO_HEADER_ROW, column=2).value == "Notes"
    # Column 3 should now be blank (Quantity was removed)
    assert main.cell(row=PORTFOLIO_HEADER_ROW, column=3).value is None


# ---------------------------------------------------------------------------
# _owned_for & read_portfolio_symbols
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("symbol,portfolio,expected", [
    ("KBC.BR",  {"KBC.BR"},        "Yes"),
    ("KBC.BR",  {"KBC"},           "Yes"),   # suffix-stripped match
    ("AAPL",    {"AAPL"},          "Yes"),
    ("AAPL",    {"GOOGL"},         "No"),
    ("BRK-B",   {"BRK-B"},         "Yes"),   # hyphen form
    ("KBC.BR",  set(),             "No"),
])
def test_owned_for(symbol, portfolio, expected):
    assert _owned_for(symbol, portfolio) == expected


def test_read_portfolio_symbols_skips_blanks_and_normalizes(tmp_path):
    """Reads col A of the portfolio range, uppercase + strip, skip None/empty."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws.cell(row=PORTFOLIO_FIRST_ROW,     column=1, value="  kbc.br  ")  # whitespace + lower
    ws.cell(row=PORTFOLIO_FIRST_ROW + 1, column=1, value=None)            # blank
    ws.cell(row=PORTFOLIO_FIRST_ROW + 2, column=1, value="AAPL")
    ws.cell(row=PORTFOLIO_FIRST_ROW + 3, column=1, value="")              # empty string
    assert read_portfolio_symbols(ws) == {"KBC.BR", "AAPL"}


# ---------------------------------------------------------------------------
# _pick_test_target — get-quotes test-mode selection
# ---------------------------------------------------------------------------

def _market(rows):
    """rows: list of (row_idx, symbol, currency, indexes_csv)"""
    return rows


def test_pick_test_target_prefers_portfolio_symbol_in_market():
    rows = _market([(2, "ABI.BR", "EUR", "BEL20"), (3, "AAPL", "USD", "SP500")])
    assert _pick_test_target(rows, {"AAPL"})[1] == "AAPL"


def test_pick_test_target_handles_suffix_stripped_portfolio_entry():
    """User wrote 'KBC' in Main; Market has 'KBC.BR' — should still match."""
    rows = _market([(2, "ABI.BR", "EUR", "BEL20"), (3, "KBC.BR", "EUR", "BEL20")])
    assert _pick_test_target(rows, {"KBC"})[1] == "KBC.BR"


def test_pick_test_target_falls_back_to_first_bel20():
    rows = _market([(2, "AAPL", "USD", "SP500"), (3, "ABI.BR", "EUR", "BEL20")])
    assert _pick_test_target(rows, set())[1] == "ABI.BR"


def test_pick_test_target_falls_back_to_first_row_if_no_bel20():
    rows = _market([(2, "AAPL", "USD", "SP500"), (3, "GOOGL", "USD", "SP500")])
    assert _pick_test_target(rows, set())[1] == "AAPL"


def test_pick_test_target_returns_none_when_market_empty():
    assert _pick_test_target([], set()) is None


# ---------------------------------------------------------------------------
# read_test_mode — Main!B5 cell interpretation
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("raw,expected", [
    ("TRUE",  True),
    ("True",  True),
    ("true",  True),
    ("Yes",   True),
    ("Y",     True),
    ("1",     True),
    (1,       True),
    (True,    True),
    ("FALSE", False),
    ("No",    False),
    ("0",     False),
    (0,       False),
    (False,   False),
    ("",      False),
    (None,    False),
    ("  TRUE  ", True),  # whitespace tolerated
])
def test_read_test_mode_interpretation(raw, expected):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Main"
    ws[MAIN_CELLS["TestMode"]] = raw
    assert read_test_mode(ws) is expected


def test_read_portfolio_symbols_ignores_main_section_labels(tmp_path):
    """Regression: a fresh init_workbook must have an empty portfolio set.

    Previously the portfolio range overlapped with the Jobs/Status/Metadata
    section labels in column A, so every label was being parsed as a "symbol".
    """
    path = tmp_path / "stocks.xlsx"
    init_workbook(path)
    from openpyxl import load_workbook
    wb = load_workbook(path)
    assert read_portfolio_symbols(wb["Main"]) == set()


# ---------------------------------------------------------------------------
# to_eur
# ---------------------------------------------------------------------------

def test_to_eur_basic_conversions():
    fx = {"USD": 1.07, "JPY": 165.0, "GBP": 0.85, "EUR": 1.0, "GBp": 0.85}
    assert to_eur(100.0, "USD", fx) == pytest.approx(100 / 1.07)
    assert to_eur(1000.0, "JPY", fx) == pytest.approx(1000 / 165.0)
    assert to_eur(50.0, "GBP", fx) == pytest.approx(50 / 0.85)
    assert to_eur(50.0, "EUR", fx) == 50.0


def test_to_eur_pence():
    """GBp = pence; convert to pounds first then to EUR."""
    fx = {"GBP": 0.85, "GBp": 0.85}
    # 200 pence = £2 = 2/0.85 EUR
    assert to_eur(200.0, "GBp", fx) == pytest.approx(2.0 / 0.85)


def test_to_eur_handles_none_price():
    assert to_eur(None, "USD", {"USD": 1.07}) is None


def test_to_eur_handles_missing_rate():
    fx = {"EUR": 1.0}
    assert to_eur(100.0, "JPY", fx) is None


def test_to_eur_handles_nan_rate():
    fx = {"USD": float("nan")}
    assert to_eur(100.0, "USD", fx) is None


# ---------------------------------------------------------------------------
# price_at_or_before
# ---------------------------------------------------------------------------

def _make_series():
    idx = pd.to_datetime([
        "2026-04-27", "2026-04-28", "2026-04-29",
        "2026-04-30", "2026-05-01", "2026-05-04",
    ])
    return pd.Series([100, 101, 102, 103, 104, 107], index=idx)


def test_price_at_or_before_weekend_rollback():
    s = _make_series()
    # Sat 2026-05-02 → fall back to Fri May 1 (104)
    assert price_at_or_before(s, dt.date(2026, 5, 2)) == 104.0
    # Sun 2026-05-03 → also fall back to Fri May 1
    assert price_at_or_before(s, dt.date(2026, 5, 3)) == 104.0


def test_price_at_or_before_exact_match():
    s = _make_series()
    assert price_at_or_before(s, dt.date(2026, 5, 4)) == 107.0
    assert price_at_or_before(s, dt.date(2026, 4, 27)) == 100.0


def test_price_at_or_before_target_after_history():
    s = _make_series()
    # If target is in the future, just return the last known close
    assert price_at_or_before(s, dt.date(2030, 1, 1)) == 107.0


def test_price_at_or_before_target_before_history():
    s = _make_series()
    assert price_at_or_before(s, dt.date(2020, 1, 1)) is None


def test_price_at_or_before_empty_series():
    assert price_at_or_before(pd.Series(dtype=float), dt.date(2026, 5, 4)) is None


def test_price_at_or_before_none_series():
    assert price_at_or_before(None, dt.date(2026, 5, 4)) is None


# ---------------------------------------------------------------------------
# Speedups A/C/F
# ---------------------------------------------------------------------------

def test_fetch_all_info_parallel_returns_all_results(monkeypatch):
    """Speedup A: with ThreadPoolExecutor, every ticker still ends up in
    the result dict regardless of completion order."""
    import stocks_report as sr
    monkeypatch.setattr(sr, "fetch_ticker_info",
                        lambda t, session=None: {"longName": t, "sector": "X"})
    tickers = [f"T{i}" for i in range(75)]
    out = sr.fetch_all_info(tickers, max_workers=8)
    assert set(out.keys()) == set(tickers)
    assert all(out[t]["longName"] == t for t in tickers)


def test_fetch_all_info_progress_fires_per_completion(monkeypatch):
    """Speedup A: progress callback fires for each completed ticker — used
    by the workbook Status cell to show live counts."""
    import stocks_report as sr
    monkeypatch.setattr(sr, "fetch_ticker_info", lambda t, session=None: {})
    seen: list[int] = []
    sr.fetch_all_info(
        [f"T{i}" for i in range(20)],
        progress=lambda done, total, sym: seen.append(done),
        max_workers=4,
    )
    assert sorted(seen) == list(range(1, 21))


def test_is_recently_updated_within_threshold():
    """Speedup C: a fresh ISO timestamp is reported as recent."""
    import stocks_report as sr
    now = dt.datetime(2026, 5, 13, 12, 0, 0, tzinfo=dt.UTC)
    iso = (now - dt.timedelta(hours=2)).strftime("%Y-%m-%dT%H:%M:%SZ")
    assert sr.is_recently_updated(iso, threshold_hours=4.0, now=now) is True


def test_is_recently_updated_outside_threshold():
    import stocks_report as sr
    now = dt.datetime(2026, 5, 13, 12, 0, 0, tzinfo=dt.UTC)
    iso = (now - dt.timedelta(hours=5)).strftime("%Y-%m-%dT%H:%M:%SZ")
    assert sr.is_recently_updated(iso, threshold_hours=4.0, now=now) is False


def test_is_recently_updated_empty_or_garbage():
    import stocks_report as sr
    assert sr.is_recently_updated(None, 4.0) is False
    assert sr.is_recently_updated("", 4.0) is False
    assert sr.is_recently_updated("not-a-date", 4.0) is False


def test_info_cache_roundtrip(tmp_path):
    """Speedup F: put_many then get_fresh returns the same payload."""
    import stocks_report as sr
    cache = sr.InfoCache(tmp_path / "cache.sqlite")
    try:
        cache.put_many({
            "AAPL": {"currency": "USD", "trailingPE": 30.0, "forwardPE": 25.0,
                     "longName": "Apple Inc.", "sector": "Technology",
                     "description": "Designs phones."},
        })
        got = cache.get_fresh(["AAPL", "MSFT"], ttl_seconds=86400)
    finally:
        cache.close()
    assert "MSFT" not in got
    assert got["AAPL"]["longName"] == "Apple Inc."
    assert got["AAPL"]["trailingPE"] == 30.0


def test_info_cache_respects_ttl(tmp_path):
    """Speedup F: entries older than TTL are not returned."""
    import stocks_report as sr
    cache = sr.InfoCache(tmp_path / "cache.sqlite")
    try:
        cache.put_many({"OLD": {"currency": "EUR"}})
        # TTL = 0 seconds means everything is stale.
        got = cache.get_fresh(["OLD"], ttl_seconds=0)
    finally:
        cache.close()
    assert got == {}


def test_fetch_all_info_with_cache_uses_cache(tmp_path, monkeypatch):
    """Speedup F: when cache covers a ticker, fetch_ticker_info is NOT called for it."""
    import stocks_report as sr
    # Pre-warm the cache for AAPL.
    cache = sr.InfoCache(tmp_path / "cache.sqlite")
    try:
        cache.put_many({"AAPL": {"currency": "USD", "longName": "Apple Inc."}})
    finally:
        cache.close()
    call_log: list[str] = []
    def fake_fetch(t, session=None):
        call_log.append(t)
        return {"currency": "??", "longName": f"FRESH-{t}"}
    monkeypatch.setattr(sr, "fetch_ticker_info", fake_fetch)
    info_map, hits = sr.fetch_all_info_with_cache(
        ["AAPL", "MSFT"], cache_path=tmp_path / "cache.sqlite",
        ttl_seconds=86400, max_workers=2,
    )
    assert hits == 1
    assert call_log == ["MSFT"], "AAPL must NOT trigger a fresh fetch"
    assert info_map["AAPL"]["longName"] == "Apple Inc."
    assert info_map["MSFT"]["longName"] == "FRESH-MSFT"


def test_fetch_all_info_with_cache_persists_misses(tmp_path, monkeypatch):
    """Speedup F: fresh fetches get persisted so the next call hits cache."""
    import stocks_report as sr
    monkeypatch.setattr(sr, "fetch_ticker_info",
                        lambda t, session=None: {"longName": f"persisted-{t}"})
    cache_path = tmp_path / "cache.sqlite"
    sr.fetch_all_info_with_cache(["A", "B"], cache_path=cache_path, ttl_seconds=86400)
    # Second call should be a 100% cache hit; fetch_ticker_info MUST NOT fire.
    monkeypatch.setattr(sr, "fetch_ticker_info",
                        lambda t, session=None: pytest.fail(f"unexpected fetch of {t}"))
    info_map, hits = sr.fetch_all_info_with_cache(
        ["A", "B"], cache_path=cache_path, ttl_seconds=86400,
    )
    assert hits == 2
    assert info_map["A"]["longName"] == "persisted-A"
    assert info_map["B"]["longName"] == "persisted-B"


def test_info_cache_path_lives_next_to_workbook(tmp_path):
    """Cache file lives in the workbook's directory, not cwd."""
    import stocks_report as sr
    wb = tmp_path / "deep" / "stocks_picker_v03.xlsm"
    wb.parent.mkdir(parents=True)
    assert sr._info_cache_path_for(wb).parent == wb.parent
    assert sr._info_cache_path_for(None).is_absolute()


# ---------------------------------------------------------------------------
# Batch 1 speedups: shared Yahoo session, retry-on-401, parallel watchlists + prices
# ---------------------------------------------------------------------------

def test_fetch_ticker_info_retries_when_first_call_returns_empty(monkeypatch):
    """Symptom of a transient 401 Invalid Crumb is t.info returning {} —
    we retry once before accepting that result."""
    import stocks_report as sr
    calls = {"n": 0}

    class _FakeTicker:
        def __init__(self, sym, session=None): self.sym = sym
        @property
        def info(self):
            calls["n"] += 1
            if calls["n"] == 1:
                return {}  # simulate 401 → empty
            return {"longName": "Apple Inc.", "currency": "USD",
                    "trailingPE": 30.0, "forwardPE": 25.0,
                    "sector": "Technology", "longBusinessSummary": "Phones"}

    monkeypatch.setattr(sr.yf, "Ticker", _FakeTicker)
    monkeypatch.setattr(sr.time, "sleep", lambda s: None)
    out = sr.fetch_ticker_info("AAPL")
    assert calls["n"] == 2, "should have retried once on empty"
    assert out["longName"] == "Apple Inc."
    assert out["trailingPE"] == 30.0


def test_fetch_ticker_info_no_retry_when_first_call_has_data(monkeypatch):
    """Successful first call must NOT trigger a retry — keep the latency
    overhead off the hot path."""
    import stocks_report as sr
    calls = {"n": 0}

    class _FakeTicker:
        def __init__(self, sym, session=None): pass
        @property
        def info(self):
            calls["n"] += 1
            return {"longName": "X", "currency": "EUR"}

    monkeypatch.setattr(sr.yf, "Ticker", _FakeTicker)
    sr.fetch_ticker_info("X")
    assert calls["n"] == 1


def test_fetch_ticker_info_accepts_empty_after_retry(monkeypatch):
    """Delisted/invalid tickers return empty even on retry — we accept that
    rather than spinning forever."""
    import stocks_report as sr

    class _FakeTicker:
        def __init__(self, sym, session=None): pass
        @property
        def info(self): return {}

    monkeypatch.setattr(sr.yf, "Ticker", _FakeTicker)
    monkeypatch.setattr(sr.time, "sleep", lambda s: None)
    out = sr.fetch_ticker_info("DEADTKR")
    # All fields default to empty/None — no crash.
    assert out["longName"] == ""
    assert out["trailingPE"] is None


def test_fetch_all_info_passes_session_to_ticker(monkeypatch):
    """The shared session must reach the underlying yf.Ticker constructor —
    otherwise the 401 fix is defeated."""
    import stocks_report as sr
    seen_sessions: list = []

    class _FakeTicker:
        def __init__(self, sym, session=None):
            seen_sessions.append(session)
        @property
        def info(self):
            return {"longName": "ok", "currency": "USD"}

    monkeypatch.setattr(sr.yf, "Ticker", _FakeTicker)
    # Skip the real cookie warmup
    monkeypatch.setattr(sr, "get_yahoo_session", lambda: "SHARED_SENTINEL")
    sr.fetch_all_info(["A", "B", "C"], max_workers=2)
    # Every fetch_ticker_info call should have received the shared session.
    assert seen_sessions and all(s == "SHARED_SENTINEL" for s in seen_sessions)


def test_fetch_close_prices_runs_chunks_in_parallel(monkeypatch):
    """Speedup: at least 2 chunks should be in flight at the same time."""
    import stocks_report as sr
    in_flight = {"max": 0, "current": 0}
    lock = __import__("threading").Lock()

    def fake_download(tickers, start, end, **kw):
        with lock:
            in_flight["current"] += 1
            in_flight["max"] = max(in_flight["max"], in_flight["current"])
        import time as _t
        _t.sleep(0.1)
        with lock:
            in_flight["current"] -= 1
        # Build a multi-ticker close DataFrame so the slicing path works
        idx = pd.date_range("2026-05-01", periods=3, freq="D")
        df = pd.DataFrame({t: [1.0, 2.0, 3.0] for t in tickers}, index=idx)
        return pd.concat({"Close": df}, axis=1)  # MultiIndex columns

    monkeypatch.setattr(sr.yf, "download", fake_download)
    tickers = [f"T{i}" for i in range(30)]
    out = sr.fetch_close_prices(tickers, dt.date(2026, 5, 1), dt.date(2026, 5, 4),
                                chunk_size=10, max_workers=4)
    assert in_flight["max"] >= 2, "should have overlapped chunks"
    assert not out.empty


def test_fetch_close_prices_passes_session(monkeypatch):
    """Shared session flows into yf.download so the chunk fetcher reuses cookies."""
    import stocks_report as sr
    seen = {"sessions": []}

    def fake_download(tickers, start, end, **kw):
        seen["sessions"].append(kw.get("session"))
        idx = pd.date_range("2026-05-01", periods=2, freq="D")
        df = pd.DataFrame({t: [1.0, 2.0] for t in tickers}, index=idx)
        return pd.concat({"Close": df}, axis=1)

    monkeypatch.setattr(sr.yf, "download", fake_download)
    sentinel = object()
    sr.fetch_close_prices(["A", "B", "C", "D"], dt.date(2026, 5, 1), dt.date(2026, 5, 3),
                          chunk_size=2, max_workers=2, session=sentinel)
    assert all(s is sentinel for s in seen["sessions"])


def test_fetch_all_watchlists_runs_in_parallel(monkeypatch):
    """The 6 watchlists should fire concurrently, not serially."""
    import stocks_report as sr
    in_flight = {"max": 0, "current": 0}
    lock = __import__("threading").Lock()

    def slow_screener(scrid, sess, crumb):
        with lock:
            in_flight["current"] += 1
            in_flight["max"] = max(in_flight["max"], in_flight["current"])
        import time as _t
        _t.sleep(0.05)
        with lock:
            in_flight["current"] -= 1
        return {scrid.upper()}

    def slow_dataroma(path):
        with lock:
            in_flight["current"] += 1
            in_flight["max"] = max(in_flight["max"], in_flight["current"])
        import time as _t
        _t.sleep(0.05)
        with lock:
            in_flight["current"] -= 1
        return {"DR-" + path[-3:]}

    def slow_activists():
        with lock:
            in_flight["current"] += 1
            in_flight["max"] = max(in_flight["max"], in_flight["current"])
        import time as _t
        _t.sleep(0.05)
        with lock:
            in_flight["current"] -= 1
        return {"ACT"}

    monkeypatch.setattr(sr, "_yahoo_screener_session",
                        lambda: (object(), "crumb"))
    monkeypatch.setattr(sr, "fetch_yahoo_screener", slow_screener)
    monkeypatch.setattr(sr, "fetch_dataroma_tickers", slow_dataroma)
    monkeypatch.setattr(sr, "fetch_dataroma_activist_aggregate", slow_activists)
    out = sr.fetch_all_watchlists(max_workers=6)
    assert in_flight["max"] >= 2, "watchlists must overlap"
    assert out  # something landed


def test_get_yahoo_session_returns_session_even_on_warmup_failure(monkeypatch):
    """Network failure during cookie seeding must not crash the run — we
    return a usable session anyway. yfinance will retry the auth itself."""
    import stocks_report as sr

    class _BoomSession:
        def __init__(self): self.headers = {}
        def get(self, *a, **kw): raise RuntimeError("network down")

    monkeypatch.setattr(sr.requests, "Session", _BoomSession)
    s = sr.get_yahoo_session()
    assert s is not None


if __name__ == "__main__":
    sys.exit(pytest.main([__file__, "-v"]))
